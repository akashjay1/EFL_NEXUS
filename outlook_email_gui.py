"""
Outlook Email Sender (GUI) - Google Sheets Cloud Log Version
------------------------------------------------------------
Sends email through the user's real, already-installed classic Outlook
desktop app via COM automation (pywin32) - not a separate SMTP relay - so the
message goes out using the user's actual account, signature, and Sent Items
folder.

Signature handling:
- A plain Tkinter text box cannot hold rich text (fonts, colors, logos), so
  pasting a signature into it would always strip its formatting.
- Instead of pasting anything, this app reads the user's real default
  signature straight off disk (the .htm file Outlook itself maintains under
  %APPDATA%\\Microsoft\\Signatures) and stitches the typed message in above
  it, all as one HTML email body.

Attachments:
- Multiple files can be attached at once (PDF and/or Excel), shown in a
  list, with the ability to remove individual files before sending.

Recipient group templates:
- Each template stores a group name plus its own 'To' and 'CC' recipient
  lists, saved in templates.xlsx (Excel) next to this script.
- A dropdown lets the user pick a saved template, which fills in To/CC.
- "Add New Template" opens a small form to save a new group.

GDN / GRN type selection:
- Two toggle buttons let the user explicitly mark the email being composed
  as a "GDN" or a "GRN" (or neither). Pressing one of these buttons also
  renames the Subject field to "GDN -" / "GRN -"; the previously typed subject
  is remembered and restored if the type is cleared or switched back off.

Send tracking & Activity Log:
- Sending is executed on an asynchronous background thread with dedicated
  COM runtime initialization (pythoncom.CoInitialize) so the UI never freezes.
- A real-time Activity Log displays comprehensive dispatch records including
  Gate Pass Numbers, Warehouse Names, Subjects, Recipients, Attachments,
  and live Google Sheet cloud synchronization status.
- Whenever an email is sent with GDN or GRN selected, log data is sent
  silently over HTTP to a Google Apps Script Web App attached to an online
  Google Sheet.
"""

import os
import re
import sys
import html
import json
import datetime
import threading
import requests
import webbrowser
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

try:
    import win32com.client
    import pythoncom
    WIN32COM_AVAILABLE = True
except ImportError:
    WIN32COM_AVAILABLE = False

try:
    import winreg
    WINREG_AVAILABLE = True
except ImportError:
    WINREG_AVAILABLE = False

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# --------------------------------------------------------------------------
# Shared visual style
# --------------------------------------------------------------------------

BG_COLOR = "#faf8f2"
PANEL_COLOR = "#ffffff"
ACCENT_COLOR = "#0b1420"
ACCENT_COLOR_DARK = "#162e4c"
AURORA_CYAN = "#00e5ff"
AURORA_MINT = "#49cf9e"
TEXT_COLOR = "#0f172a"
MUTED_COLOR = "#64748b"
DANGER_COLOR = "#ef4444"
FONT_FAMILY = "Segoe UI"
FONT_NORMAL = (FONT_FAMILY, 10)
FONT_BOLD = (FONT_FAMILY, 10, "bold")
FONT_HEADER = (FONT_FAMILY, 14, "bold")

OUTLOOK_MAIL_ITEM = 0  # olMailItem


def _get_base_dir():
    """Where templates.xlsx and config.json should live."""
    if getattr(sys, "frozen", False):
        return os.path.dirname(os.path.abspath(sys.executable))
    return os.path.dirname(os.path.abspath(__file__))


BASE_DIR = _get_base_dir()
TEMPLATES_XLSX_PATH = os.path.join(BASE_DIR, "templates.xlsx")
SENT_LOG_XLSX_PATH = os.path.join(BASE_DIR, "sent_log.xlsx")
CONFIG_JSON_PATH = os.path.join(BASE_DIR, "config.json")

TEMPLATE_PLACEHOLDER = "-- Select Template --"
DEFAULT_BODY_TEXT = "Dear all,\n\nFYI."

SEND_TYPE_GDN = "GDN"
SEND_TYPE_GRN = "GRN"

ATTACHMENT_FILETYPES = [
    ("PDF and Excel files", "*.pdf;*.xlsx;*.xls"),
    ("PDF files", "*.pdf"),
    ("Excel files", "*.xlsx;*.xls"),
    ("All files", "*.*"),
]

SIGNATURE_REGISTRY_PATHS = [
    r"Software\Microsoft\Office\16.0\Common\MailSettings",
    r"Software\Microsoft\Office\15.0\Common\MailSettings",
]


def style_button(btn, kind="primary"):
    colors = {
        "primary": (ACCENT_COLOR, "#ffffff", ACCENT_COLOR_DARK),
        "danger": (DANGER_COLOR, "#ffffff", "#b91c1c"),
        "secondary": ("#e2e8f0", TEXT_COLOR, "#cbd5e1"),
        "success": ("#16a34a", "#ffffff", "#15803d"),
    }
    bg, fg, active_bg = colors.get(kind, colors["primary"])
    btn.configure(
        bg=bg, fg=fg, activebackground=active_bg, activeforeground=fg,
        font=FONT_BOLD, relief="flat", bd=0, padx=14, pady=6, cursor="hand2",
    )


# --------------------------------------------------------------------------
# Config Store (JSON-based settings)
# --------------------------------------------------------------------------

class ConfigStore:
    """Handles saving and loading user-defined settings (Google Sheets, Web App, Korber Credentials)."""

    def __init__(self, path):
        self.path = path
        self.config = self.load()

    def load(self):
        if os.path.exists(self.path):
            try:
                with open(self.path, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception:
                pass
        return {
            "webapp_url": "",
            "sheet_url": "",
            "korber_user": "",
            "korber_pass": "",
            "korber_url": "https://lopwaprodweb.koerbercloud.com/core/Default.html"
        }

    def save(self, webapp_url=None, sheet_url=None, korber_user=None, korber_pass=None, korber_url=None, **kwargs):
        if webapp_url is not None:
            self.config["webapp_url"] = str(webapp_url).strip()
        if sheet_url is not None:
            self.config["sheet_url"] = str(sheet_url).strip()
        if korber_user is not None:
            self.config["korber_user"] = str(korber_user).strip()
        if korber_pass is not None:
            self.config["korber_pass"] = str(korber_pass).strip()
        if korber_url is not None:
            self.config["korber_url"] = str(korber_url).strip()
        for k, v in kwargs.items():
            self.config[k] = str(v).strip() if v is not None else ""
        try:
            with open(self.path, "w", encoding="utf-8") as f:
                json.dump(self.config, f, indent=4)
        except Exception as e:
            print(f"Failed to save config: {e}")

    def get_webapp_url(self):
        return self.config.get("webapp_url", "")

    def get_sheet_url(self):
        return self.config.get("sheet_url", "")

    def get_korber_user(self):
        return self.config.get("korber_user", "") or os.environ.get("KORBER_USER", "")

    def get_korber_pass(self):
        return self.config.get("korber_pass", "") or os.environ.get("KORBER_PASS", "")

    def get_korber_url(self):
        return self.config.get("korber_url", "") or os.environ.get("KORBER_URL", "https://lopwaprodweb.koerbercloud.com/core/Default.html")


# --------------------------------------------------------------------------
# Recipient helpers
# --------------------------------------------------------------------------

def normalize_recipients(raw_text):
    """Outlook expects recipients separated by semicolons."""
    if not raw_text:
        return ""
    text = raw_text.replace(",", ";")
    parts = [p.strip() for p in text.split(";")]
    parts = [p for p in parts if p]
    return "; ".join(parts)


# --------------------------------------------------------------------------
# Default signature (read directly from disk)
# --------------------------------------------------------------------------

def _fix_signature_image_paths(html_text, files_dir):
    if not files_dir or not os.path.isdir(files_dir):
        return html_text

    def repl(match):
        attr, quote, src = match.group(1), match.group(2), match.group(3)
        if src.lower().startswith(("http://", "https://", "file://", "cid:", "data:")):
            return match.group(0)
        abs_path = os.path.join(files_dir, src)
        if os.path.isfile(abs_path):
            uri = "file:///" + abs_path.replace(os.sep, "/")
            return f'{attr}={quote}{uri}{quote}'
        return match.group(0)

    return re.sub(r'(src)=(["\'])([^"\']+)\2', repl, html_text, flags=re.IGNORECASE)


def get_default_signature_html():
    appdata = os.environ.get("APPDATA")
    if not appdata:
        return ""
    sig_dir = os.path.join(appdata, "Microsoft", "Signatures")
    if not os.path.isdir(sig_dir):
        return ""

    sig_name = None
    if WINREG_AVAILABLE:
        for path in SIGNATURE_REGISTRY_PATHS:
            try:
                with winreg.OpenKey(winreg.HKEY_CURRENT_USER, path) as key:
                    value, _ = winreg.QueryValueEx(key, "NewSignature")
                    if value:
                        sig_name = value
                        break
            except OSError:
                continue

    if not sig_name:
        try:
            htm_files = [f for f in os.listdir(sig_dir) if f.lower().endswith(".htm")]
        except OSError:
            return ""
        if not htm_files:
            return ""
        htm_files.sort(
            key=lambda f: os.path.getmtime(os.path.join(sig_dir, f)), reverse=True)
        sig_name = os.path.splitext(htm_files[0])[0]

    sig_path = os.path.join(sig_dir, f"{sig_name}.htm")
    if not os.path.isfile(sig_path):
        return ""

    try:
        with open(sig_path, "r", encoding="utf-8", errors="ignore") as f:
            signature_html = f.read()
    except OSError:
        return ""

    files_dir = os.path.join(sig_dir, f"{sig_name}_files")
    return _fix_signature_image_paths(signature_html, files_dir)


# --------------------------------------------------------------------------
# Recipient-group template storage (Excel)
# --------------------------------------------------------------------------

class TemplateStore:
    def __init__(self, path):
        self.path = path
        self.templates = self.load()

    def load(self):
        if not OPENPYXL_AVAILABLE or not os.path.exists(self.path):
            return []
        try:
            wb = load_workbook(self.path)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return []
            header = [str(c).strip().lower() if c else "" for c in rows[0]]
            try:
                name_idx = header.index("name")
                to_idx = header.index("to")
                cc_idx = header.index("cc")
            except ValueError:
                name_idx, to_idx, cc_idx = 0, 1, 2

            results = []
            for r in rows[1:]:
                if not r or not any(r):
                    continue
                name = str(r[name_idx]).strip() if len(r) > name_idx and r[name_idx] else ""
                to_addr = str(r[to_idx]).strip() if len(r) > to_idx and r[to_idx] else ""
                cc_addr = str(r[cc_idx]).strip() if len(r) > cc_idx and r[cc_idx] else ""
                if name:
                    results.append({
                        "name": name,
                        "to": normalize_recipients(to_addr),
                        "cc": normalize_recipients(cc_addr),
                    })
            return sorted(results, key=lambda t: t["name"].lower())
        except Exception as e:
            print(f"Error loading templates: {e}")
            return []

    def save(self):
        if not OPENPYXL_AVAILABLE:
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "Templates"
        ws.append(["Name", "To", "CC"])
        for t in self.templates:
            ws.append([t["name"], t["to"], t["cc"]])
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 40
        wb.save(self.path)

    def add_or_update_template(self, name, to_addr, cc_addr):
        clean_name = name.strip()
        clean_to = normalize_recipients(to_addr)
        clean_cc = normalize_recipients(cc_addr)
        for t in self.templates:
            if t["name"].lower() == clean_name.lower():
                t["name"] = clean_name
                t["to"] = clean_to
                t["cc"] = clean_cc
                self.save()
                return
        self.templates.append({"name": clean_name, "to": clean_to, "cc": clean_cc})
        self.templates.sort(key=lambda t: t["name"].lower())
        self.save()

    def get_names(self):
        return [t["name"] for t in self.templates]

    def get_template(self, name):
        for t in self.templates:
            if t["name"].lower() == name.strip().lower():
                return t
        return None


# --------------------------------------------------------------------------
# Send-tracking store (Excel + Google Apps Script Web App)
# --------------------------------------------------------------------------

class SentLogStore:
    def __init__(self, config_store, excel_path=SENT_LOG_XLSX_PATH):
        self.config_store = config_store
        self.excel_path = excel_path

    @staticmethod
    def _normalize_date_str(val):
        """Normalize date value/string to YYYY-MM-DD."""
        if val is None:
            return ""
        s = str(val).strip()
        if not s:
            return ""
        import re
        m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", s)
        if m:
            month, day, year = int(m.group(1)), int(m.group(2)), m.group(3)
            return f"{year}-{month:02d}-{day:02d}"
        return s

    @staticmethod
    def _normalize_time_str(val):
        """Normalize time value/string to HH:MM:SS."""
        if val is None:
            return ""
        s = str(val).strip()
        if not s:
            return ""
        import re
        m = re.match(r"^(\d{1,2}):(\d{1,2})(?::(\d{1,2}))?$", s)
        if m:
            hour = int(m.group(1))
            minute = int(m.group(2))
            second = int(m.group(3)) if m.group(3) is not None else 0
            return f"{hour:02d}:{minute:02d}:{second:02d}"
        return s

    @staticmethod
    def _apply_sheet_formatting(ws):
        """Apply Calibri 11, proper alignments, and number formats matching the screenshot."""
        font_header = Font(name="Calibri", size=11, bold=True)
        font_data = Font(name="Calibri", size=11)
        align_right = Alignment(horizontal="right", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")

        # Header row (Row 1)
        for col_idx in range(1, 9):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = font_header
            if col_idx in (4, 8):
                cell.alignment = align_left
            else:
                cell.alignment = align_right

        # Data rows (Row 2+)
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, 9):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = font_data
                if col_idx in (1, 5):  # GDN Date, GRN Date
                    cell.number_format = 'yyyy-mm-dd'
                    cell.alignment = align_right
                elif col_idx in (2, 6):  # GDN Time, GRN Time
                    cell.number_format = 'hh:mm:ss'
                    cell.alignment = align_right
                elif col_idx in (3, 7):  # GDN Number, GRN Number
                    cell.number_format = '@'
                    cell.alignment = align_right
                elif col_idx in (4, 8):  # GDN Type, GRN Type
                    cell.number_format = '@'
                    cell.alignment = align_left

        for col, width in zip("ABCDEFGH", (14, 12, 16, 20, 14, 12, 16, 20)):
            ws.column_dimensions[col].width = width

    def sort_local_records(self):
        if not OPENPYXL_AVAILABLE or not os.path.exists(self.excel_path):
            return
        try:
            wb = load_workbook(self.excel_path)
            ws = wb.active

            gdn_rows = []
            grn_rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                if any(row[:4]):
                    d = self._normalize_date_str(row[0])
                    t = self._normalize_time_str(row[1])
                    num = str(row[2] if row[2] is not None else "").strip()
                    tp = str(row[3] if row[3] is not None else "").strip()
                    if d or t or num or tp:
                        gdn_rows.append([d, t, num, tp])
                if len(row) > 4 and any(row[4:8]):
                    d = self._normalize_date_str(row[4])
                    t = self._normalize_time_str(row[5])
                    num = str(row[6] if row[6] is not None else "").strip()
                    tp = str(row[7] if row[7] is not None else "").strip()
                    if d or t or num or tp:
                        grn_rows.append([d, t, num, tp])

            gdn_rows.sort(key=lambda r: (r[0], r[1]))
            grn_rows.sort(key=lambda r: (r[0], r[1]))

            max_r = ws.max_row
            if max_r > 1:
                ws.delete_rows(2, max_r)

            total_rows = max(len(gdn_rows), len(grn_rows))
            for i in range(total_rows):
                gdn_part = gdn_rows[i] if i < len(gdn_rows) else ["", "", "", ""]
                grn_part = grn_rows[i] if i < len(grn_rows) else ["", "", "", ""]
                ws.append(gdn_part + grn_part)

            self._apply_sheet_formatting(ws)
            wb.save(self.excel_path)
        except Exception as e:
            print(f"Error sorting local sent_log.xlsx: {e}")

    @staticmethod
    def _split_subject(send_type, subject_text):
        text = (subject_text or "").strip()
        marker = f"{send_type} -"
        idx = text.find(marker)
        if idx == -1:
            return text, send_type
        prefix_end = idx + len(marker)
        if prefix_end < len(text) and text[prefix_end] == " ":
            prefix_end += 1
        number_part = text[prefix_end:].strip()
        type_part = text[:prefix_end].rstrip()
        if type_part.endswith("-"):
            type_part = type_part[:-1].rstrip()
        return number_part, type_part

    def _log_to_excel(self, send_type, date_str, time_str, number_part, type_part):
        if not OPENPYXL_AVAILABLE:
            return
        try:
            if os.path.exists(self.excel_path):
                wb = load_workbook(self.excel_path)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Sent Log"
                ws.append([
                    "GDN Date", "GDN Time", "GDN Number", "GDN Type",
                    "GRN Date", "GRN Time", "GRN Number", "GRN Type"
                ])

            gdn_rows = []
            grn_rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                if any(row[:4]):
                    d = self._normalize_date_str(row[0])
                    t = self._normalize_time_str(row[1])
                    num = str(row[2] if row[2] is not None else "").strip()
                    tp = str(row[3] if row[3] is not None else "").strip()
                    if d or t or num or tp:
                        gdn_rows.append([d, t, num, tp])
                if len(row) > 4 and any(row[4:8]):
                    d = self._normalize_date_str(row[4])
                    t = self._normalize_time_str(row[5])
                    num = str(row[6] if row[6] is not None else "").strip()
                    tp = str(row[7] if row[7] is not None else "").strip()
                    if d or t or num or tp:
                        grn_rows.append([d, t, num, tp])

            new_record = [
                self._normalize_date_str(date_str),
                self._normalize_time_str(time_str),
                str(number_part or "").strip(),
                str(type_part or "").strip()
            ]
            if send_type == SEND_TYPE_GDN:
                gdn_rows.append(new_record)
                gdn_rows.sort(key=lambda r: (r[0], r[1]))
            elif send_type == SEND_TYPE_GRN:
                grn_rows.append(new_record)
                grn_rows.sort(key=lambda r: (r[0], r[1]))

            max_r = ws.max_row
            if max_r > 1:
                ws.delete_rows(2, max_r)

            total_rows = max(len(gdn_rows), len(grn_rows))
            for i in range(total_rows):
                gdn_part = gdn_rows[i] if i < len(gdn_rows) else ["", "", "", ""]
                grn_part = grn_rows[i] if i < len(grn_rows) else ["", "", "", ""]
                ws.append(gdn_part + grn_part)

            self._apply_sheet_formatting(ws)
            wb.save(self.excel_path)
        except Exception as e:
            print(f"Excel logging failed: {e}")

    def log_send(self, send_type, subject_text):
        if send_type not in [SEND_TYPE_GDN, SEND_TYPE_GRN]:
            return

        now = datetime.datetime.now()
        date_str = now.strftime("%Y-%m-%d")
        time_str = now.strftime("%H:%M:%S")
        number_part, type_part = self._split_subject(send_type, subject_text)

        # 1. Update local sent_log.xlsx (in order of date)
        self._log_to_excel(send_type, date_str, time_str, number_part, type_part)

        # 2. Update Google Sheet Web App
        url = self.config_store.get_webapp_url()
        if not url:
            return

        try:
            payload = {
                "type": send_type,
                "date": date_str,
                "time": time_str,
                "number": number_part,
                "type_part": type_part
            }
            requests.post(url, json=payload, timeout=10)
        except Exception as e:
            print(f"Logging failed: {e}")

    def get_counts(self):
        url = self.config_store.get_webapp_url()
        if not url:
            return {"GDN": 0, "GRN": 0}

        try:
            response = requests.get(url, timeout=5)
            if response.status_code == 200:
                return response.json()
        except Exception:
            pass
        return {"GDN": 0, "GRN": 0}

    def get_recent_records(self, limit=10):
        """Read latest records from local sent_log.xlsx for activity log history."""
        if not OPENPYXL_AVAILABLE or not os.path.exists(self.excel_path):
            return []
        try:
            wb = load_workbook(self.excel_path, data_only=True)
            ws = wb.active
            records = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                # GDN record
                if any(row[:4]) and str(row[0] or "").strip():
                    records.append({
                        "type": "GDN",
                        "date": self._normalize_date_str(row[0]),
                        "time": self._normalize_time_str(row[1]),
                        "number": str(row[2] or "").strip(),
                        "warehouse": str(row[3] or "").replace("GDN", "").strip().rstrip("-").strip() or "General"
                    })
                # GRN record
                if len(row) > 4 and any(row[4:8]) and str(row[4] or "").strip():
                    records.append({
                        "type": "GRN",
                        "date": self._normalize_date_str(row[4]),
                        "time": self._normalize_time_str(row[5]),
                        "number": str(row[6] or "").strip(),
                        "warehouse": str(row[7] or "").replace("GRN", "").strip().rstrip("-").strip() or "General"
                    })
            records.sort(key=lambda r: (r.get("date", ""), r.get("time", "")), reverse=True)
            return records[:limit]
        except Exception as e:
            print(f"Error reading recent records: {e}")
            return []


# --------------------------------------------------------------------------
# Outlook automation
# --------------------------------------------------------------------------

class OutlookClient:
    def __init__(self):
        self._app = None

    def _get_app(self):
        if self._app is None:
            if not WIN32COM_AVAILABLE:
                raise RuntimeError(
                    "pywin32 is not installed. Install it with:\n\n"
                    "pip install pywin32")
            self._app = win32com.client.Dispatch("Outlook.Application")
        return self._app

    def build_mail(self, to_addr, cc_addr, subject, body_text,
                    attachment_paths, include_signature=True):
        app = self._get_app()
        mail = app.CreateItem(OUTLOOK_MAIL_ITEM)

        try:
            mail.To = to_addr
            if cc_addr:
                mail.CC = cc_addr
            mail.Subject = subject
        except Exception as exc:
            raise RuntimeError(f"Could not set recipients/subject: {exc}")

        signature_html = ""
        if include_signature:
            try:
                signature_html = get_default_signature_html()
            except Exception:
                signature_html = ""

        try:
            mail.HTMLBody = self._compose_html(body_text, signature_html)
        except Exception as exc:
            raise RuntimeError(f"Could not set the email body: {exc}")

        for path in attachment_paths:
            if not os.path.isfile(path):
                raise RuntimeError(f"Attachment file not found on disk: {path}")
            try:
                mail.Attachments.Add(path)
            except Exception as exc:
                raise RuntimeError(
                    f"Could not attach '{os.path.basename(path)}': {exc}")

        return mail

    @staticmethod
    def _compose_html(body_text, signature_html):
        message_html = (
            '<div style="font-family:Calibri,Arial,sans-serif;'
            'font-size:11pt;color:#000000;">'
            + html.escape(body_text).replace("\n", "<br>")
            + "</div><br>"
        )
        if signature_html:
            match = re.search(r"(<body[^>]*>)", signature_html, re.IGNORECASE)
            if match:
                insertion_point = match.end()
                return (signature_html[:insertion_point] + message_html
                         + signature_html[insertion_point:])
            return message_html + signature_html
        return f"<html><body>{message_html}</body></html>"

    def preview(self, mail):
        mail.Display(True)

    def send(self, mail):
        mail.Send()


# --------------------------------------------------------------------------
# Settings Dialog
# --------------------------------------------------------------------------

class SettingsDialog(tk.Toplevel):
    def __init__(self, master, config_store, on_save):
        super().__init__(master)
        self.config_store = config_store
        self.on_save = on_save
        self.title("Settings")
        self.configure(bg=BG_COLOR)
        self.resizable(False, False)
        self.transient(master)
        self.grab_set()
        self._build_ui()
        self._center_window(master)

    def _center_window(self, master):
        self.update_idletasks()
        master.update_idletasks()
        mw = master.winfo_width() or 800
        mh = master.winfo_height() or 600
        mx = master.winfo_rootx()
        my = master.winfo_rooty()
        dw = self.winfo_reqwidth()
        dh = self.winfo_reqheight()
        x = max(0, mx + (mw - dw) // 2)
        y = max(0, my + (mh - dh) // 2)
        self.geometry(f"+{x}+{y}")

    def _build_ui(self):
        tk.Label(self, text="Google Sheets Settings", font=FONT_HEADER, bg=BG_COLOR,
                  fg=TEXT_COLOR).grid(row=0, column=0, columnspan=2,
                                       padx=20, pady=(16, 10), sticky="w")

        tk.Label(self, text="Web App URL:", font=FONT_NORMAL, bg=BG_COLOR,
                  fg=TEXT_COLOR).grid(row=1, column=0, sticky="w", padx=20, pady=6)
        self.webapp_entry = tk.Entry(self, font=FONT_NORMAL, width=50)
        self.webapp_entry.insert(0, self.config_store.get_webapp_url())
        self.webapp_entry.grid(row=1, column=1, padx=20, pady=6)

        tk.Label(self, text="Google Sheet URL:", font=FONT_NORMAL, bg=BG_COLOR,
                  fg=TEXT_COLOR).grid(row=2, column=0, sticky="w", padx=20, pady=6)
        self.sheet_entry = tk.Entry(self, font=FONT_NORMAL, width=50)
        self.sheet_entry.insert(0, self.config_store.get_sheet_url())
        self.sheet_entry.grid(row=2, column=1, padx=20, pady=6)

        btn_frame = tk.Frame(self, bg=BG_COLOR)
        btn_frame.grid(row=3, column=0, columnspan=2, pady=16)

        save_btn = tk.Button(btn_frame, text="Save Settings", command=self._save)
        style_button(save_btn, "primary")
        save_btn.pack(side="left", padx=6)

        cancel_btn = tk.Button(btn_frame, text="Cancel", command=self.destroy)
        style_button(cancel_btn, "secondary")
        cancel_btn.pack(side="left", padx=6)

    def _save(self):
        webapp_url = self.webapp_entry.get().strip()
        sheet_url = self.sheet_entry.get().strip()

        self.config_store.save(webapp_url, sheet_url)
        self.on_save()
        messagebox.showinfo("Saved", "Settings saved successfully.")
        self.destroy()


# --------------------------------------------------------------------------
# Add New Template dialog
# --------------------------------------------------------------------------

class AddTemplateDialog(tk.Toplevel):
    def __init__(self, master, on_submit):
        super().__init__(master)
        self.on_submit = on_submit
        self.title("Add New Template")
        self.configure(bg=BG_COLOR)
        self.resizable(False, False)
        self.transient(master)
        self.grab_set()
        self._build_ui()
        self._center_window(master)

    def _center_window(self, master):
        self.update_idletasks()
        master.update_idletasks()
        mw = master.winfo_width() or 800
        mh = master.winfo_height() or 600
        mx = master.winfo_rootx()
        my = master.winfo_rooty()
        dw = self.winfo_reqwidth()
        dh = self.winfo_reqheight()
        x = max(0, mx + (mw - dw) // 2)
        y = max(0, my + (mh - dh) // 2)
        self.geometry(f"+{x}+{y}")

    def _build_ui(self):
        tk.Label(self, text="Add New Template", font=FONT_HEADER, bg=BG_COLOR,
                  fg=TEXT_COLOR).grid(row=0, column=0, columnspan=2,
                                       padx=20, pady=(16, 10), sticky="w")

        tk.Label(self, text="Template name:", font=FONT_NORMAL, bg=BG_COLOR,
                  fg=TEXT_COLOR).grid(row=1, column=0, sticky="w", padx=20, pady=6)
        self.name_entry = tk.Entry(self, font=FONT_NORMAL, width=40)
        self.name_entry.grid(row=1, column=1, padx=20, pady=6)

        tk.Label(self, text="To:", font=FONT_NORMAL, bg=BG_COLOR,
                  fg=TEXT_COLOR).grid(row=2, column=0, sticky="w", padx=20, pady=6)
        self.to_entry = tk.Entry(self, font=FONT_NORMAL, width=40)
        self.to_entry.grid(row=2, column=1, padx=20, pady=6)

        tk.Label(self, text="CC:", font=FONT_NORMAL, bg=BG_COLOR,
                  fg=TEXT_COLOR).grid(row=3, column=0, sticky="w", padx=20, pady=6)
        self.cc_entry = tk.Entry(self, font=FONT_NORMAL, width=40)
        self.cc_entry.grid(row=3, column=1, padx=20, pady=6)

        tk.Label(self, text="(separate multiple recipients with , or ;)",
                  font=(FONT_FAMILY, 8), bg=BG_COLOR, fg=MUTED_COLOR).grid(
            row=4, column=1, sticky="w", padx=20)

        btn_frame = tk.Frame(self, bg=BG_COLOR)
        btn_frame.grid(row=5, column=0, columnspan=2, pady=16)

        save_btn = tk.Button(btn_frame, text="Save Template", command=self._save)
        style_button(save_btn, "primary")
        save_btn.pack(side="left", padx=6)

        cancel_btn = tk.Button(btn_frame, text="Cancel", command=self.destroy)
        style_button(cancel_btn, "secondary")
        cancel_btn.pack(side="left", padx=6)

    def _save(self):
        name = self.name_entry.get().strip()
        to_addr = normalize_recipients(self.to_entry.get().strip())
        cc_addr = normalize_recipients(self.cc_entry.get().strip())

        if not name:
            messagebox.showwarning("Missing name", "Please enter a template name.")
            return
        if not to_addr:
            messagebox.showwarning("Missing recipients",
                                    "Please enter at least one 'To:' recipient.")
            return

        self.on_submit(name, to_addr, cc_addr)
        self.destroy()


# --------------------------------------------------------------------------
# Main application window
# --------------------------------------------------------------------------

class OutlookEmailApp:
    def __init__(self, root, container=None, standalone=True, on_open_settings=None):
        self.root = root
        self.on_open_settings = on_open_settings
        self.target = container if container is not None else root
        self.standalone = standalone

        if standalone:
            root.title("Outlook Email Sender")
            root.geometry("1180x820")
            root.configure(bg=BG_COLOR)
            root.minsize(950, 680)
            try:
                root.state('zoomed')
            except Exception:
                pass
            for icon_name in ("icon_2.ico", "icon.ico", "favicon.ico"):
                icon_path = os.path.join(BASE_DIR, icon_name)
                if not os.path.exists(icon_path) and getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
                    icon_path = os.path.join(sys._MEIPASS, icon_name)
                if os.path.exists(icon_path):
                    try:
                        root.iconbitmap(default=icon_path)
                        break
                    except Exception:
                        try:
                            root.iconbitmap(icon_path)
                            break
                        except Exception:
                            pass

        self.config_store = ConfigStore(CONFIG_JSON_PATH)
        self.outlook = OutlookClient()
        self.template_store = TemplateStore(TEMPLATES_XLSX_PATH)
        self.sent_log = SentLogStore(self.config_store)
        self.attachment_paths = []

        self._subject_before_type = None
        self._last_prefix = None
        self._is_sending = False
        self.session_dispatch_count = 0

        self._build_ui(self.target)
        self._refresh_counts_label()
        self._populate_recent_activity_logs()

        if not WIN32COM_AVAILABLE:
            messagebox.showwarning(
                "pywin32 not found",
                "pywin32 is not installed, so this app cannot talk to "
                "Outlook yet.\n\nInstall it with:\n\npip install pywin32")
        if not OPENPYXL_AVAILABLE:
            messagebox.showwarning(
                "openpyxl not found",
                "openpyxl is not installed, so recipient templates "
                "can't be saved/loaded yet.\n\n"
                "Install it with:\n\npip install openpyxl")

    # -- UI construction ---------------------------------------------------

    def _build_ui(self, target):
        # --- Header banner ---
        HEADER_BG = ACCENT_COLOR
        HEADER_BTN_BG = "#162e4c"
        HEADER_BTN_BG_HOVER = "#1f3f66"

        header = tk.Frame(target, bg=HEADER_BG)
        header.pack(fill="x", side="top")

        header_top_row = tk.Frame(header, bg=HEADER_BG)
        header_top_row.pack(fill="x", padx=20, pady=(16, 0))

        tk.Label(
            header_top_row,
            text="Outlook Email Sender",
            font=(FONT_FAMILY, 15, "bold"),
            bg=HEADER_BG,
            fg="#ffffff"
        ).pack(side="left")

        def _make_header_button(parent, text, command):
            btn = tk.Label(
                parent, text=text, bg=HEADER_BTN_BG, fg="#ffffff",
                font=(FONT_FAMILY, 8, "bold"), cursor="hand2",
                padx=8, pady=4,
            )
            btn.bind("<Button-1>", lambda e: command())
            btn.bind("<Enter>", lambda e: btn.config(bg=HEADER_BTN_BG_HOVER))
            btn.bind("<Leave>", lambda e: btn.config(bg=HEADER_BTN_BG))
            return btn

        # Header action buttons (top right)
        _make_header_button(header_top_row, "⚙ Settings", self._open_settings_dialog).pack(
            side="right"
        )
        _make_header_button(header_top_row, "View Records", self._open_records_window).pack(
            side="right", padx=(0, 6)
        )

        self.counts_var = tk.StringVar(value="GDN sent: 0    GRN sent: 0")
        tk.Label(
            header_top_row,
            textvariable=self.counts_var,
            font=(FONT_FAMILY, 9),
            bg=HEADER_BG,
            fg="#94a3b8"
        ).pack(side="right", padx=(20, 10))

        tk.Label(
            header,
            text="GDN / GRN direct dispatch engine with live activity logging",
            font=(FONT_FAMILY, 9),
            bg=HEADER_BG,
            fg="#94a3b8"
        ).pack(anchor="w", padx=20, pady=(0, 14))

        # --- Main Split Container (Left: Composer, Right: Activity Log) ---
        main_split = tk.Frame(target, bg=BG_COLOR)
        main_split.pack(fill="both", expand=True, padx=18, pady=14)

        # -------------------------------------------------------------
        # LEFT PANEL: Email Composer
        # -------------------------------------------------------------
        composer_outer = tk.Frame(main_split, bg=BG_COLOR)
        composer_outer.pack(side="left", fill="both", expand=True, padx=(0, 12))

        body = tk.Frame(composer_outer, bg=BG_COLOR)
        body.pack(fill="both", expand=True)
        body.grid_columnconfigure(1, weight=1)

        # 1. Template picker
        tk.Label(body, text="Template:", font=FONT_NORMAL, bg=BG_COLOR,
                  fg=TEXT_COLOR).grid(row=0, column=0, sticky="w", pady=4)
        template_frame = tk.Frame(body, bg=BG_COLOR)
        template_frame.grid(row=0, column=1, sticky="ew", padx=(10, 0), pady=4)
        template_frame.grid_columnconfigure(0, weight=1)

        self.template_var = tk.StringVar(value=TEMPLATE_PLACEHOLDER)
        self.template_combo = ttk.Combobox(
            template_frame, textvariable=self.template_var,
            state="readonly")
        self.template_combo.grid(row=0, column=0, sticky="ew")
        self.template_combo.bind("<<ComboboxSelected>>", self._on_template_selected)

        add_template_btn = tk.Button(template_frame, text="Add New Template",
                                      command=self._open_add_template_dialog)
        style_button(add_template_btn, "secondary")
        add_template_btn.grid(row=0, column=1, padx=(10, 0))

        self._refresh_template_dropdown()

        # 2. To
        tk.Label(body, text="To:", font=FONT_NORMAL, bg=BG_COLOR,
                  fg=TEXT_COLOR).grid(row=1, column=0, sticky="w", pady=4)
        self.to_entry = tk.Entry(body, font=FONT_NORMAL)
        self.to_entry.grid(row=1, column=1, sticky="ew", padx=(10, 0), pady=4)

        # 3. CC
        tk.Label(body, text="CC:", font=FONT_NORMAL, bg=BG_COLOR,
                  fg=TEXT_COLOR).grid(row=2, column=0, sticky="w", pady=4)
        self.cc_entry = tk.Entry(body, font=FONT_NORMAL)
        self.cc_entry.grid(row=2, column=1, sticky="ew", padx=(10, 0), pady=4)

        tk.Label(body, text="(multiple recipients: separate with , or ;)",
                  font=(FONT_FAMILY, 8), bg=BG_COLOR, fg=MUTED_COLOR).grid(
            row=3, column=1, sticky="w", padx=(10, 0))

        # 4. Subject
        tk.Label(body, text="Subject:", font=FONT_NORMAL, bg=BG_COLOR,
                  fg=TEXT_COLOR).grid(row=4, column=0, sticky="w", pady=(8, 4))
        self.subject_entry = tk.Entry(body, font=FONT_NORMAL)
        self.subject_entry.grid(row=4, column=1, sticky="ew", padx=(10, 0), pady=(8, 4))

        # 5. GDN / GRN type selection
        tk.Label(body, text="Type:", font=FONT_NORMAL, bg=BG_COLOR,
                  fg=TEXT_COLOR).grid(row=5, column=0, sticky="w", pady=(0, 4))
        type_frame = tk.Frame(body, bg=BG_COLOR)
        type_frame.grid(row=5, column=1, sticky="w", padx=(10, 0), pady=(0, 4))

        self.selected_type_var = tk.StringVar(value="")

        self.gdn_btn = tk.Button(type_frame, text="GDN",
                                  command=lambda: self._select_type(SEND_TYPE_GDN))
        style_button(self.gdn_btn, "secondary")
        self.gdn_btn.pack(side="left")

        self.grn_btn = tk.Button(type_frame, text="GRN",
                                  command=lambda: self._select_type(SEND_TYPE_GRN))
        style_button(self.grn_btn, "secondary")
        self.grn_btn.pack(side="left", padx=(6, 0))

        clear_type_btn = tk.Button(type_frame, text="Clear",
                                    command=lambda: self._select_type(None))
        style_button(clear_type_btn, "secondary")
        clear_type_btn.pack(side="left", padx=(6, 0))

        tk.Label(type_frame,
                  text="(sets 'GDN -' / 'GRN -' prefix & logs gate pass)",
                  font=(FONT_FAMILY, 8), bg=BG_COLOR, fg=MUTED_COLOR).pack(
            side="left", padx=(10, 0))

        # 6. Message
        tk.Label(body, text="Message:", font=FONT_NORMAL, bg=BG_COLOR,
                  fg=TEXT_COLOR).grid(row=6, column=0, sticky="nw", pady=(8, 4))
        msg_frame = tk.Frame(body, bg=BG_COLOR)
        msg_frame.grid(row=6, column=1, sticky="nsew", padx=(10, 0), pady=(8, 4))
        body.grid_rowconfigure(6, weight=1)

        self.message_text = tk.Text(msg_frame, font=FONT_NORMAL, wrap="word",
                                     height=8, relief="solid", bd=1)
        self.message_text.insert("1.0", DEFAULT_BODY_TEXT)
        self.message_text.pack(side="left", fill="both", expand=True)
        msg_scroll = tk.Scrollbar(msg_frame, command=self.message_text.yview)
        msg_scroll.pack(side="right", fill="y")
        self.message_text.configure(yscrollcommand=msg_scroll.set)

        # 7. Signature toggle
        sig_frame = tk.Frame(body, bg=BG_COLOR)
        sig_frame.grid(row=7, column=1, sticky="w", padx=(10, 0), pady=(4, 0))
        self.include_signature_var = tk.BooleanVar(value=True)
        sig_check = tk.Checkbutton(
            sig_frame, text="Include my default Outlook signature",
            variable=self.include_signature_var, font=FONT_NORMAL, bg=BG_COLOR,
            fg=TEXT_COLOR, activebackground=BG_COLOR, selectcolor=PANEL_COLOR)
        sig_check.pack(side="left")

        # 8. Attachments
        tk.Label(body, text="Attachments:", font=FONT_NORMAL, bg=BG_COLOR,
                  fg=TEXT_COLOR).grid(row=8, column=0, sticky="nw", pady=(8, 4))
        attach_frame = tk.Frame(body, bg=BG_COLOR)
        attach_frame.grid(row=8, column=1, sticky="ew", padx=(10, 0), pady=(8, 4))
        attach_frame.grid_columnconfigure(0, weight=1)

        list_frame = tk.Frame(attach_frame, bg=BG_COLOR)
        list_frame.grid(row=0, column=0, sticky="ew")
        self.attachment_listbox = tk.Listbox(
            list_frame, font=FONT_NORMAL, height=3, relief="solid", bd=1,
            selectbackground=ACCENT_COLOR)
        self.attachment_listbox.pack(side="left", fill="both", expand=True)
        attach_scroll = tk.Scrollbar(list_frame, command=self.attachment_listbox.yview)
        attach_scroll.pack(side="right", fill="y")
        self.attachment_listbox.configure(yscrollcommand=attach_scroll.set)

        attach_btn_frame = tk.Frame(attach_frame, bg=BG_COLOR)
        attach_btn_frame.grid(row=1, column=0, sticky="w", pady=(6, 0))

        attach_btn = tk.Button(attach_btn_frame, text="Attach PDF(s) / Excel(s)...",
                                command=self._attach_files)
        style_button(attach_btn, "secondary")
        attach_btn.pack(side="left")

        remove_btn = tk.Button(attach_btn_frame, text="Remove Selected",
                                command=self._remove_selected_attachment)
        style_button(remove_btn, "danger")
        remove_btn.pack(side="left", padx=(6, 0))

        clear_btn = tk.Button(attach_btn_frame, text="Clear All",
                               command=self._clear_attachments)
        style_button(clear_btn, "secondary")
        clear_btn.pack(side="left", padx=(6, 0))

        # Bottom Actions Frame
        action_frame = tk.Frame(composer_outer, bg=BG_COLOR)
        action_frame.pack(fill="x", pady=(12, 0))

        self.preview_btn = tk.Button(action_frame, text="Preview in Outlook",
                                     command=self._preview_email)
        style_button(self.preview_btn, "secondary")
        self.preview_btn.pack(side="left")

        self.send_btn = tk.Button(action_frame, text="Send", command=self._send_email)
        style_button(self.send_btn, "primary")
        self.send_btn.pack(side="left", padx=10)

        # Inline Toast Status Label
        self.toast_label = tk.Label(
            action_frame, text="", font=(FONT_FAMILY, 9, "bold"),
            bg=BG_COLOR, fg="#16a34a"
        )
        self.toast_label.pack(side="left", padx=10)

        # -------------------------------------------------------------
        # RIGHT PANEL: Live Activity & Dispatch Log
        # -------------------------------------------------------------
        activity_panel = tk.Frame(
            main_split, bg="#ffffff", highlightbackground="#e2e8f0",
            highlightthickness=1, width=420
        )
        activity_panel.pack(side="right", fill="both", expand=False)
        activity_panel.pack_propagate(False)

        # Log Top Bar
        log_header = tk.Frame(activity_panel, bg="#ffffff", padx=14, pady=10)
        log_header.pack(fill="x")

        tk.Label(
            log_header, text="📋 Activity & Dispatch Log",
            font=(FONT_FAMILY, 11, "bold"), bg="#ffffff", fg="#0f172a"
        ).pack(side="left")

        clear_log_btn = tk.Label(
            log_header, text="Clear Log", font=(FONT_FAMILY, 8, "bold"),
            bg="#f1f5f9", fg="#64748b", padx=8, pady=3, cursor="hand2"
        )
        clear_log_btn.pack(side="right")
        clear_log_btn.bind("<Button-1>", lambda e: self._clear_activity_log())
        clear_log_btn.bind("<Enter>", lambda e: clear_log_btn.config(bg="#e2e8f0", fg="#0f172a"))
        clear_log_btn.bind("<Leave>", lambda e: clear_log_btn.config(bg="#f1f5f9", fg="#64748b"))

        # Status & KPI Banner
        status_bar = tk.Frame(activity_panel, bg="#f8fafc", padx=14, pady=8, highlightbackground="#e2e8f0", highlightthickness=1)
        status_bar.pack(fill="x", padx=12, pady=(0, 8))

        self.activity_status_dot = tk.Label(
            status_bar, text="●", font=(FONT_FAMILY, 9, "bold"),
            bg="#f8fafc", fg="#16a34a"
        )
        self.activity_status_dot.pack(side="left", padx=(0, 4))

        self.activity_status_text = tk.Label(
            status_bar, text="System Ready", font=(FONT_FAMILY, 8, "bold"),
            bg="#f8fafc", fg="#334155"
        )
        self.activity_status_text.pack(side="left")

        self.activity_kpi_text = tk.Label(
            status_bar, text="Session: 0", font=(FONT_FAMILY, 8),
            bg="#f8fafc", fg="#64748b"
        )
        self.activity_kpi_text.pack(side="right")

        # Rich Activity Log Text Box
        log_box_frame = tk.Frame(activity_panel, bg="#ffffff", padx=12, pady=4)
        log_box_frame.pack(fill="both", expand=True, pady=(0, 10))

        self.activity_text = tk.Text(
            log_box_frame, font=("Segoe UI", 9), wrap="word",
            bg="#ffffff", fg="#0f172a", relief="solid", bd=1,
            padx=10, pady=8
        )
        self.activity_text.pack(side="left", fill="both", expand=True)

        log_scroll = tk.Scrollbar(log_box_frame, command=self.activity_text.yview)
        log_scroll.pack(side="right", fill="y")
        self.activity_text.configure(yscrollcommand=log_scroll.set)

        # Configure Log Syntax Tags
        self.activity_text.tag_configure("tag_time", foreground="#64748b", font=("Segoe UI", 8, "bold"))
        self.activity_text.tag_configure("tag_success", foreground="#16a34a", font=("Segoe UI", 9, "bold"))
        self.activity_text.tag_configure("tag_sending", foreground="#d97706", font=("Segoe UI", 9, "bold"))
        self.activity_text.tag_configure("tag_error", foreground="#dc2626", font=("Segoe UI", 9, "bold"))
        self.activity_text.tag_configure("tag_info", foreground="#0284c7", font=("Segoe UI", 9, "bold"))
        self.activity_text.tag_configure("tag_label", foreground="#475569", font=("Segoe UI", 8, "bold"))
        self.activity_text.tag_configure("tag_val", foreground="#0f172a", font=("Segoe UI", 8))
        self.activity_text.tag_configure("tag_val_bold", foreground="#0f172a", font=("Segoe UI", 8, "bold"))
        self.activity_text.tag_configure("tag_val_accent", foreground="#0284c7", font=("Segoe UI", 8, "bold"))
        self.activity_text.tag_configure("tag_sep", foreground="#cbd5e1")

        self.activity_text.configure(state="disabled")

    # -- Activity Log Handling ------------------------------------------------

    def _clear_activity_log(self):
        self.activity_text.configure(state="normal")
        self.activity_text.delete("1.0", tk.END)
        self.activity_text.configure(state="disabled")

    def _append_activity_log(self, status, gate_pass, warehouse, send_type, subject,
                              to_addr=None, cc_addr=None, attachments=None, details=None,
                              cloud_logged=True, is_history=False):
        now_str = datetime.datetime.now().strftime("%H:%M:%S")

        self.activity_text.configure(state="normal")

        # Header with Timestamp & Status
        self.activity_text.insert(tk.END, f"[{now_str}] ", "tag_time")
        if status == "SUCCESS":
            self.activity_text.insert(tk.END, "✓ SUCCESSFULLY SENT\n", "tag_success")
        elif status == "SENDING":
            self.activity_text.insert(tk.END, "⏳ SENDING DISPATCH...\n", "tag_sending")
        elif status == "FAILED":
            self.activity_text.insert(tk.END, "✗ SEND FAILED\n", "tag_error")
        elif status == "HISTORY":
            self.activity_text.insert(tk.END, "● PREVIOUS DISPATCH\n", "tag_info")
        else:
            self.activity_text.insert(tk.END, f"{status}\n", "tag_info")

        # Core Metadata Rows
        self.activity_text.insert(tk.END, "  • Gate Pass No : ", "tag_label")
        self.activity_text.insert(tk.END, f"{gate_pass}\n", "tag_val_accent")

        self.activity_text.insert(tk.END, "  • Warehouse    : ", "tag_label")
        self.activity_text.insert(tk.END, f"{warehouse}\n", "tag_val_bold")

        if send_type:
            self.activity_text.insert(tk.END, "  • Type         : ", "tag_label")
            self.activity_text.insert(tk.END, f"{send_type}\n", "tag_val")

        if subject:
            self.activity_text.insert(tk.END, "  • Subject      : ", "tag_label")
            self.activity_text.insert(tk.END, f"{subject}\n", "tag_val")

        if to_addr:
            recipients_summary = to_addr
            if cc_addr:
                cc_count = len([c for c in cc_addr.split(";") if c.strip()])
                recipients_summary += f" (+{cc_count} CC)"
            self.activity_text.insert(tk.END, "  • Recipients   : ", "tag_label")
            self.activity_text.insert(tk.END, f"{recipients_summary}\n", "tag_val")

        if attachments is not None:
            att_count = len(attachments) if isinstance(attachments, list) else int(attachments)
            att_text = f"{att_count} file(s) attached" if att_count > 0 else "None"
            self.activity_text.insert(tk.END, "  • Attachments  : ", "tag_label")
            self.activity_text.insert(tk.END, f"{att_text}\n", "tag_val")

        if status == "SUCCESS":
            cloud_txt = "✓ Synced to Google Sheet & Excel" if cloud_logged else "Saved to local sent_log.xlsx"
            self.activity_text.insert(tk.END, "  • Cloud Sync   : ", "tag_label")
            self.activity_text.insert(tk.END, f"{cloud_txt}\n", "tag_success")
        elif status == "FAILED" and details:
            self.activity_text.insert(tk.END, "  • Error Details: ", "tag_label")
            self.activity_text.insert(tk.END, f"{details}\n", "tag_error")

        self.activity_text.insert(tk.END, "─" * 46 + "\n", "tag_sep")
        self.activity_text.see(tk.END)
        self.activity_text.configure(state="disabled")

    def _populate_recent_activity_logs(self):
        def _read():
            records = self.sent_log.get_recent_records(limit=4)
            if not records:
                return

            def _apply():
                for r in reversed(records):
                    self._append_activity_log(
                        status="HISTORY",
                        gate_pass=r.get("number", "N/A"),
                        warehouse=r.get("warehouse", "General"),
                        send_type=r.get("type", ""),
                        subject=f"{r.get('warehouse', '')} {r.get('type', '')} - {r.get('number', '')}".strip(),
                        is_history=True
                    )

            try:
                self.root.after(0, _apply)
            except Exception:
                pass

        threading.Thread(target=_read, daemon=True).start()

    def _extract_gate_pass_and_warehouse(self, send_type, subject, template_name):
        warehouse = "General / Direct"
        if template_name and template_name != TEMPLATE_PLACEHOLDER:
            warehouse = template_name.strip()

        gate_pass = ""
        if send_type:
            number_part, type_part = SentLogStore._split_subject(send_type, subject)
            if number_part:
                gate_pass = number_part.strip()
            if warehouse == "General / Direct" and type_part:
                cleaned_wh = type_part.replace(send_type, "").strip().rstrip("-").strip()
                if cleaned_wh:
                    warehouse = cleaned_wh

        if not gate_pass:
            if " - " in (subject or ""):
                parts = subject.split(" - ", 1)
                gate_pass = parts[1].strip()
                if warehouse == "General / Direct" and parts[0].strip():
                    warehouse = parts[0].strip()
            else:
                gate_pass = (subject or "").strip() or "N/A"

        return gate_pass or "N/A", warehouse or "General / Direct"

    def _show_send_success_toast(self, gate_pass, warehouse):
        msg = f"✓ Sent: {gate_pass} ({warehouse})"
        self.toast_label.config(text=msg, fg="#16a34a")
        self.root.after(4500, lambda: self.toast_label.config(text="") if hasattr(self, 'toast_label') and self.toast_label.winfo_exists() else None)

    # -- Settings handling ----------------------------------------------------

    def _open_settings_dialog(self):
        if self.on_open_settings is not None:
            self.on_open_settings()
        else:
            SettingsDialog(self.root, self.config_store, on_save=self._refresh_counts_label)

    # -- Template handling ----------------------------------------------------

    def _refresh_template_dropdown(self, select_name=None):
        names = self.template_store.get_names()
        self.template_combo.configure(values=[TEMPLATE_PLACEHOLDER] + names)
        if select_name and select_name in names:
            self.template_var.set(select_name)
            self._on_template_selected()
        else:
            self.template_var.set(TEMPLATE_PLACEHOLDER)

    def _on_template_selected(self, event=None):
        name = self.template_var.get()
        if not name or name == TEMPLATE_PLACEHOLDER:
            return
        template = self.template_store.get_template(name)
        if not template:
            return
        self.to_entry.delete(0, tk.END)
        self.to_entry.insert(0, template["to"])
        self.cc_entry.delete(0, tk.END)
        self.cc_entry.insert(0, template["cc"])

        current_type = self.selected_type_var.get()
        if current_type:
            self._apply_subject_prefix(current_type)

    def _open_add_template_dialog(self):
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror(
                "openpyxl not installed",
                "Install openpyxl to save templates:\n\npip install openpyxl")
            return

        def submit(name, to_addr, cc_addr):
            self.template_store.add_or_update_template(name, to_addr, cc_addr)
            self._refresh_template_dropdown(select_name=name)

        AddTemplateDialog(self.root, on_submit=submit)

    # -- GDN/GRN type selection -------------------------------------------------

    def _build_subject_prefix(self, type_name):
        template_name = self.template_var.get()
        if template_name and template_name != TEMPLATE_PLACEHOLDER:
            return f"{template_name} {type_name} - "
        return f"{type_name} - "

    def _apply_subject_prefix(self, type_name):
        current_text = self.subject_entry.get()
        if self._last_prefix and current_text.startswith(self._last_prefix):
            number_part = current_text[len(self._last_prefix):]
        else:
            number_part = ""
        prefix = self._build_subject_prefix(type_name)
        self.subject_entry.delete(0, tk.END)
        self.subject_entry.insert(0, prefix + number_part)
        self._last_prefix = prefix

    def _select_type(self, type_name):
        turning_off = type_name is not None and self.selected_type_var.get() == type_name
        new_type = "" if turning_off else (type_name or "")
        self.selected_type_var.set(new_type)

        if new_type:
            if self._subject_before_type is None:
                self._subject_before_type = self.subject_entry.get()
            self._apply_subject_prefix(new_type)
        else:
            if self._subject_before_type is not None:
                self.subject_entry.delete(0, tk.END)
                self.subject_entry.insert(0, self._subject_before_type)
            self._subject_before_type = None
            self._last_prefix = None

        style_button(self.gdn_btn,
                     "primary" if self.selected_type_var.get() == SEND_TYPE_GDN
                     else "secondary")
        style_button(self.grn_btn,
                     "primary" if self.selected_type_var.get() == SEND_TYPE_GRN
                     else "secondary")

    # -- Send-tracking (GDN/GRN log) -------------------------------------------

    def _refresh_counts_label(self):
        def _fetch():
            counts = self.sent_log.get_counts()
            text = f"GDN sent: {counts.get('GDN', 0)}    GRN sent: {counts.get('GRN', 0)}"
            try:
                self.root.after(0, lambda: self.counts_var.set(text))
            except Exception:
                pass

        threading.Thread(target=_fetch, daemon=True).start()

    def open_records_window(self):
        self._open_records_window()

    def _open_records_window(self):
        sheet_url = self.config_store.get_sheet_url()
        if not sheet_url:
            messagebox.showwarning(
                "Missing URL",
                "Please set your Google Sheet URL in ⚙ Settings first."
            )
            return

        try:
            threading.Thread(target=self.sent_log.get_counts, daemon=True).start()
        except Exception:
            pass

        try:
            webbrowser.open(sheet_url)
        except Exception as exc:
            messagebox.showerror(
                "Browser Error",
                f"Could not open the web browser:\n\n{exc}"
            )

    # -- Attachment handling -------------------------------------------------

    def _attach_files(self):
        paths = filedialog.askopenfilenames(
            title="Select PDF and/or Excel files to attach",
            filetypes=ATTACHMENT_FILETYPES,
        )
        if not paths:
            return
        for path in paths:
            if path not in self.attachment_paths:
                self.attachment_paths.append(path)
        self._refresh_attachment_list()

    def _remove_selected_attachment(self):
        selection = list(self.attachment_listbox.curselection())
        if not selection:
            return
        for index in reversed(selection):
            del self.attachment_paths[index]
        self._refresh_attachment_list()

    def _clear_attachments(self):
        self.attachment_paths = []
        self._refresh_attachment_list()

    def _refresh_attachment_list(self):
        self.attachment_listbox.delete(0, tk.END)
        for path in self.attachment_paths:
            self.attachment_listbox.insert(tk.END, os.path.basename(path))

    # -- Gathering / validating form data ------------------------------------

    def _gather_fields(self):
        to_addr = normalize_recipients(self.to_entry.get().strip())
        cc_addr = normalize_recipients(self.cc_entry.get().strip())
        subject = self.subject_entry.get().strip()
        body = self.message_text.get("1.0", "end-1c")
        return to_addr, cc_addr, subject, body

    def _validate(self, to_addr, subject):
        if not to_addr:
            messagebox.showwarning("Missing recipient",
                                    "Please enter at least one 'To:' recipient.")
            return False
        if not subject:
            if not messagebox.askyesno(
                    "No subject",
                    "The subject line is empty. Send anyway?"):
                return False
        return True

    # -- Actions (Asynchronous Non-Blocking Send) ---------------------------

    def _preview_email(self):
        to_addr, cc_addr, subject, body = self._gather_fields()
        if not self._validate(to_addr, subject):
            return
        try:
            mail = self.outlook.build_mail(
                to_addr, cc_addr, subject, body, self.attachment_paths,
                include_signature=self.include_signature_var.get())
            self.outlook.preview(mail)
        except Exception as exc:
            messagebox.showerror("Outlook error",
                                  f"Could not open the preview in Outlook:\n\n{exc}")

    def _send_email(self):
        if self._is_sending:
            return

        to_addr, cc_addr, subject, body = self._gather_fields()
        if not self._validate(to_addr, subject):
            return

        if self.attachment_paths:
            attachment_summary = "\n".join(
                f"  - {os.path.basename(p)}" for p in self.attachment_paths)
        else:
            attachment_summary = "  (none)"
        send_type = self.selected_type_var.get() or None
        confirm_text = (
            f"To: {to_addr}\n"
            f"CC: {cc_addr if cc_addr else '(none)'}\n"
            f"Subject: {subject if subject else '(none)'}\n"
            f"Type: {send_type if send_type else '(none)'}\n"
            f"Attachments:\n{attachment_summary}\n\n"
            "Send this email now?"
        )
        if not messagebox.askyesno("Confirm send", confirm_text):
            return

        template_name = self.template_var.get()
        gate_pass_num, warehouse_name = self._extract_gate_pass_and_warehouse(
            send_type, subject, template_name
        )

        # Lock UI into Sending state
        self._is_sending = True
        self.send_btn.config(text="⏳ Sending...", state="disabled", bg="#94a3b8")
        self.preview_btn.config(state="disabled")
        self.activity_status_dot.config(fg="#d97706")
        self.activity_status_text.config(text="Sending in background...", fg="#d97706")

        # Initial Sending Log entry
        self._append_activity_log(
            status="SENDING",
            gate_pass=gate_pass_num,
            warehouse=warehouse_name,
            send_type=send_type,
            subject=subject,
            to_addr=to_addr,
            cc_addr=cc_addr,
            attachments=list(self.attachment_paths)
        )

        attachments_copy = list(self.attachment_paths)
        include_sig = self.include_signature_var.get()

        # Non-blocking async execution
        threading.Thread(
            target=self._send_email_worker,
            args=(to_addr, cc_addr, subject, body, attachments_copy, include_sig,
                  send_type, gate_pass_num, warehouse_name),
            daemon=True
        ).start()

    def _send_email_worker(self, to_addr, cc_addr, subject, body, attachment_paths,
                           include_sig, send_type, gate_pass_num, warehouse_name):
        try:
            if WIN32COM_AVAILABLE:
                pythoncom.CoInitialize()

            app = win32com.client.Dispatch("Outlook.Application")
            mail = app.CreateItem(OUTLOOK_MAIL_ITEM)
            mail.To = to_addr
            if cc_addr:
                mail.CC = cc_addr
            mail.Subject = subject

            signature_html = ""
            if include_sig:
                try:
                    signature_html = get_default_signature_html()
                except Exception:
                    signature_html = ""

            mail.HTMLBody = OutlookClient._compose_html(body, signature_html)

            for path in attachment_paths:
                if os.path.isfile(path):
                    mail.Attachments.Add(path)

            mail.Send()

            cloud_logged = False
            if send_type:
                self.sent_log.log_send(send_type, subject)
                cloud_logged = bool(self.config_store.get_webapp_url())

            def on_success():
                self._is_sending = False
                style_button(self.send_btn, "primary")
                self.send_btn.config(text="Send", state="normal")
                self.preview_btn.config(state="normal")
                self.activity_status_dot.config(fg="#16a34a")
                self.activity_status_text.config(text="System Ready", fg="#334155")
                self.session_dispatch_count += 1
                self.activity_kpi_text.config(text=f"Session: {self.session_dispatch_count}")

                self._append_activity_log(
                    status="SUCCESS",
                    gate_pass=gate_pass_num,
                    warehouse=warehouse_name,
                    send_type=send_type,
                    subject=subject,
                    to_addr=to_addr,
                    cc_addr=cc_addr,
                    attachments=attachment_paths,
                    cloud_logged=cloud_logged
                )

                self._refresh_counts_label()
                self._show_send_success_toast(gate_pass_num, warehouse_name)
                self._reset_after_send()

            self.root.after(0, on_success)

        except Exception as exc:
            err_msg = str(exc)

            def on_error(err=err_msg):
                self._is_sending = False
                style_button(self.send_btn, "primary")
                self.send_btn.config(text="Send", state="normal")
                self.preview_btn.config(state="normal")
                self.activity_status_dot.config(fg="#dc2626")
                self.activity_status_text.config(text="Send Failed", fg="#dc2626")

                self._append_activity_log(
                    status="FAILED",
                    gate_pass=gate_pass_num,
                    warehouse=warehouse_name,
                    send_type=send_type,
                    subject=subject,
                    to_addr=to_addr,
                    cc_addr=cc_addr,
                    attachments=attachment_paths,
                    details=err
                )
                messagebox.showerror("Outlook Send Error", f"Could not send email:\n\n{err}")

            self.root.after(0, on_error)

        finally:
            if WIN32COM_AVAILABLE:
                try:
                    pythoncom.CoUninitialize()
                except Exception:
                    pass

    def _reset_after_send(self):
        self.to_entry.delete(0, tk.END)
        self.cc_entry.delete(0, tk.END)
        self.subject_entry.delete(0, tk.END)
        self.template_var.set(TEMPLATE_PLACEHOLDER)
        self._clear_attachments()
        self._subject_before_type = None
        self._select_type(None)


# --------------------------------------------------------------------------

if __name__ == "__main__":
    try:
        import ctypes
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID('efl.nexus.outlook')
    except Exception:
        pass

    root = tk.Tk()
    app = OutlookEmailApp(root)
    root.mainloop()
