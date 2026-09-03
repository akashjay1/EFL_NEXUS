import sys
import html
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import os
import json
from datetime import datetime
import shutil
import threading
import warnings
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
from PIL import Image, ImageDraw, ImageFont, ImageTk

warnings.filterwarnings('ignore')

class TextToColumnsDialog:
    """Dialog for configuring text to columns with draggable split markers and details panel"""
    def __init__(self, parent, title, column_data, column_name=None, app=None):
        self.parent = parent
        self.title = title
        self.column_data = column_data
        self.column_name = column_name
        self.app = app  # reference to ReconciliationApp, used to match its theme
        self.result = None
        self.widths = []
        self.split_positions = []
        self.dragging = False
        self.drag_index = -1
        self.char_width = 9
        self.x_start = 50
        self.y_start = 20
        self.line_height = 25
        self.preview_rows = min(10, len(column_data))
        
        self.dialog = tk.Toplevel(parent)
        self.dialog.title(title)
        self.dialog.transient(parent)
        self.dialog.resizable(True, True)
        self.dialog.minsize(1000, 700)
        self._size_dialog(parent)
        self.dialog.grab_set()
        
        self.create_widgets()
        self._apply_dialog_theme()
        
    def _size_dialog(self, parent):
        """Open the dialog large enough to use right away (as a normal sized
        window, not maximized), centered over the parent window."""
        self.dialog.update_idletasks()
        parent.update_idletasks()
        pw = parent.winfo_width() or 1200
        ph = parent.winfo_height() or 850
        px = parent.winfo_rootx()
        py = parent.winfo_rooty()
        w = max(1000, int(pw * 0.85))
        h = max(700, int(ph * 0.85))
        x = max(0, px + (pw - w) // 2)
        y = max(0, py + (ph - h) // 2)
        self.dialog.geometry(f"{w}x{h}+{x}+{y}")
        
    def _apply_dialog_theme(self):
        """Match this dialog's plain tk widgets to the main app's current
        light/dark theme (ttk widgets already follow the shared ttk.Style)."""
        if self.app is None:
            return
        palette = self.app.DARK_PALETTE if self.app.dark_mode.get() else self.app.LIGHT_PALETTE
        
        self.dialog.configure(bg=palette['bg'])
        self.instruction_label.configure(bg=palette['bg'], fg=palette['fg'])
        self.canvas.configure(bg=palette['panel_bg'], highlightcolor=palette['subtle_fg'])
        
        for text_widget in (self.result_text, self.sample_text, self.split_info_text):
            state = text_widget.cget('state')
            text_widget.configure(state=tk.NORMAL, bg=palette['entry_bg'], fg=palette['entry_fg'],
                                   insertbackground=palette['entry_fg'])
            text_widget.configure(state=state)
        
        self.warning_label.configure(bg=palette['bg'])
        
    def create_widgets(self):
        # Main container with paned window
        main_paned = ttk.PanedWindow(self.dialog, orient=tk.VERTICAL)
        main_paned.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Top frame - Preview
        top_frame = ttk.Frame(main_paned)
        main_paned.add(top_frame, weight=2)
        
        # Bottom frame - Details and Preview
        bottom_frame = ttk.Frame(main_paned)
        main_paned.add(bottom_frame, weight=1)
        
        # ===== TOP FRAME: Preview with Splits =====
        self.instruction_label = tk.Label(
            top_frame,
            text="Drag the ▼ arrows to set split points | Click on text to add splits | Click on arrow to remove",
            font=('Segoe UI', 10)
        )
        self.instruction_label.pack(pady=5)
        
        preview_frame = ttk.LabelFrame(top_frame, text="Drag arrows to split text", padding="10")
        preview_frame.pack(fill=tk.BOTH, expand=True)
        
        self.canvas_frame = ttk.Frame(preview_frame)
        self.canvas_frame.pack(fill=tk.BOTH, expand=True)
        
        self.canvas = tk.Canvas(
            self.canvas_frame,
            bg='white',
            height=300,
            highlightthickness=1,
            highlightcolor='gray'
        )
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        scrollbar = ttk.Scrollbar(self.canvas_frame, orient="vertical", command=self.canvas.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.canvas.configure(yscrollcommand=scrollbar.set)
        
        # Bind mouse events
        self.canvas.bind("<Button-1>", self.on_mouse_down)
        self.canvas.bind("<B1-Motion>", self.on_mouse_drag)
        self.canvas.bind("<ButtonRelease-1>", self.on_mouse_up)
        self.canvas.bind("<Motion>", self.on_mouse_move)
        
        # ===== BOTTOM FRAME: Details Panel =====
        details_paned = ttk.PanedWindow(bottom_frame, orient=tk.HORIZONTAL)
        details_paned.pack(fill=tk.BOTH, expand=True)
        
        # Left: Column Details
        details_frame = ttk.LabelFrame(details_paned, text="📋 Column Details", padding="10")
        details_paned.add(details_frame, weight=1)
        
        self.create_details_panel(details_frame)
        
        # Right: Result Preview
        result_frame = ttk.LabelFrame(details_paned, text="📊 Result Preview", padding="10")
        details_paned.add(result_frame, weight=1)
        
        self.result_text = tk.Text(
            result_frame,
            height=8,
            font=('Consolas', 9),
            wrap=tk.NONE,
            state=tk.DISABLED
        )
        self.result_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        result_scrollbar = ttk.Scrollbar(result_frame, orient="vertical", command=self.result_text.yview)
        result_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.result_text.config(yscrollcommand=result_scrollbar.set)
        
        # ===== BUTTONS =====
        button_frame = ttk.Frame(self.dialog)
        button_frame.pack(fill=tk.X, padx=20, pady=10)
        
        ttk.Button(
            button_frame,
            text="Clear Splits",
            command=self.clear_splits
        ).pack(side=tk.LEFT, padx=(0, 10))
        
        ttk.Button(
            button_frame,
            text="Auto-Detect",
            command=self.auto_detect_splits
        ).pack(side=tk.LEFT, padx=(0, 10))
        
        ttk.Button(
            button_frame,
            text="Refresh Preview",
            command=self.preview_result
        ).pack(side=tk.LEFT, padx=(0, 10))
        
        ttk.Button(
            button_frame,
            text="Apply Split",
            command=self.apply_split
        ).pack(side=tk.RIGHT, padx=(0, 10))
        
        ttk.Button(
            button_frame,
            text="Cancel",
            command=self.dialog.destroy
        ).pack(side=tk.RIGHT)
        
        # Status bar
        self.status_var = tk.StringVar(value="Ready - Drag arrows to set split points")
        status_bar = ttk.Label(
            self.dialog,
            textvariable=self.status_var,
            relief=tk.SUNKEN,
            anchor=tk.W,
            padding=(5, 2)
        )
        status_bar.pack(fill=tk.X, padx=10, pady=(0, 10))
        
        # Initial render
        self.render_preview()
        self.update_details()
        
    def create_details_panel(self, parent):
        """Create the details panel with column information"""
        # Column info
        info_frame = ttk.Frame(parent)
        info_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Column name
        name_frame = ttk.Frame(info_frame)
        name_frame.pack(fill=tk.X, pady=2)
        ttk.Label(name_frame, text="Column Name:", font=('Segoe UI', 9, 'bold')).pack(side=tk.LEFT)
        ttk.Label(name_frame, text=self.column_name or "Unknown", foreground='blue').pack(side=tk.LEFT, padx=(5, 0))
        
        # Total rows
        rows_frame = ttk.Frame(info_frame)
        rows_frame.pack(fill=tk.X, pady=2)
        ttk.Label(rows_frame, text="Total Rows:", font=('Segoe UI', 9, 'bold')).pack(side=tk.LEFT)
        ttk.Label(rows_frame, text=str(len(self.column_data)), foreground='green').pack(side=tk.LEFT, padx=(5, 0))
        
        # Sample data
        ttk.Label(parent, text="Sample Data (first 5 rows):", font=('Segoe UI', 9, 'bold')).pack(anchor=tk.W, pady=(10, 5))
        
        sample_frame = ttk.Frame(parent)
        sample_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.sample_text = tk.Text(
            sample_frame,
            height=5,
            font=('Consolas', 9),
            wrap=tk.NONE,
            bg='#F5F5F5'
        )
        self.sample_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        sample_scrollbar = ttk.Scrollbar(sample_frame, orient="vertical", command=self.sample_text.yview)
        sample_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.sample_text.config(yscrollcommand=sample_scrollbar.set)
        
        # Populate sample
        self.sample_text.insert(tk.END, "Row | Data\n")
        self.sample_text.insert(tk.END, "-" * 60 + "\n")
        for i, text in enumerate(self.column_data[:5]):
            self.sample_text.insert(tk.END, f"{i+1:3d} | {text}\n")
        self.sample_text.config(state=tk.DISABLED)
        
        # Split info
        split_info_frame = ttk.LabelFrame(parent, text="Split Information", padding="10")
        split_info_frame.pack(fill=tk.X, pady=(10, 0))
        
        self.split_info_text = tk.Text(
            split_info_frame,
            height=4,
            font=('Consolas', 9),
            wrap=tk.WORD,
            bg='#FFF8E1'
        )
        self.split_info_text.pack(fill=tk.X)
        self.split_info_text.insert(tk.END, "No splits configured yet.\nDrag the ▼ arrows above to set split points.")
        self.split_info_text.config(state=tk.DISABLED)
        
        # Warning label
        self.warning_label = tk.Label(
            parent,
            text="",
            font=('Segoe UI', 9, 'bold'),
            fg='red',
            wraplength=400
        )
        self.warning_label.pack(fill=tk.X, pady=(10, 0))
        
    def update_details(self):
        """Update the details panel with current split information"""
        if not self.split_positions:
            self.split_info_text.config(state=tk.NORMAL)
            self.split_info_text.delete(1.0, tk.END)
            self.split_info_text.insert(tk.END, "No splits configured yet.\nDrag the ▼ arrows above to set split points.")
            self.split_info_text.config(state=tk.DISABLED)
            self.warning_label.config(text="")
            return
        
        max_length = self.get_max_length()
        widths = []
        prev = 0
        for pos in sorted(self.split_positions):
            widths.append(pos - prev)
            prev = pos
        widths.append(max_length - prev)
        
        self.split_info_text.config(state=tk.NORMAL)
        self.split_info_text.delete(1.0, tk.END)
        
        # Show split positions and widths
        self.split_info_text.insert(tk.END, f"Split Points: {len(self.split_positions)}\n")
        self.split_info_text.insert(tk.END, f"Positions: {', '.join(map(str, sorted(self.split_positions)))}\n")
        self.split_info_text.insert(tk.END, f"Column Widths: {', '.join(map(str, widths))}\n")
        self.split_info_text.insert(tk.END, f"Total Columns: {len(widths)}\n")
        
        # Show column preview
        self.split_info_text.insert(tk.END, "\nColumn Preview (first row):\n")
        if self.column_data:
            text = str(self.column_data[0])
            parts = []
            pos = 0
            for width in widths[:-1]:
                if pos < len(text):
                    parts.append(f"[{text[pos:pos+width].strip()}]")
                    pos += width
                else:
                    parts.append("[]")
            if pos < len(text):
                parts.append(f"[{text[pos:].strip()}]")
            else:
                parts.append("[]")
            self.split_info_text.insert(tk.END, " | ".join(parts))
        
        self.split_info_text.config(state=tk.DISABLED)
        
        # Check for potential issues
        warnings = []
        if len(widths) > 10:
            warnings.append("⚠️ Many columns created (>10) - verify this is intended")
        if any(w < 2 for w in widths[:-1]):
            warnings.append("⚠️ Very narrow columns detected (width < 2 chars)")
        if widths and widths[-1] > 50:
            warnings.append("ℹ️ Last column is very wide - may contain remaining data")
        
        if warnings:
            self.warning_label.config(text="\n".join(warnings))
        else:
            self.warning_label.config(text="✅ Split configuration looks good")
        
        # Update status
        self.status_var.set(f"Ready - {len(widths)} columns, widths: {', '.join(map(str, widths))}")
        
    def get_char_position(self, x):
        """Convert pixel x position to character position"""
        x -= self.x_start
        if x < 0:
            return 0
        return int(x / self.char_width)
    
    def get_max_length(self):
        """Get maximum length of sample data"""
        sample_data = self.column_data[:self.preview_rows]
        return max([len(str(text)) for text in sample_data]) if sample_data else 0
    
    def render_preview(self, highlight_positions=None):
        """Render the preview with split markers"""
        self.canvas.delete("all")
        
        if not self.column_data:
            self.canvas.create_text(10, 10, text="No data to preview", anchor=tk.NW)
            return
        
        sample_data = self.column_data[:self.preview_rows]
        max_length = self.get_max_length()
        
        # Draw column headers
        self.canvas.create_text(self.x_start - 10, self.y_start - 5, text="#", anchor=tk.E, font=('Segoe UI', 9, 'bold'))
        self.canvas.create_text(self.x_start, self.y_start - 5, text="Data", anchor=tk.W, font=('Segoe UI', 9, 'bold'))
        
        # Draw horizontal line
        self.canvas.create_line(
            self.x_start - 20, self.y_start + 5,
            self.x_start + max_length * self.char_width + 100, self.y_start + 5,
            fill='gray', width=1
        )
        
        # Draw sample data
        for i, text in enumerate(sample_data):
            y_pos = self.y_start + (i + 1) * self.line_height + 5
            text_str = str(text)
            
            self.canvas.create_text(
                self.x_start - 10, y_pos,
                text=str(i+1), anchor=tk.E,
                font=('Segoe UI', 8), fill='gray'
            )
            
            self.canvas.create_text(
                self.x_start, y_pos,
                text=text_str, anchor=tk.W,
                font=('Consolas', 9)
            )
            
            # Draw split markers for this row
            for pos in self.split_positions:
                if pos < len(text_str):
                    marker_x = self.x_start + pos * self.char_width
                    self.canvas.create_text(
                        marker_x, y_pos + 15,
                        text="▼",
                        fill='#E74C3C',
                        font=('Segoe UI', 10, 'bold')
                    )
                    self.canvas.create_line(
                        marker_x, self.y_start + 10,
                        marker_x, y_pos + 20,
                        fill='#E74C3C',
                        width=1,
                        dash=(3, 3)
                    )
        
        # Draw top markers (draggable)
        for i, pos in enumerate(self.split_positions):
            marker_x = self.x_start + pos * self.char_width
            self.canvas.create_rectangle(
                marker_x - 4, self.y_start - 18,
                marker_x + 4, self.y_start - 8,
                fill='#E74C3C', outline='#C0392B'
            )
            self.canvas.create_text(
                marker_x, self.y_start - 20,
                text="▼",
                fill='#E74C3C',
                font=('Segoe UI', 12, 'bold'),
                tags=(f"split_{i}",)
            )
            self.canvas.create_rectangle(
                marker_x - 8, self.y_start - 25,
                marker_x + 8, self.y_start - 5,
                fill='', outline='#E74C3C',
                width=2,
                tags=(f"handle_{i}",)
            )
        
        # Update canvas scroll region
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        
        # Update details
        self.update_details()
        
    def on_mouse_move(self, event):
        """Handle mouse movement"""
        pos = self.get_char_position(event.x)
        max_length = self.get_max_length()
        if 0 <= pos <= max_length:
            self.status_var.set(f"Position: {pos} | Click to add split at this position")
        else:
            self.status_var.set("Ready - Drag arrows to set split points")
        
        # Show cursor change when near a split handle
        for i, split_pos in enumerate(self.split_positions):
            marker_x = self.x_start + split_pos * self.char_width
            if abs(event.x - marker_x) < 15 and abs(event.y - self.y_start) < 30:
                self.canvas.config(cursor="sb_h_double_arrow")
                return
        self.canvas.config(cursor="")
        
    def on_mouse_down(self, event):
        """Handle mouse click"""
        # Check if clicking on a split handle
        for i, split_pos in enumerate(self.split_positions):
            marker_x = self.x_start + split_pos * self.char_width
            if abs(event.x - marker_x) < 15 and abs(event.y - self.y_start) < 30:
                # Remove this split
                self.split_positions.pop(i)
                self.split_positions.sort()
                self.render_preview()
                self.preview_result()
                return
        
        # Add new split
        pos = self.get_char_position(event.x)
        max_length = self.get_max_length()
        if 0 < pos < max_length:
            if pos not in self.split_positions:
                self.split_positions.append(pos)
                self.split_positions.sort()
                self.render_preview()
                self.preview_result()
                self.status_var.set(f"✅ Split added at position {pos}")
    
    def on_mouse_drag(self, event):
        """Handle dragging"""
        pos = self.get_char_position(event.x)
        max_length = self.get_max_length()
        
        # Find closest split to drag
        closest = None
        closest_dist = float('inf')
        for i, split_pos in enumerate(self.split_positions):
            marker_x = self.x_start + split_pos * self.char_width
            dist = abs(event.x - marker_x)
            if dist < closest_dist:
                closest_dist = dist
                closest = i
        
        if closest is not None and closest_dist < 30:
            pos = max(1, min(pos, max_length - 1))
            self.split_positions[closest] = pos
            self.split_positions.sort()
            self.render_preview()
            self.preview_result()
            self.status_var.set(f"🔄 Moving split to position {pos}")
    
    def on_mouse_up(self, event):
        """Handle mouse release"""
        self.canvas.config(cursor="")
        self.status_var.set("✅ Split position updated")
        
    def clear_splits(self):
        """Clear all split points"""
        self.split_positions = []
        self.render_preview()
        self.result_text.config(state=tk.NORMAL)
        self.result_text.delete(1.0, tk.END)
        self.result_text.config(state=tk.DISABLED)
        self.status_var.set("🧹 All splits cleared")
        
    def auto_detect_splits(self):
        """Auto-detect common split patterns"""
        if not self.column_data:
            return
        
        # Look for common delimiters in the first row
        text = str(self.column_data[0])
        
        # Check for spaces, tabs, multiple spaces
        import re
        patterns = [
            (r'\t', 'Tab'),  # Tabs
            (r'  +', 'Multiple spaces'),  # Multiple spaces
            (r',', 'Comma'),  # Commas
            (r'\|', 'Pipe'),  # Pipes
        ]
        
        for pattern, name in patterns:
            matches = list(re.finditer(pattern, text))
            if len(matches) >= 2:  # At least 2 delimiters
                positions = [m.start() for m in matches]
                self.split_positions = positions
                self.split_positions.sort()
                self.render_preview()
                self.preview_result()
                self.status_var.set(f"🔍 Auto-detected {len(positions)} splits using {name} delimiter")
                return
        
        # If no delimiters found, try to detect fixed width pattern
        # Look for consistent spacing in first few rows
        if len(self.column_data) >= 3:
            lengths = [len(str(text)) for text in self.column_data[:3]]
            if all(l > 10 for l in lengths):
                # Suggest splitting at 10, 20, 30 for wide data
                max_len = max(lengths)
                suggested = list(range(10, max_len, 10))
                if len(suggested) >= 2:
                    self.split_positions = suggested
                    self.split_positions.sort()
                    self.render_preview()
                    self.preview_result()
                    self.status_var.set(f"🔍 Auto-detected fixed-width splits at positions: {suggested}")
                    return
        
        self.status_var.set("⚠️ Could not auto-detect splits. Please set them manually.")
        
    def preview_result(self):
        """Preview the split result"""
        if not self.split_positions:
            self.result_text.config(state=tk.NORMAL)
            self.result_text.delete(1.0, tk.END)
            self.result_text.insert(tk.END, "No splits configured. Drag the arrows or click on text to add splits.")
            self.result_text.config(state=tk.DISABLED)
            return
        
        self.result_text.config(state=tk.NORMAL)
        self.result_text.delete(1.0, tk.END)
        
        # Calculate widths
        widths = []
        prev = 0
        for pos in sorted(self.split_positions):
            width = pos - prev
            if width > 0:
                widths.append(width)
            prev = pos
        widths.append(100)
        
        # Show headers
        header = " | ".join([f"Col {i+1:2d}" for i in range(len(widths))])
        separator = "-" * (len(header) + 10)
        self.result_text.insert(tk.END, f"Columns: {len(widths)}\n")
        self.result_text.insert(tk.END, f"Widths: {', '.join(map(str, widths))}\n")
        self.result_text.insert(tk.END, f"{separator}\n")
        self.result_text.insert(tk.END, f"{header}\n")
        self.result_text.insert(tk.END, f"{separator}\n")
        
        # Split and display data
        for i, text in enumerate(self.column_data[:10]):
            parts = []
            pos = 0
            for width in widths[:-1]:
                if pos < len(text):
                    parts.append(text[pos:pos+width].strip())
                    pos += width
                else:
                    parts.append("")
            if pos < len(text):
                parts.append(text[pos:].strip())
            else:
                parts.append("")
            
            result_line = " | ".join([f"{part:20s}" for part in parts])
            self.result_text.insert(tk.END, f"{i+1:2d}. {result_line}\n")
        
        self.result_text.see(tk.END)
        self.result_text.config(state=tk.DISABLED)
        
        # Update details
        self.update_details()
        
    def apply_split(self):
        """Apply the split and close dialog"""
        if not self.split_positions:
            messagebox.showwarning("Warning", "Please add split points by dragging the arrows or clicking on text.")
            return
        
        try:
            widths = []
            prev = 0
            for pos in sorted(self.split_positions):
                width = pos - prev
                if width > 0:
                    widths.append(width)
                prev = pos
            widths.append(100)
            
            split_data = []
            for text in self.column_data:
                parts = []
                pos = 0
                for width in widths[:-1]:
                    if pos < len(text):
                        parts.append(text[pos:pos+width].strip())
                        pos += width
                    else:
                        parts.append("")
                if pos < len(text):
                    parts.append(text[pos:].strip())
                else:
                    parts.append("")
                split_data.append(parts)
            
            self.widths = widths
            self.result = split_data
            self.dialog.destroy()
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to apply split: {str(e)}")


class ConcatenateColumnsDialog:
    """Dialog for combining two or more columns into a single new column.
    Shows a full data preview (like the main window), lets the user pick
    which columns to merge and in what order, choose a separator and a
    name for the new column, and preview the result before applying."""
    def __init__(self, parent, title, df, app=None):
        self.parent = parent
        self.title = title
        self.df = df
        self.app = app  # reference to ReconciliationApp, used to match its theme
        self.result = None
        self.available_columns = [str(c) for c in df.columns]
        self.selected_columns = []  # ordered list of columns chosen for concatenation
        self._header_badge_imgs = {}

        self.dialog = tk.Toplevel(parent)
        self.dialog.title(title)
        self.dialog.transient(parent)
        self.dialog.resizable(True, True)
        self.dialog.minsize(1000, 700)
        self._size_dialog(parent)
        self.dialog.grab_set()

        self.create_widgets()
        self._apply_dialog_theme()

    def _size_dialog(self, parent):
        """Open the dialog large enough to use right away, centered over the parent window."""
        self.dialog.update_idletasks()
        parent.update_idletasks()
        pw = parent.winfo_width() or 1200
        ph = parent.winfo_height() or 850
        px = parent.winfo_rootx()
        py = parent.winfo_rooty()
        w = max(1000, int(pw * 0.85))
        h = max(700, int(ph * 0.85))
        x = max(0, px + (pw - w) // 2)
        y = max(0, py + (ph - h) // 2)
        self.dialog.geometry(f"{w}x{h}+{x}+{y}")

    def _apply_dialog_theme(self):
        """Match this dialog's plain tk widgets to the main app's current light/dark theme."""
        if self.app is None:
            return
        palette = self.app.DARK_PALETTE if self.app.dark_mode.get() else self.app.LIGHT_PALETTE

        self.dialog.configure(bg=palette['bg'])
        self.instruction_label.configure(bg=palette['bg'], fg=palette['fg'])

        for text_widget in (self.preview_text,):
            state = text_widget.cget('state')
            text_widget.configure(state=tk.NORMAL, bg=palette['entry_bg'], fg=palette['entry_fg'],
                                   insertbackground=palette['entry_fg'])
            text_widget.configure(state=state)

        for listbox in (self.available_listbox, self.selected_listbox):
            listbox.configure(bg=palette['entry_bg'], fg=palette['entry_fg'],
                              selectbackground='#2563EB', selectforeground='#ffffff')

        self.data_tree.tag_configure('oddrow', background=palette['oddrow'])
        self.data_tree.tag_configure('evenrow', background=palette['evenrow'])
        self.highlight_selected_headers()

    def create_widgets(self):
        main_paned = ttk.PanedWindow(self.dialog, orient=tk.VERTICAL)
        main_paned.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # ===== TOP: Full data preview (same style as the main window) =====
        top_frame = ttk.Frame(main_paned)
        main_paned.add(top_frame, weight=2)

        self.instruction_label = tk.Label(
            top_frame,
            text="Preview of the sheet below. Pick columns to combine by clicking headers or using the lists below. Choose a separator and name for the new column.",
            font=('Segoe UI', 10),
            wraplength=900,
            justify=tk.LEFT
        )
        self.instruction_label.pack(pady=5, anchor=tk.W)

        data_preview_frame = ttk.LabelFrame(top_frame, text="📊 Data Preview (Click column header to select/deselect)", padding="10")
        data_preview_frame.pack(fill=tk.BOTH, expand=True)

        self.data_tree = self.build_data_preview(data_preview_frame)
        self.populate_data_preview()

        # ===== BOTTOM: Column selection + options + result preview =====
        bottom_paned = ttk.PanedWindow(main_paned, orient=tk.HORIZONTAL)
        main_paned.add(bottom_paned, weight=1)

        # --- Left: column picker (available <-> selected) ---
        picker_frame = ttk.LabelFrame(bottom_paned, text="📋 Choose Columns (in order)", padding="10")
        bottom_paned.add(picker_frame, weight=1)

        lists_row = ttk.Frame(picker_frame)
        lists_row.pack(fill=tk.BOTH, expand=True)

        avail_col = ttk.Frame(lists_row)
        avail_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        ttk.Label(avail_col, text="Available Columns").pack(anchor=tk.W)
        self.available_listbox = tk.Listbox(avail_col, selectmode=tk.EXTENDED, exportselection=False, height=10)
        self.available_listbox.pack(fill=tk.BOTH, expand=True)
        self.available_listbox.bind("<Double-1>", lambda e: self.add_selected_columns())
        self.available_listbox.bind("<<ListboxSelect>>", self._on_available_select)
        for col in self.available_columns:
            self.available_listbox.insert(tk.END, col)

        btn_col = ttk.Frame(lists_row)
        btn_col.pack(side=tk.LEFT, fill=tk.Y, padx=8)
        ttk.Button(btn_col, text="Add ▶", command=self.add_selected_columns).pack(pady=(20, 4))
        ttk.Button(btn_col, text="◀ Remove", command=self.remove_selected_columns).pack(pady=4)
        ttk.Button(btn_col, text="▲ Move Up", command=lambda: self.move_selected(-1)).pack(pady=4)
        ttk.Button(btn_col, text="▼ Move Down", command=lambda: self.move_selected(1)).pack(pady=4)

        sel_col = ttk.Frame(lists_row)
        sel_col.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        ttk.Label(sel_col, text="Columns to Concatenate (in order)").pack(anchor=tk.W)
        self.selected_listbox = tk.Listbox(sel_col, selectmode=tk.EXTENDED, exportselection=False, height=10)
        self.selected_listbox.pack(fill=tk.BOTH, expand=True)
        self.selected_listbox.bind("<Double-1>", lambda e: self.remove_selected_columns())
        self.selected_listbox.bind("<<ListboxSelect>>", self._on_selected_select)

        # --- Right: options + result preview ---
        options_frame = ttk.LabelFrame(bottom_paned, text="⚙️ Options & Preview", padding="10")
        bottom_paned.add(options_frame, weight=1)

        sep_row = ttk.Frame(options_frame)
        sep_row.pack(fill=tk.X, pady=4)
        ttk.Label(sep_row, text="Separator:", width=16).pack(side=tk.LEFT)
        self.separator_var = tk.StringVar(value="-")
        self.separator_combo = ttk.Combobox(
            sep_row,
            textvariable=self.separator_var,
            values=["-", " ", "_", ",", ", ", "|", "/", "(none)"],
            width=10
        )
        self.separator_combo.pack(side=tk.LEFT)

        name_row = ttk.Frame(options_frame)
        name_row.pack(fill=tk.X, pady=4)
        ttk.Label(name_row, text="New Column Name:", width=16).pack(side=tk.LEFT)
        self.new_column_name_var = tk.StringVar(value="Concatenated")
        ttk.Entry(name_row, textvariable=self.new_column_name_var).pack(side=tk.LEFT, fill=tk.X, expand=True)

        drop_row = ttk.Frame(options_frame)
        drop_row.pack(fill=tk.X, pady=4)
        self.drop_originals_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            drop_row,
            text="Remove original columns after concatenating",
            variable=self.drop_originals_var
        ).pack(side=tk.LEFT)

        ttk.Button(options_frame, text="Refresh Preview", command=self.preview_result).pack(pady=(6, 4), anchor=tk.W)

        ttk.Label(options_frame, text="Result Preview:", font=('Segoe UI', 9, 'bold')).pack(anchor=tk.W)
        preview_container = ttk.Frame(options_frame)
        preview_container.pack(fill=tk.BOTH, expand=True)
        self.preview_text = tk.Text(preview_container, height=8, font=('Consolas', 9), wrap=tk.NONE, state=tk.DISABLED)
        self.preview_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        preview_scroll = ttk.Scrollbar(preview_container, orient="vertical", command=self.preview_text.yview)
        preview_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.preview_text.config(yscrollcommand=preview_scroll.set)

        # ===== BUTTONS =====
        button_frame = ttk.Frame(self.dialog)
        button_frame.pack(fill=tk.X, padx=20, pady=10)

        ttk.Button(button_frame, text="Cancel", command=self.dialog.destroy).pack(side=tk.RIGHT)
        ttk.Button(button_frame, text="Apply Concatenate", command=self.apply_concatenate).pack(side=tk.RIGHT, padx=(0, 10))

        self.status_var = tk.StringVar(value="Ready - pick at least two columns to combine")
        status_bar = ttk.Label(self.dialog, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W, padding=(5, 2))
        status_bar.pack(fill=tk.X, padx=10, pady=(0, 10))

    def _on_available_select(self, event=None):
        selections = self.available_listbox.curselection()
        if selections:
            col = self.available_listbox.get(selections[-1])
            self.scroll_to_column(col)

    def _on_selected_select(self, event=None):
        selections = self.selected_listbox.curselection()
        if selections:
            col = self.selected_listbox.get(selections[-1])
            self.scroll_to_column(col)

    def scroll_to_column(self, col):
        """Scroll the data_tree horizontally so that the given column is visible."""
        try:
            columns = list(self.data_tree['columns'])
            if col not in columns:
                return
            idx = columns.index(col)
            total_cols = len(columns)
            if total_cols > 0:
                fraction = max(0.0, min(1.0, idx / total_cols))
                self.data_tree.xview_moveto(fraction)
        except Exception:
            pass

    def make_header_badge(self, text, bg_color="#2563EB", fg_color="#ffffff", height=20, icon="🔗"):
        """Create a rounded colored background badge image for Treeview heading."""
        full_text = f"{icon} {text}" if icon else str(text)
        try:
            font = ImageFont.truetype("segoeuib.ttf", 11)
        except Exception:
            try:
                font = ImageFont.truetype("arialbd.ttf", 11)
            except Exception:
                try:
                    font = ImageFont.truetype("segoeui.ttf", 11)
                except Exception:
                    font = ImageFont.load_default()

        try:
            bbox = font.getbbox(full_text)
            text_w = bbox[2] - bbox[0]
            text_h = bbox[3] - bbox[1]
        except Exception:
            text_w = len(full_text) * 7
            text_h = 12

        width = max(text_w + 16, 50)
        img = Image.new("RGBA", (width, height), (0, 0, 0, 0))
        draw = ImageDraw.Draw(img)

        # Draw filled colored rectangle with rounded corners
        draw.rounded_rectangle([0, 1, width - 1, height - 1], radius=3, fill=bg_color)

        # Draw text
        text_x = 8
        text_y = max(1, (height - text_h) // 2 - 1)
        draw.text((text_x, text_y), full_text, fill=fg_color, font=font)

        return ImageTk.PhotoImage(img), width

    def highlight_selected_headers(self):
        """Visually mark the selected columns in the preview Treeview header with colored badges and ordering."""
        tree = self.data_tree
        if tree is None:
            return
        columns = tree['columns']
        if not columns:
            return

        self._header_badge_imgs.clear()

        for col in columns:
            if col in self.selected_columns:
                order_num = self.selected_columns.index(col) + 1
                img, badge_w = self.make_header_badge(
                    col,
                    bg_color="#2563EB",
                    fg_color="#ffffff",
                    icon=f"🔗 #{order_num}"
                )
                extra = badge_w + 8
                try:
                    current_w = int(tree.column(col, 'width'))
                    if current_w < extra:
                        tree.column(col, width=extra)
                except Exception:
                    pass
                self._header_badge_imgs[col] = img
                tree.heading(col, image=img, text='', anchor=tk.W, command=lambda c=col: self.toggle_column_from_preview(c))
            else:
                tree.heading(col, text=col, image='', anchor=tk.W, command=lambda c=col: self.toggle_column_from_preview(c))

    def toggle_column_from_preview(self, col):
        """Toggle a column in the concatenation selection when clicking directly on its header in the preview table."""
        if col in self.selected_columns:
            self.selected_columns.remove(col)
        else:
            self.selected_columns.append(col)
        self.refresh_lists()

    def build_data_preview(self, parent):
        """Build a scrollable Treeview showing the full sheet, mirroring the main window's preview panes."""
        container = ttk.Frame(parent)
        container.pack(fill=tk.BOTH, expand=True)

        vsb = ttk.Scrollbar(container, orient="vertical")
        hsb = ttk.Scrollbar(container, orient="horizontal")

        tree = ttk.Treeview(
            container,
            show="headings",
            height=8,
            yscrollcommand=vsb.set,
            xscrollcommand=hsb.set,
            selectmode="browse"
        )

        vsb.config(command=tree.yview)
        hsb.config(command=tree.xview)

        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        container.rowconfigure(0, weight=1)
        container.columnconfigure(0, weight=1)

        tree.tag_configure('oddrow', background='#F5F5F5')
        tree.tag_configure('evenrow', background='#FFFFFF')

        return tree

    def populate_data_preview(self):
        """Fill the data preview Treeview with the sheet's data (first 200 rows)."""
        tree = self.data_tree
        tree.delete(*tree.get_children())
        columns = self.available_columns
        tree['columns'] = columns

        for col in columns:
            tree.heading(col, text=col, anchor=tk.W, command=lambda c=col: self.toggle_column_from_preview(c))
            col_str = str(col)
            try:
                data_max_len = int(self.df[col].astype(str).str.len().max()) if len(self.df) else len(col_str)
            except Exception:
                data_max_len = len(col_str)
            header_w = len(col_str) * 7.2 + 18
            data_w = min(200, data_max_len * 6.8 + 14)
            width = max(50, int(max(header_w, data_w)))
            tree.column(col, width=width, minwidth=40, anchor=tk.W, stretch=False)

        max_rows = 200
        preview_df = self.df.head(max_rows)
        for i, (_, row) in enumerate(preview_df.iterrows()):
            values = ["" if pd.isna(v) else str(v) for v in row.tolist()]
            tag = 'evenrow' if i % 2 == 0 else 'oddrow'
            tree.insert("", tk.END, iid=str(i), values=values, tags=(tag,))

        self.highlight_selected_headers()

    def add_selected_columns(self):
        """Move the highlighted available columns into the selected (ordered) list."""
        picks = [self.available_listbox.get(i) for i in self.available_listbox.curselection()]
        for col in picks:
            if col not in self.selected_columns:
                self.selected_columns.append(col)
        self.refresh_lists()

    def remove_selected_columns(self):
        """Move the highlighted selected columns back to the available list."""
        picks = set(self.selected_listbox.get(i) for i in self.selected_listbox.curselection())
        self.selected_columns = [c for c in self.selected_columns if c not in picks]
        self.refresh_lists()

    def move_selected(self, direction):
        """Shift the highlighted item(s) in the selected list up (-1) or down (+1)."""
        indices = list(self.selected_listbox.curselection())
        if not indices:
            return
        if direction < 0:
            indices = sorted(indices)
        else:
            indices = sorted(indices, reverse=True)
        for idx in indices:
            new_idx = idx + direction
            if 0 <= new_idx < len(self.selected_columns):
                self.selected_columns[idx], self.selected_columns[new_idx] = \
                    self.selected_columns[new_idx], self.selected_columns[idx]
        self.refresh_lists()
        # Re-select the moved items at their new positions
        self.selected_listbox.selection_clear(0, tk.END)
        for idx in indices:
            new_idx = max(0, min(len(self.selected_columns) - 1, idx + direction))
            self.selected_listbox.selection_set(new_idx)

    def refresh_lists(self):
        """Redraw both listboxes from self.selected_columns / self.available_columns and update header highlights."""
        self.selected_listbox.delete(0, tk.END)
        for col in self.selected_columns:
            self.selected_listbox.insert(tk.END, col)

        self.available_listbox.delete(0, tk.END)
        for col in self.available_columns:
            if col not in self.selected_columns:
                self.available_listbox.insert(tk.END, col)

        count = len(self.selected_columns)
        self.status_var.set(f"{count} column(s) selected" + ("" if count >= 2 else " - pick at least two columns to combine"))
        self.highlight_selected_headers()
        self.preview_result()

    def _resolve_separator(self):
        sep = self.separator_var.get()
        return "" if sep == "(none)" else sep

    def preview_result(self):
        """Show a small sample of what the concatenated column will look like."""
        self.preview_text.config(state=tk.NORMAL)
        self.preview_text.delete(1.0, tk.END)

        if len(self.selected_columns) < 2:
            self.preview_text.insert(tk.END, "Select at least two columns to see a preview.")
            self.preview_text.config(state=tk.DISABLED)
            return

        separator = self._resolve_separator()
        new_name = self.new_column_name_var.get().strip() or "Concatenated"

        sample = self.df.head(10)
        lines = [f"{new_name}", "-" * 40]
        for _, row in sample.iterrows():
            parts = []
            for col in self.selected_columns:
                val = row[col]
                parts.append("" if pd.isna(val) else str(val))
            lines.append(separator.join(parts))

        self.preview_text.insert(tk.END, "\n".join(lines))
        self.preview_text.config(state=tk.DISABLED)
        self.status_var.set(f"Previewing '{new_name}' from {len(self.selected_columns)} column(s)")

    def apply_concatenate(self):
        """Validate the selection and hand the configuration back to the caller."""
        if len(self.selected_columns) < 2:
            messagebox.showwarning("Warning", "Please select at least two columns to concatenate.")
            return

        new_name = self.new_column_name_var.get().strip()
        if not new_name:
            messagebox.showwarning("Warning", "Please enter a name for the new column.")
            return

        drop_originals = self.drop_originals_var.get()
        # Columns that will still exist afterward (everything except dropped originals)
        surviving_columns = [c for c in self.available_columns if not (drop_originals and c in self.selected_columns)]
        if new_name in surviving_columns and new_name not in self.selected_columns:
            messagebox.showerror("Error", f"A column named '{new_name}' already exists. Please choose a different name.")
            return

        self.result = {
            'columns': list(self.selected_columns),
            'separator': self._resolve_separator(),
            'new_column_name': new_name,
            'drop_originals': drop_originals
        }
        self.dialog.destroy()


class SendVarianceDialog(tk.Toplevel):
    """Dialog for composing and sending shipment variance emails to site managers via Outlook."""
    def __init__(self, parent, app=None, initial_attachment=None):
        super().__init__(parent)
        self.app = app
        self.title("Send Variance Report")
        self.transient(parent)
        self.grab_set()
        self.resizable(True, True)
        self.minsize(680, 620)

        # Palette
        palette = self.app.DARK_PALETTE if (self.app and self.app.dark_mode.get()) else (
            self.app.LIGHT_PALETTE if self.app else {
                'bg': '#faf8f2', 'panel_bg': '#ffffff', 'fg': '#0f172a',
                'entry_bg': '#ffffff', 'entry_fg': '#0f172a', 'button_fg': '#ffffff'
            }
        )
        self.palette = palette
        self.configure(bg=palette['bg'])

        # Separate Excel sheet for variance mail site managers template
        if getattr(sys, "frozen", False):
            base_dir = os.path.dirname(os.path.abspath(sys.executable))
        else:
            base_dir = os.path.dirname(os.path.abspath(__file__))
        self.templates_path = os.path.join(base_dir, "variance_templates.xlsx")

        self.outlook_mod = None
        self.template_store = None
        self.outlook_client = None
        try:
            import outlook_email_gui
            self.outlook_mod = outlook_email_gui
            self.template_store = outlook_email_gui.TemplateStore(self.templates_path)
            # Create variance_templates.xlsx with headers if it doesn't exist yet
            if not os.path.exists(self.templates_path):
                self.template_store.save()
            self.outlook_client = outlook_email_gui.OutlookClient()
        except Exception:
            pass

        self.attachment_paths = []
        if initial_attachment and os.path.isfile(initial_attachment):
            self.attachment_paths.append(initial_attachment)

        self.template_var = tk.StringVar(value="-- Select Client / Site Manager --")
        self.include_signature_var = tk.BooleanVar(value=True)

        self.create_widgets()
        self._size_and_center(parent)

    def _size_and_center(self, parent):
        self.update_idletasks()
        parent.update_idletasks()
        pw = parent.winfo_width() or 1000
        ph = parent.winfo_height() or 800
        px = parent.winfo_rootx()
        py = parent.winfo_rooty()
        w = max(680, min(800, int(pw * 0.75)))
        h = max(620, min(750, int(ph * 0.85)))
        x = max(0, px + (pw - w) // 2)
        y = max(0, py + (ph - h) // 2)
        self.geometry(f"{w}x{h}+{x}+{y}")

    def create_widgets(self):
        container = ttk.Frame(self, padding="16")
        container.pack(fill=tk.BOTH, expand=True)

        # Header
        header = ttk.Frame(container)
        header.pack(fill=tk.X, pady=(0, 10))
        ttk.Label(header, text="📧 Send Variance Report", font=('Segoe UI', 13, 'bold')).pack(anchor=tk.W)
        ttk.Label(
            header,
            text="Compose and email the shipment variance report to the client's site manager via Outlook.",
            font=('Segoe UI', 9), foreground='gray'
        ).pack(anchor=tk.W, pady=(2, 0))

        # --- Template selection ---
        tpl_frame = ttk.LabelFrame(container, text="Site Manager Template", padding="10")
        tpl_frame.pack(fill=tk.X, pady=(0, 10))

        tpl_row = ttk.Frame(tpl_frame)
        tpl_row.pack(fill=tk.X)
        ttk.Label(tpl_row, text="Client / Preset:", font=('Segoe UI', 9, 'bold'), width=14).pack(side=tk.LEFT)

        tpl_names = ["-- Select Client / Site Manager --"]
        if self.template_store:
            tpl_names += self.template_store.get_names()

        self.template_combo = ttk.Combobox(
            tpl_row, textvariable=self.template_var, values=tpl_names, state="readonly"
        )
        self.template_combo.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8))
        self.template_combo.bind("<<ComboboxSelected>>", self._on_template_selected)

        ttk.Button(
            tpl_row, text="+ Add New Preset", command=self._open_add_template_dialog
        ).pack(side=tk.RIGHT)

        # --- Recipients & Subject ---
        fields_frame = ttk.LabelFrame(container, text="Recipients & Subject", padding="10")
        fields_frame.pack(fill=tk.X, pady=(0, 10))

        to_row = ttk.Frame(fields_frame)
        to_row.pack(fill=tk.X, pady=3)
        ttk.Label(to_row, text="To:", font=('Segoe UI', 9, 'bold'), width=12).pack(side=tk.LEFT)
        self.to_entry = ttk.Entry(to_row)
        self.to_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)

        cc_row = ttk.Frame(fields_frame)
        cc_row.pack(fill=tk.X, pady=3)
        ttk.Label(cc_row, text="CC:", font=('Segoe UI', 9, 'bold'), width=12).pack(side=tk.LEFT)
        self.cc_entry = ttk.Entry(cc_row)
        self.cc_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)

        subj_row = ttk.Frame(fields_frame)
        subj_row.pack(fill=tk.X, pady=3)
        ttk.Label(subj_row, text="Subject:", font=('Segoe UI', 9, 'bold'), width=12).pack(side=tk.LEFT)
        self.subject_entry = ttk.Entry(subj_row)
        self.subject_entry.insert(0, "Shipment Variance Report")
        self.subject_entry.pack(side=tk.LEFT, fill=tk.X, expand=True)

        # --- Message Body ---
        body_frame = ttk.LabelFrame(container, text="Message Body", padding="10")
        body_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

        # Toolbar above message box
        msg_tools = ttk.Frame(body_frame)
        msg_tools.pack(fill=tk.X, pady=(0, 6))

        ttk.Label(msg_tools, text="Message Text & Table:", font=('Segoe UI', 9, 'bold')).pack(side=tk.LEFT)

        ttk.Button(
            msg_tools,
            text="⚡ Insert the Variance",
            command=self._insert_variance_from_report
        ).pack(side=tk.RIGHT, padx=(4, 0))

        text_scroll_frame = ttk.Frame(body_frame)
        text_scroll_frame.pack(fill=tk.BOTH, expand=True)

        self.body_text = tk.Text(
            text_scroll_frame, font=('Consolas', 9), wrap=tk.NONE, height=6,
            bg=self.palette['entry_bg'], fg=self.palette['entry_fg'],
            insertbackground=self.palette['entry_fg']
        )
        self.body_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        body_scrollbar = ttk.Scrollbar(text_scroll_frame, orient="vertical", command=self.body_text.yview)
        body_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        body_hscroll = ttk.Scrollbar(body_frame, orient="horizontal", command=self.body_text.xview)
        body_hscroll.pack(fill=tk.X)
        self.body_text.config(yscrollcommand=body_scrollbar.set, xscrollcommand=body_hscroll.set)

        if self.outlook_mod and hasattr(self.outlook_mod, "setup_text_tags"):
            self.outlook_mod.setup_text_tags(self.body_text)

        def _on_paste(event=None):
            self._paste_excel_table()
            return "break"

        self.body_text.bind("<Control-v>", _on_paste)
        self.body_text.bind("<Control-V>", _on_paste)

        default_body = (
            "Dear ,\n\n"
            "Please check and advise the variance below.\n"
        )
        self.body_text.insert("1.0", default_body)

        sig_row = ttk.Frame(body_frame)
        sig_row.pack(fill=tk.X, pady=(6, 0))
        ttk.Checkbutton(
            sig_row, text="Include default Outlook signature",
            variable=self.include_signature_var
        ).pack(anchor=tk.W)

        # --- Attachments ---
        attach_frame = ttk.LabelFrame(container, text="📎 Attachments", padding="10")
        attach_frame.pack(fill=tk.X, pady=(0, 10))

        attach_list_frame = ttk.Frame(attach_frame)
        attach_list_frame.pack(fill=tk.X)

        self.attach_listbox = tk.Listbox(
            attach_list_frame, height=3, font=('Segoe UI', 9),
            bg=self.palette['entry_bg'], fg=self.palette['entry_fg']
        )
        self.attach_listbox.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8))

        attach_btn_frame = ttk.Frame(attach_list_frame)
        attach_btn_frame.pack(side=tk.RIGHT, fill=tk.Y)

        ttk.Button(
            attach_btn_frame, text="+ Attach Files...", command=self._attach_files
        ).pack(fill=tk.X, pady=1)
        ttk.Button(
            attach_btn_frame, text="Remove", command=self._remove_attachment
        ).pack(fill=tk.X, pady=1)
        ttk.Button(
            attach_btn_frame, text="Clear All", command=self._clear_attachments
        ).pack(fill=tk.X, pady=1)

        self._refresh_attachment_list()

        # --- Bottom Buttons ---
        btn_bar = ttk.Frame(container)
        btn_bar.pack(fill=tk.X, pady=(4, 0))

        ttk.Button(
            btn_bar, text="👁 Preview in Outlook", command=self._preview_email
        ).pack(side=tk.LEFT, padx=(0, 8))

        ttk.Button(
            btn_bar, text="🚀 Send Email", command=self._send_email
        ).pack(side=tk.LEFT)

        ttk.Button(
            btn_bar, text="Cancel", command=self.destroy
        ).pack(side=tk.RIGHT)

    def _on_template_selected(self, event=None):
        name = self.template_var.get()
        if not name or name == "-- Select Client / Site Manager --":
            return
        if not self.template_store:
            return
        tpl = self.template_store.get_template(name)
        if tpl:
            self.to_entry.delete(0, tk.END)
            self.to_entry.insert(0, tpl.get("to", ""))
            self.cc_entry.delete(0, tk.END)
            self.cc_entry.insert(0, tpl.get("cc", ""))
            self.subject_entry.delete(0, tk.END)
            self.subject_entry.insert(0, f"{name} - Shipment Variance Report")

    def _setup_tags(self):
        self.body_text.tag_configure("tbl_header", background="#1c3f60", foreground="#ffffff", font=("Consolas", 9, "bold"))
        self.body_text.tag_configure("tbl_border", foreground="#64748b", font=("Consolas", 9))
        self.body_text.tag_configure("tbl_match", background="#d4edda", foreground="#155724", font=("Consolas", 9, "bold"))
        self.body_text.tag_configure("tbl_variance", background="#f8d7da", foreground="#721c24", font=("Consolas", 9, "bold"))
        self.body_text.tag_configure("tbl_missing", background="#fff3cd", foreground="#856404", font=("Consolas", 9, "bold"))
        self.body_text.tag_configure("tbl_cell", font=("Consolas", 9))

    def _insert_grid_table(self, grid):
        max_cols = max(len(row) for row in grid)
        if max_cols < 1:
            return

        norm_grid = [r + [""] * (max_cols - len(r)) for r in grid]
        col_widths = [max(len(str(row[c]).strip()) for row in norm_grid) for c in range(max_cols)]
        col_widths = [max(w, 8) for w in col_widths]

        def make_border(left, mid, right, fill="─"):
            return left + mid.join(fill * (w + 2) for w in col_widths) + right

        top_border = make_border("┌", "┬", "┐")
        mid_border = make_border("├", "┼", "┤")
        bot_border = make_border("└", "┴", "┘")

        self._setup_tags()

        self.body_text.insert(tk.INSERT, "\n")
        self.body_text.insert(tk.INSERT, top_border + "\n", "tbl_border")

        for r_idx, row in enumerate(norm_grid):
            is_header = (r_idx == 0)
            self.body_text.insert(tk.INSERT, "│", "tbl_border")
            for c_idx, cell in enumerate(row):
                val = str(cell).strip()
                cell_padded = f" {val.ljust(col_widths[c_idx])} "
                if is_header:
                    tag = "tbl_header"
                elif "MATCH" in val.upper() or val == "Exact match":
                    tag = "tbl_match"
                elif "VARIANCE" in val.upper():
                    tag = "tbl_variance"
                elif "MISSING" in val.upper() or "EXTRA" in val.upper():
                    tag = "tbl_missing"
                else:
                    tag = "tbl_cell"

                self.body_text.insert(tk.INSERT, cell_padded, tag)
                self.body_text.insert(tk.INSERT, "│", "tbl_border")
            self.body_text.insert(tk.INSERT, "\n")

            if is_header and len(norm_grid) > 1:
                self.body_text.insert(tk.INSERT, mid_border + "\n", "tbl_border")

        self.body_text.insert(tk.INSERT, bot_border + "\n\n", "tbl_border")

    def _paste_excel_table(self):
        """Pastes Excel table from clipboard with full borders, header styling, and cell coloring."""
        plain_text = ""
        try:
            import win32clipboard
            win32clipboard.OpenClipboard()
            if win32clipboard.IsClipboardFormatAvailable(win32clipboard.CF_UNICODETEXT):
                plain_text = win32clipboard.GetClipboardData(win32clipboard.CF_UNICODETEXT)
        except Exception:
            pass
        finally:
            try:
                import win32clipboard
                win32clipboard.CloseClipboard()
            except Exception:
                pass

        if not plain_text:
            try:
                plain_text = self.body_text.clipboard_get()
            except Exception:
                return

        if not plain_text:
            return

        lines = [l.rstrip("\r\n") for l in plain_text.strip().split("\n") if l.strip()]
        if not lines:
            return

        grid = [line.split("\t") for line in lines]
        max_cols = max(len(row) for row in grid)

        if max_cols >= 2 or (len(grid) > 1 and "\t" in plain_text):
            self._insert_grid_table(grid)
        else:
            self.body_text.insert(tk.INSERT, plain_text)

    def _insert_variance_from_report(self):
        """Loads and formats the entire variance sheet content directly from the reconciliation report without filtering."""
        excel_path = None
        for p in self.attachment_paths:
            if os.path.isfile(p) and p.lower().endswith((".xlsx", ".xls", ".xlsm", ".xlsb")):
                excel_path = p
                break

        if not excel_path:
            excel_path = filedialog.askopenfilename(
                title="Select Reconciliation Report with Variance Sheet",
                filetypes=[("Excel files", "*.xlsx;*.xls;*.xlsm;*.xlsb"), ("All files", "*.*")]
            )
            if not excel_path:
                return
            if excel_path not in self.attachment_paths:
                self.attachment_paths.append(excel_path)
                self._refresh_attachment_list()

        try:
            import pandas as pd
            xl = pd.ExcelFile(excel_path)
            target_sheet = None
            for s in xl.sheet_names:
                if "variance" in s.lower():
                    target_sheet = s
                    break
            if not target_sheet:
                target_sheet = xl.sheet_names[0]

            df = pd.read_excel(excel_path, sheet_name=target_sheet)
            df = df.fillna("")

            # Convert all rows and columns to grid table without filtering or truncation
            grid = [df.columns.tolist()] + df.astype(str).values.tolist()
            self._insert_grid_table(grid)
            messagebox.showinfo("Variance Inserted", f"Successfully inserted {len(grid)-1} rows from '{target_sheet}' sheet into message body.")
        except Exception as e:
            messagebox.showerror("Read Error", f"Could not read variance sheet from Excel file:\n\n{e}")

    def _open_add_template_dialog(self):
        if not self.outlook_mod or not getattr(self.outlook_mod, "OPENPYXL_AVAILABLE", False):
            messagebox.showerror(
                "openpyxl required",
                "openpyxl is required to save recipient templates.\nPlease run: pip install openpyxl"
            )
            return

        def submit(name, to_addr, cc_addr):
            self.template_store.add_or_update_template(name, to_addr, cc_addr)
            names = ["-- Select Client / Site Manager --"] + self.template_store.get_names()
            self.template_combo.configure(values=names)
            self.template_var.set(name)
            self._on_template_selected()

        self.outlook_mod.AddTemplateDialog(self, on_submit=submit)

    def _attach_files(self):
        filetypes = [
            ("Excel and PDF files", "*.xlsx;*.xls;*.pdf"),
            ("Excel files", "*.xlsx;*.xls"),
            ("PDF files", "*.pdf"),
            ("All files", "*.*")
        ]
        paths = filedialog.askopenfilenames(title="Select Files to Attach", filetypes=filetypes)
        if paths:
            for p in paths:
                if p not in self.attachment_paths:
                    self.attachment_paths.append(p)
            self._refresh_attachment_list()

    def _remove_attachment(self):
        sel = list(self.attach_listbox.curselection())
        if sel:
            for idx in reversed(sel):
                del self.attachment_paths[idx]
            self._refresh_attachment_list()

    def _clear_attachments(self):
        self.attachment_paths.clear()
        self._refresh_attachment_list()

    def _refresh_attachment_list(self):
        self.attach_listbox.delete(0, tk.END)
        for p in self.attachment_paths:
            self.attach_listbox.insert(tk.END, f"📄 {os.path.basename(p)}")

    def _convert_ascii_table_to_html(self, table_lines):
        data_rows = []
        for line in table_lines:
            line_str = line.strip()
            if not line_str or line_str.startswith("┌") or line_str.startswith("├") or line_str.startswith("└") or line_str.startswith("+"):
                continue
            if "│" in line_str:
                cells = [c.strip() for c in line_str.split("│")[1:-1]]
                if any(cells):
                    data_rows.append(cells)

        if not data_rows:
            return ""

        max_cols = max(len(r) for r in data_rows)
        html_rows = []

        for r_idx, row in enumerate(data_rows):
            is_header = (r_idx == 0)
            td_tags = []
            padded_row = row + [""] * (max_cols - len(row))
            for c_idx, cell in enumerate(padded_row):
                val = cell.strip()
                if is_header:
                    th_style = (
                        "background-color: #1c3f60; color: #ffffff; font-weight: bold; "
                        "padding: 8px 14px; border: 1px solid #0d1826; text-align: left; font-size: 10pt;"
                    )
                    td_tags.append(f'<th style="{th_style}">{html.escape(val)}</th>')
                else:
                    val_upper = val.upper()
                    bg_color = "#ffffff"
                    fg_color = "#1f2937"
                    font_weight = "normal"

                    if "MATCH" in val_upper or val == "Exact match" or "EXACT MATCH" in val_upper:
                        bg_color = "#d4edda"
                        fg_color = "#155724"
                        font_weight = "bold"
                    elif "MAJOR VARIANCE" in val_upper or "LARGE VARIANCE" in val_upper:
                        bg_color = "#f8d7da"
                        fg_color = "#721c24"
                        font_weight = "bold"
                    elif "VARIANCE" in val_upper or "MODERATE VARIANCE" in val_upper or "MINOR VARIANCE" in val_upper:
                        bg_color = "#fff3cd"
                        fg_color = "#856404"
                        font_weight = "bold"
                    elif "EXTRA" in val_upper or "MISSING" in val_upper:
                        bg_color = "#e2e3e5"
                        fg_color = "#383d41"
                        font_weight = "bold"
                    elif r_idx % 2 == 1:
                        bg_color = "#f8fafc"

                    td_style = (
                        f"background-color: {bg_color}; color: {fg_color}; font-weight: {font_weight}; "
                        f"padding: 6px 12px; border: 1px solid #cbd5e1; font-size: 10pt;"
                    )
                    td_tags.append(f'<td style="{td_style}">{html.escape(val)}</td>')

            html_rows.append("<tr>" + "".join(td_tags) + "</tr>")

        table_style = (
            "border-collapse: collapse; width: auto; margin: 12px 0; "
            "font-family: Calibri, Segoe UI, Arial, sans-serif; "
            "border: 1px solid #0d1826;"
        )
        return f'<table border="1" cellpadding="0" cellspacing="0" style="{table_style}">{"".join(html_rows)}</table>'

    def _compose_variance_html(self, body_text, signature_html):
        lines = body_text.split("\n")
        html_parts = []
        in_table = False
        table_lines = []

        for line in lines:
            is_tbl_line = (
                line.startswith("┌")
                or line.startswith("+---")
                or ("│" in line and (line.startswith("│") or line.endswith("│")))
                or line.startswith("├")
                or line.startswith("└")
                or line.startswith("+")
            )

            if is_tbl_line:
                if not in_table:
                    in_table = True
                    table_lines = []
                table_lines.append(line)
            else:
                if in_table:
                    html_table = self._convert_ascii_table_to_html(table_lines)
                    html_parts.append(html_table)
                    in_table = False
                    table_lines = []

                if line.strip():
                    html_parts.append(f'<div style="margin: 4px 0;">{html.escape(line)}</div>')
                else:
                    html_parts.append('<div style="height: 12px;"></div>')

        if in_table:
            html_table = self._convert_ascii_table_to_html(table_lines)
            html_parts.append(html_table)

        message_html = (
            '<div style="font-family:Calibri,Arial,sans-serif;'
            'font-size:11pt;color:#000000;">'
            + "".join(html_parts)
            + "</div><br>"
        )

        if signature_html:
            match = re.search(r"(<body[^>]*>)", signature_html, re.IGNORECASE)
            if match:
                insertion_point = match.end()
                return (signature_html[:insertion_point] + message_html
                         + signature_html[insertion_point:])
            return message_html + signature_html
        return f"<html><body>{message_html}</body></html>"

    def _gather_and_validate(self):
        if not self.outlook_client:
            messagebox.showerror(
                "Outlook Error",
                "pywin32 is not installed or Outlook is unavailable.\n\nInstall with: pip install pywin32"
            )
            return None

        to_addr = (self.to_entry.get() or "").strip()
        cc_addr = (self.cc_entry.get() or "").strip()
        subject = (self.subject_entry.get() or "").strip()
        body = self.body_text.get("1.0", "end-1c")

        if not to_addr:
            messagebox.showwarning("Missing Recipient", "Please enter at least one 'To:' recipient email.")
            return None

        if not subject:
            if not messagebox.askyesno("Empty Subject", "The subject line is empty. Continue anyway?"):
                return None

        return to_addr, cc_addr, subject, body

    def _preview_email(self):
        data = self._gather_and_validate()
        if not data:
            return
        to_addr, cc_addr, subject, body = data
        try:
            signature_html = ""
            if self.include_signature_var.get() and self.outlook_mod and hasattr(self.outlook_mod, "get_default_signature_html"):
                try:
                    signature_html = self.outlook_mod.get_default_signature_html()
                except Exception:
                    pass

            app = self.outlook_client._get_app()
            mail = app.CreateItem(0)
            mail.To = to_addr
            if cc_addr:
                mail.CC = cc_addr
            mail.Subject = subject
            mail.HTMLBody = self._compose_variance_html(body, signature_html)

            for path in self.attachment_paths:
                if os.path.isfile(path):
                    mail.Attachments.Add(path)

            mail.Display(True)
        except Exception as e:
            messagebox.showerror("Outlook Error", f"Could not preview email in Outlook:\n\n{e}")

    def _send_email(self):
        data = self._gather_and_validate()
        if not data:
            return
        to_addr, cc_addr, subject, body = data

        attach_names = "\n".join(f"  • {os.path.basename(p)}" for p in self.attachment_paths) or "  (none)"
        confirm = (
            f"To: {to_addr}\n"
            f"CC: {cc_addr or '(none)'}\n"
            f"Subject: {subject}\n"
            f"Attachments:\n{attach_names}\n\n"
            "Send this variance email via Outlook now?"
        )
        if not messagebox.askyesno("Confirm Send", confirm):
            return

        try:
            signature_html = ""
            if self.include_signature_var.get() and self.outlook_mod and hasattr(self.outlook_mod, "get_default_signature_html"):
                try:
                    signature_html = self.outlook_mod.get_default_signature_html()
                except Exception:
                    pass

            app = self.outlook_client._get_app()
            mail = app.CreateItem(0)
            mail.To = to_addr
            if cc_addr:
                mail.CC = cc_addr
            mail.Subject = subject
            mail.HTMLBody = self._compose_variance_html(body, signature_html)

            for path in self.attachment_paths:
                if os.path.isfile(path):
                    mail.Attachments.Add(path)

            mail.Send()
            messagebox.showinfo("Email Sent", "Variance email was successfully sent via Outlook.")
            if self.app:
                self.app.add_log(f"📧 Variance email sent to: {to_addr} (Subject: {subject})", "SUCCESS")
            self.destroy()
        except Exception as e:
            messagebox.showerror("Outlook Error", f"Could not send email:\n\n{e}")


class ReconciliationApp:
    SETTINGS_FILE = os.path.join(os.path.expanduser("~"), ".load_reconciliation_tool_settings.json")

    def __init__(self, root, container=None, standalone=True):
        """root: the Tk instance (used for dialogs / winfo_* / etc).
        container: the frame this tool's UI should be built into. Defaults
        to root itself, which reproduces the original standalone behavior.
        standalone: when False (i.e. embedded in the multi-tool shell),
        window-level chrome (title, size/position, close handling) is left
        alone since the shell owns it, and the in-page Settings button is
        hidden since Settings now lives in the shell's sidebar."""
        self.root = root
        self.embedded = not standalone
        self.theme_target = container if container is not None else root
        if standalone:
            self.root.title("Load Reconciliation v2.1")
            self.root.minsize(1200, 750)
        
        # Load any previously saved settings before building the UI so the
        # window size/theme can be applied from the very first frame.
        self.settings_data = self.load_settings()
        self.remember_settings = tk.BooleanVar(value=self.settings_data.get('remember_settings', False))
        self.dark_mode = tk.BooleanVar(value=self.settings_data.get('dark_mode', False))
        
        if standalone:
            if self.remember_settings.get() and self.settings_data.get('window_geometry'):
                try:
                    self.root.geometry(self.settings_data['window_geometry'])
                except tk.TclError:
                    self.root.geometry("1500x950")
                if self.settings_data.get('window_state') == 'zoomed':
                    try:
                        self.root.state('zoomed')
                    except tk.TclError:
                        pass
            else:
                self.root.geometry("1500x950")
                try:
                    self.root.state('zoomed')  # start maximized on Windows so nothing is hidden below the fold
                except tk.TclError:
                    pass
        
        # Set colors
        self.colors = {
            'primary': '#0d1b2a',
            'secondary': '#00b7ff',
            'success': '#10b981',
            'danger': '#ef4444',
            'warning': '#f59e0b',
            'bg': '#faf8f2',
            'white': '#ffffff',
            'info': '#00e5ff'
        }
        
        # Light / dark palettes used by apply_theme()
        self.LIGHT_PALETTE = {
            'bg': '#faf8f2', 'panel_bg': '#ffffff', 'fg': '#0f172a', 'subtle_fg': '#64748b',
            'entry_bg': '#ffffff', 'entry_fg': '#0f172a',
            'tree_bg': '#ffffff', 'tree_fg': '#0f172a', 'tree_heading_bg': '#e2e8f0',
            'oddrow': '#f8fafc', 'evenrow': '#ffffff',
            'title_fg': '#0f172a', 'button_fg': '#ffffff',
        }
        self.DARK_PALETTE = {
            'bg': '#0b1420', 'panel_bg': '#13233a', 'fg': '#ececec', 'subtle_fg': '#94a3b8',
            'entry_bg': '#162e4c', 'entry_fg': '#ececec',
            'tree_bg': '#13233a', 'tree_fg': '#ececec', 'tree_heading_bg': '#162e4c',
            'oddrow': '#0f1b2d', 'evenrow': '#13233a',
            'title_fg': '#00e5ff', 'button_fg': '#ffffff',
        }
        
        # Variables
        self.history_file_path = tk.StringVar()
        self.plan_file_path = tk.StringVar()
        self.output_folder = tk.StringVar(value=os.path.join(os.path.expanduser("~"), "Desktop"))
        self.archive_folder = tk.StringVar(value="C:\\Reconciliation\\Archive")
        self.progress_var = tk.DoubleVar()
        self.status_var = tk.StringVar(value="Ready")
        self.history_file_info = tk.StringVar(value="No file selected")
        self.plan_file_info = tk.StringVar(value="No file selected")
        self.error_message = tk.StringVar(value="")
        
        # Column selection variables
        self.history_load_id_col = tk.StringVar(value="")
        self.history_qty_col = tk.StringVar(value="")
        self.plan_load_id_col = tk.StringVar(value="")
        self.plan_qty_col = tk.StringVar(value="")
        
        # Divide Plan Qty option
        self.plan_divide_enabled = tk.BooleanVar(value=False)
        self.plan_divide_value = tk.StringVar(value="1")
        
        # Store column lists
        self.history_columns = []
        self.plan_columns = []
        
        # Store original data for text-to-columns
        self.history_original_data = None
        self.plan_original_data = None
        self.history_df_loaded = None
        self.plan_df_loaded = None
        
        # Store concatenate-columns configuration so it can be reapplied to
        # the freshly re-read file when reconciliation actually runs.
        # Lists so multiple concatenations (done one after another) are all replayed.
        self.history_concat_configs = []
        self.plan_concat_configs = []
        self.last_output_file = None
        
        # Additional options
        self.auto_open_file = tk.BooleanVar(value=True)
        self.verbose_logging = tk.BooleanVar(value=False)
        self.archive_var = tk.BooleanVar(value=True)
        
        # Text to columns options
        self.history_split_enabled = tk.BooleanVar(value=False)
        self.plan_split_enabled = tk.BooleanVar(value=False)
        self.history_split_column = tk.StringVar(value="")
        self.plan_split_column = tk.StringVar(value="")
        self.history_split_widths = ""
        self.plan_split_widths = ""
        
        # Log storage
        self.log_messages = []
        
        # Preview panes
        self.history_tree = None
        self.plan_tree = None
        self.history_preview_count = tk.StringVar(value="No data loaded")
        self.plan_preview_count = tk.StringVar(value="No data loaded")
        self.preview_status_var = tk.StringVar(value="👆 Click a row in either preview below to see its values here")
        self._history_marker_imgs = {}
        self._plan_marker_imgs = {}
        self.PREVIEW_MAX_ROWS = 200
        
        # Apply remembered folder/option settings now that their variables exist
        if self.remember_settings.get():
            self.output_folder.set(self.settings_data.get('output_folder', self.output_folder.get()))
            self.archive_folder.set(self.settings_data.get('archive_folder', self.archive_folder.get()))
            self.archive_var.set(self.settings_data.get('archive_after_processing', self.archive_var.get()))
            self.auto_open_file.set(self.settings_data.get('auto_open_file', self.auto_open_file.get()))
            self.verbose_logging.set(self.settings_data.get('verbose_logging', self.verbose_logging.get()))
        
        # Create UI
        self.create_widgets()
        
        # Apply the saved (or default) theme now that all widgets exist
        self.apply_theme()
        
        # Save settings on close (X button) as well as via the Exit button.
        # When embedded in the shell, the shell owns the window-close
        # protocol, so this tool just saves settings when the shell exits
        # (see the shell's own on_close handler) rather than here.
        if standalone:
            self.root.protocol("WM_DELETE_WINDOW", self.on_app_close)
        
        # Re-highlight preview columns whenever a column selection changes
        self.history_load_id_col.trace_add('write', lambda *a: self.highlight_preview_columns('history'))
        self.history_qty_col.trace_add('write', lambda *a: self.highlight_preview_columns('history'))
        self.plan_load_id_col.trace_add('write', lambda *a: self.highlight_preview_columns('plan'))
        self.plan_qty_col.trace_add('write', lambda *a: self.highlight_preview_columns('plan'))
    
    # ===================== SETTINGS PERSISTENCE =====================
    
    def load_settings(self):
        """Load saved settings from disk, if any. Never raises."""
        try:
            if os.path.exists(self.SETTINGS_FILE):
                with open(self.SETTINGS_FILE, 'r') as f:
                    return json.load(f)
        except Exception:
            pass
        return {}
    
    def save_settings(self):
        """Persist current settings to disk. Theme + the remember flag itself
        are always saved; window size/folders/options are only saved when
        'Remember my last used settings' is enabled, so turning it back on
        later restores whatever was last used."""
        data = {
            'remember_settings': self.remember_settings.get(),
            'dark_mode': self.dark_mode.get(),
        }
        try:
            data['window_geometry'] = self.root.geometry()
            data['window_state'] = self.root.state()
        except tk.TclError:
            pass
        data['output_folder'] = self.output_folder.get()
        data['archive_folder'] = self.archive_folder.get()
        data['archive_after_processing'] = self.archive_var.get()
        data['auto_open_file'] = self.auto_open_file.get()
        data['verbose_logging'] = self.verbose_logging.get()
        
        try:
            with open(self.SETTINGS_FILE, 'w') as f:
                json.dump(data, f, indent=2)
        except Exception:
            pass
    
    def on_app_close(self):
        """Save settings and close the app (used by the X button and Exit).
        When embedded in the multi-tool shell, "closing" this tool must not
        destroy the shared root window -- it just saves settings and hands
        control back to the shell (e.g. returns to the Dashboard) instead."""
        self.save_settings()
        if self.embedded:
            if hasattr(self, "on_embedded_exit") and callable(self.on_embedded_exit):
                self.on_embedded_exit()
        else:
            self.root.destroy()
    
    def open_settings_dialog(self):
        """Open the ⚙️ Settings popup with the remember/theme switches."""
        dialog = tk.Toplevel(self.root)
        dialog.title("Settings")
        dialog.transient(self.root)
        dialog.resizable(False, False)
        dialog.grab_set()
        
        palette = self.DARK_PALETTE if self.dark_mode.get() else self.LIGHT_PALETTE
        dialog.configure(bg=palette['bg'])
        
        container = ttk.Frame(dialog, padding="20")
        container.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(container, text="⚙️ Settings", font=('Segoe UI', 14, 'bold')).pack(anchor=tk.W, pady=(0, 15))
        
        remember_row = ttk.Frame(container)
        remember_row.pack(fill=tk.X, pady=6)
        ttk.Checkbutton(
            remember_row,
            text="💾 Remember my last used settings (window size, folders, options)",
            variable=self.remember_settings,
            command=self.save_settings
        ).pack(anchor=tk.W)
        
        theme_row = ttk.Frame(container)
        theme_row.pack(fill=tk.X, pady=6)
        ttk.Checkbutton(
            theme_row,
            text="🌙 Dark theme",
            variable=self.dark_mode,
            command=self._on_theme_toggle
        ).pack(anchor=tk.W)
        
        ttk.Separator(container, orient=tk.HORIZONTAL).pack(fill=tk.X, pady=(15, 10))
        
        ttk.Button(container, text="Close", command=dialog.destroy).pack(anchor=tk.E)
        
        dialog.update_idletasks()
        rx, ry = self.root.winfo_rootx(), self.root.winfo_rooty()
        rw, rh = self.root.winfo_width(), self.root.winfo_height()
        dw, dh = dialog.winfo_reqwidth(), dialog.winfo_reqheight()
        dialog.geometry(f"+{rx + (rw - dw)//2}+{ry + (rh - dh)//2}")
    
    def open_send_variance_dialog(self):
        """Open the Send Variance dialog to email site managers via Outlook."""
        SendVarianceDialog(self.root, app=self, initial_attachment=self.last_output_file)
    
    def _on_theme_toggle(self):
        self.apply_theme()
        self.save_settings()
    
    def apply_theme(self):
        """Recolor the app for the current dark_mode setting. Covers ttk
        styles globally plus the handful of plain tk widgets that don't
        follow ttk styling."""
        palette = self.DARK_PALETTE if self.dark_mode.get() else self.LIGHT_PALETTE
        
        style = ttk.Style()
        try:
            style.theme_use('clam')
        except tk.TclError:
            pass
        
        style.configure('TFrame', background=palette['bg'])
        style.configure('TLabelframe', background=palette['bg'], foreground=palette['fg'])
        style.configure('TLabelframe.Label', background=palette['bg'], foreground=palette['fg'])
        style.configure('TLabel', background=palette['bg'], foreground=palette['fg'])
        style.configure('TCheckbutton', background=palette['bg'], foreground=palette['fg'])
        style.map('TCheckbutton', background=[('active', palette['bg'])])
        style.configure('TButton', background=palette['panel_bg'], foreground=palette['fg'])
        style.map('TButton', background=[('active', palette['tree_heading_bg'])])
        style.configure('TEntry', fieldbackground=palette['entry_bg'], foreground=palette['entry_fg'])
        style.configure('TCombobox', fieldbackground=palette['entry_bg'], foreground=palette['entry_fg'])
        style.map(
            'TCombobox',
            fieldbackground=[('readonly', palette['entry_bg'])],
            foreground=[('readonly', palette['entry_fg'])]
        )
        style.configure(
            'Treeview',
            background=palette['tree_bg'],
            fieldbackground=palette['tree_bg'],
            foreground=palette['tree_fg']
        )
        style.configure('Treeview.Heading', background=palette['tree_heading_bg'], foreground=palette['fg'])
        style.configure('TPanedwindow', background=palette['bg'])
        style.configure('TProgressbar', background=self.colors['secondary'])
        
        # Plain tk widgets that need manual recoloring
        try:
            self.theme_target.configure(bg=palette['bg'])
        except tk.TclError:
            pass  # ttk containers (e.g. a shell content frame) don't take bg
        if hasattr(self, 'canvas'):
            self.canvas.configure(bg=palette['bg'])
        if hasattr(self, 'title_label'):
            self.title_label.configure(bg='#0b1420', fg='#ffffff')
        if hasattr(self, 'subtitle_label'):
            self.subtitle_label.configure(bg='#0b1420', fg='#94a3b8')
        if hasattr(self, 'footer_label'):
            self.footer_label.configure(bg=palette['bg'], fg=palette['subtle_fg'])
        if hasattr(self, 'error_label'):
            self.error_label.configure(bg=palette['bg'])
        if hasattr(self, 'run_button'):
            self.run_button.configure(bg=self.colors['secondary'], fg=palette['button_fg'])
        
        # Re-tag preview row striping for the new palette
        for tree in (self.history_tree, self.plan_tree):
            if tree is not None:
                tree.tag_configure('oddrow', background=palette['oddrow'])
                tree.tag_configure('evenrow', background=palette['evenrow'])
    
    def create_widgets(self):
        # --- Header banner (matches Korber Automation & Outlook Email Sender) ---
        HEADER_BG = "#0b1420"
        HEADER_BTN_BG = "#162e4c"
        HEADER_BTN_BG_HOVER = "#1f3f66"

        header = tk.Frame(self.theme_target, bg=HEADER_BG)
        header.pack(fill="x", side="top")

        header_top_row = tk.Frame(header, bg=HEADER_BG)
        header_top_row.pack(fill="x", padx=20, pady=(16, 0))

        self.title_label = tk.Label(
            header_top_row,
            text="Load Reconciliation",
            font=('Segoe UI', 15, 'bold'),
            bg=HEADER_BG,
            fg="#ffffff"
        )
        self.title_label.pack(side="left")

        def _make_header_button(parent, text, command):
            btn = tk.Label(
                parent, text=text, bg=HEADER_BTN_BG, fg="#ffffff",
                font=("Segoe UI", 8, "bold"), cursor="hand2",
                padx=8, pady=4,
            )
            btn.bind("<Button-1>", lambda e: command())
            btn.bind("<Enter>", lambda e: btn.config(bg=HEADER_BTN_BG_HOVER))
            btn.bind("<Leave>", lambda e: btn.config(bg=HEADER_BTN_BG))
            return btn

        # Header action buttons (top right)
        if not self.embedded:
            self.settings_button = _make_header_button(
                header_top_row, "⚙ Settings", self.open_settings_dialog
            )
            self.settings_button.pack(side="right")

        self.send_variance_button = _make_header_button(
            header_top_row, "📧 Send Variance", self.open_send_variance_dialog
        )
        self.send_variance_button.pack(
            side="right", padx=(0, 6) if not self.embedded else 0
        )

        self.subtitle_label = tk.Label(
            header,
            text="Compare Actual Dispatches Against Planned Loads",
            font=('Segoe UI', 9),
            bg=HEADER_BG,
            fg="#94a3b8"
        )
        self.subtitle_label.pack(anchor="w", padx=20, pady=(0, 14))

        # Main container
        main_container = ttk.Frame(self.theme_target)
        main_container.pack(fill=tk.BOTH, expand=True)
        
        # Canvas for scrolling
        self.canvas = tk.Canvas(main_container, highlightthickness=0)
        scrollbar = ttk.Scrollbar(main_container, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=scrollbar.set)
        
        # Main frame inside canvas
        main_frame = ttk.Frame(self.canvas, padding="20")
        main_frame.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        
        self.canvas_window = self.canvas.create_window((0, 0), window=main_frame, anchor="nw")
        self.canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Keep main_frame's width matched to the visible canvas width so the
        # right column can actually expand into the blank space on resize.
        self.canvas.bind(
            "<Configure>",
            lambda e: self.canvas.itemconfig(self.canvas_window, width=e.width)
        )
        
        # ===== TWO-COLUMN LAYOUT =====
        # Left column: setup (files, output, options, progress, run controls)
        # Right column: data preview + activity log, using the space that used
        # to sit blank while everything was stacked in one column.
        left_column = ttk.Frame(main_frame)
        left_column.grid(row=0, column=0, sticky="nsew", padx=(0, 15))
        
        right_column = ttk.Frame(main_frame)
        right_column.grid(row=0, column=1, sticky="nsew")
        
        main_frame.columnconfigure(0, weight=0, minsize=520)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(0, weight=1)
        
        # ===== FILE SELECTION FRAME =====
        file_frame = ttk.LabelFrame(left_column, text="📁 Select Files", padding="15")
        file_frame.pack(fill=tk.X, pady=(0, 15))
        
        # Loading History File
        history_row = ttk.Frame(file_frame)
        history_row.pack(fill=tk.X, pady=5)
        
        ttk.Label(history_row, text="Loading History:", font=('Segoe UI', 10, 'bold')).pack(side=tk.LEFT, padx=(0, 10))
        
        self.history_entry = ttk.Entry(history_row, textvariable=self.history_file_path)
        self.history_entry.pack(side=tk.LEFT, padx=(0, 10), fill=tk.X, expand=True)
        
        ttk.Button(
            history_row,
            text="Browse",
            command=self.browse_history_file
        ).pack(side=tk.LEFT)
        
        # History file info
        history_info_row = ttk.Frame(file_frame)
        history_info_row.pack(fill=tk.X, pady=2, padx=(10, 0))
        ttk.Label(history_info_row, textvariable=self.history_file_info, font=('Segoe UI', 8), foreground='gray').pack(anchor=tk.W)
        
        # History column selection (stacked to fit the narrower left column)
        history_load_id_row = ttk.Frame(file_frame)
        history_load_id_row.pack(fill=tk.X, pady=2, padx=(10, 0))
        ttk.Label(history_load_id_row, text="🔑 Primary Key (Load ID):", font=('Segoe UI', 9, 'bold'), width=22).pack(side=tk.LEFT)
        self.history_load_id_combo = ttk.Combobox(
            history_load_id_row, 
            textvariable=self.history_load_id_col,
            state="readonly"
        )
        self.history_load_id_combo.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        history_qty_row = ttk.Frame(file_frame)
        history_qty_row.pack(fill=tk.X, pady=2, padx=(10, 0))
        ttk.Label(history_qty_row, text="🔢 Quantity Column:", font=('Segoe UI', 9), width=22).pack(side=tk.LEFT)
        self.history_qty_combo = ttk.Combobox(
            history_qty_row,
            textvariable=self.history_qty_col,
            state="readonly"
        )
        self.history_qty_combo.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # Concatenate Columns button
        history_concat_row = ttk.Frame(file_frame)
        history_concat_row.pack(fill=tk.X, pady=(4, 2), padx=(10, 0))
        self.history_concat_button = ttk.Button(
            history_concat_row,
            text="🔗 Concatenate Columns...",
            command=lambda: self.open_concatenate_dialog('history'),
            state='disabled'
        )
        self.history_concat_button.pack(side=tk.LEFT)
        
        # History Text-to-Columns options
        history_split_check_row = ttk.Frame(file_frame)
        history_split_check_row.pack(fill=tk.X, pady=(8, 2), padx=(10, 0))
        ttk.Checkbutton(
            history_split_check_row,
            text="Enable Text-to-Columns (Fixed Width)",
            variable=self.history_split_enabled,
            command=self.toggle_history_split
        ).pack(side=tk.LEFT)
        
        history_split_frame = ttk.Frame(file_frame)
        history_split_frame.pack(fill=tk.X, pady=2, padx=(10, 0))
        
        ttk.Label(history_split_frame, text="Column to Split:", width=15).pack(side=tk.LEFT)
        self.history_split_combo = ttk.Combobox(
            history_split_frame,
            textvariable=self.history_split_column,
            state="readonly"
        )
        self.history_split_combo.pack(side=tk.LEFT, padx=(0, 10), fill=tk.X, expand=True)
        self.history_split_combo['state'] = 'disabled'
        
        ttk.Button(
            history_split_frame,
            text="Configure Split",
            command=lambda: self.configure_split('history'),
            state='disabled'
        ).pack(side=tk.LEFT)
        
        # Load Plan File
        plan_row = ttk.Frame(file_frame)
        plan_row.pack(fill=tk.X, pady=5)
        
        ttk.Label(plan_row, text="Load Plan:", font=('Segoe UI', 10, 'bold')).pack(side=tk.LEFT, padx=(0, 10))
        
        self.plan_entry = ttk.Entry(plan_row, textvariable=self.plan_file_path)
        self.plan_entry.pack(side=tk.LEFT, padx=(0, 10), fill=tk.X, expand=True)
        
        ttk.Button(
            plan_row,
            text="Browse",
            command=self.browse_plan_file
        ).pack(side=tk.LEFT)
        
        # Plan file info
        plan_info_row = ttk.Frame(file_frame)
        plan_info_row.pack(fill=tk.X, pady=2, padx=(10, 0))
        ttk.Label(plan_info_row, textvariable=self.plan_file_info, font=('Segoe UI', 8), foreground='gray').pack(anchor=tk.W)
        
        # Plan column selection (stacked to fit the narrower left column)
        plan_load_id_row = ttk.Frame(file_frame)
        plan_load_id_row.pack(fill=tk.X, pady=2, padx=(10, 0))
        ttk.Label(plan_load_id_row, text="🔑 Primary Key (Load ID):", font=('Segoe UI', 9, 'bold'), width=22).pack(side=tk.LEFT)
        self.plan_load_id_combo = ttk.Combobox(
            plan_load_id_row,
            textvariable=self.plan_load_id_col,
            state="readonly"
        )
        self.plan_load_id_combo.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        plan_qty_row = ttk.Frame(file_frame)
        plan_qty_row.pack(fill=tk.X, pady=2, padx=(10, 0))
        ttk.Label(plan_qty_row, text="🔢 Quantity Column:", font=('Segoe UI', 9), width=22).pack(side=tk.LEFT)
        self.plan_qty_combo = ttk.Combobox(
            plan_qty_row,
            textvariable=self.plan_qty_col,
            state="readonly"
        )
        self.plan_qty_combo.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # Concatenate Columns button
        plan_concat_row = ttk.Frame(file_frame)
        plan_concat_row.pack(fill=tk.X, pady=(4, 2), padx=(10, 0))
        self.plan_concat_button = ttk.Button(
            plan_concat_row,
            text="🔗 Concatenate Columns...",
            command=lambda: self.open_concatenate_dialog('plan'),
            state='disabled'
        )
        self.plan_concat_button.pack(side=tk.LEFT)
        
        # Divide Plan Qty option
        plan_divide_row = ttk.Frame(file_frame)
        plan_divide_row.pack(fill=tk.X, pady=(8, 2), padx=(10, 0))
        ttk.Checkbutton(
            plan_divide_row,
            text="Divide Plan Qty by:",
            variable=self.plan_divide_enabled,
            command=self.toggle_plan_divide
        ).pack(side=tk.LEFT)
        self.plan_divide_entry = ttk.Entry(
            plan_divide_row,
            textvariable=self.plan_divide_value,
            width=10,
            state='disabled'
        )
        self.plan_divide_entry.pack(side=tk.LEFT, padx=(5, 0))
        ttk.Label(
            plan_divide_row,
            text="(Plan Qty is divided by this number before comparing to Loaded Qty)",
            font=('Segoe UI', 8),
            foreground='gray'
        ).pack(side=tk.LEFT, padx=(8, 0))
        
        # Plan Text-to-Columns options
        plan_split_check_row = ttk.Frame(file_frame)
        plan_split_check_row.pack(fill=tk.X, pady=(8, 2), padx=(10, 0))
        ttk.Checkbutton(
            plan_split_check_row,
            text="Enable Text-to-Columns (Fixed Width)",
            variable=self.plan_split_enabled,
            command=self.toggle_plan_split
        ).pack(side=tk.LEFT)
        
        plan_split_frame = ttk.Frame(file_frame)
        plan_split_frame.pack(fill=tk.X, pady=2, padx=(10, 0))
        
        ttk.Label(plan_split_frame, text="Column to Split:", width=15).pack(side=tk.LEFT)
        self.plan_split_combo = ttk.Combobox(
            plan_split_frame,
            textvariable=self.plan_split_column,
            state="readonly"
        )
        self.plan_split_combo.pack(side=tk.LEFT, padx=(0, 10), fill=tk.X, expand=True)
        self.plan_split_combo['state'] = 'disabled'
        
        ttk.Button(
            plan_split_frame,
            text="Configure Split",
            command=lambda: self.configure_split('plan'),
            state='disabled'
        ).pack(side=tk.LEFT)
        
        # Error message label (red)
        error_row = ttk.Frame(file_frame)
        error_row.pack(fill=tk.X, pady=2, padx=(10, 0))
        
        self.error_label = tk.Label(
            error_row,
            textvariable=self.error_message,
            font=('Segoe UI', 8, 'bold'),
            fg='red'
        )
        self.error_label.pack(anchor=tk.W)
        
        # ===== RIGHT COLUMN: DATA PREVIEW (top) + ACTIVITY LOG (bottom) =====
        # A vertical splitter lets the user drag to resize either section.
        right_paned = ttk.PanedWindow(right_column, orient=tk.VERTICAL)
        right_paned.pack(fill=tk.BOTH, expand=True)

        preview_frame = ttk.LabelFrame(right_paned, text="📊 Data Preview", padding="10")
        right_paned.add(preview_frame, weight=4)

        preview_paned = ttk.PanedWindow(preview_frame, orient=tk.VERTICAL)
        preview_paned.pack(fill=tk.BOTH, expand=True)

        # Top: Loading History preview
        history_preview_frame = ttk.Frame(preview_paned)
        preview_paned.add(history_preview_frame, weight=1)

        history_preview_header = ttk.Frame(history_preview_frame)
        history_preview_header.pack(fill=tk.X, pady=(0, 4))
        ttk.Label(history_preview_header, text="Loading History (Top)", font=('Segoe UI', 10, 'bold')).pack(side=tk.LEFT)
        ttk.Label(history_preview_header, textvariable=self.history_preview_count, font=('Segoe UI', 8), foreground='gray').pack(side=tk.LEFT, padx=(10, 0))
        ttk.Button(
            history_preview_header,
            text="🔄 Refresh",
            width=10,
            command=lambda: self.refresh_preview('history')
        ).pack(side=tk.RIGHT)

        self.history_tree = self.build_preview_pane(history_preview_frame)

        # Bottom: Load Plan preview
        plan_preview_frame = ttk.Frame(preview_paned)
        preview_paned.add(plan_preview_frame, weight=1)

        plan_preview_header = ttk.Frame(plan_preview_frame)
        plan_preview_header.pack(fill=tk.X, pady=(6, 4))
        ttk.Label(plan_preview_header, text="Load Plan (Bottom)", font=('Segoe UI', 10, 'bold')).pack(side=tk.LEFT)
        ttk.Label(plan_preview_header, textvariable=self.plan_preview_count, font=('Segoe UI', 8), foreground='gray').pack(side=tk.LEFT, padx=(10, 0))
        ttk.Button(
            plan_preview_header,
            text="🔄 Refresh",
            width=10,
            command=lambda: self.refresh_preview('plan')
        ).pack(side=tk.RIGHT)

        self.plan_tree = self.build_preview_pane(plan_preview_frame)

        # Shared status bar for row selection in either preview
        preview_status_bar = ttk.Label(
            preview_frame,
            textvariable=self.preview_status_var,
            relief=tk.SUNKEN,
            anchor=tk.W,
            padding=(5, 3),
            font=('Segoe UI', 9)
        )
        preview_status_bar.pack(fill=tk.X, pady=(8, 0))

        # ===== FOLDER SELECTION FRAME (left column) =====
        folder_frame = ttk.LabelFrame(left_column, text="📂 Output Settings", padding="15")
        folder_frame.pack(fill=tk.X, pady=(0, 15))
        
        output_row = ttk.Frame(folder_frame)
        output_row.pack(fill=tk.X, pady=2)
        
        ttk.Label(output_row, text="Output Folder:", font=('Segoe UI', 10)).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Entry(output_row, textvariable=self.output_folder).pack(side=tk.LEFT, padx=(0, 10), fill=tk.X, expand=True)
        ttk.Button(output_row, text="Browse", command=self.browse_output_folder).pack(side=tk.LEFT)
        
        archive_row = ttk.Frame(folder_frame)
        archive_row.pack(fill=tk.X, pady=2)
        
        ttk.Label(archive_row, text="Archive Folder:", font=('Segoe UI', 10)).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Entry(archive_row, textvariable=self.archive_folder).pack(side=tk.LEFT, padx=(0, 10), fill=tk.X, expand=True)
        ttk.Button(archive_row, text="Browse", command=self.browse_archive_folder).pack(side=tk.LEFT)
        
        # ===== OPTIONS FRAME (left column) =====
        options_frame = ttk.LabelFrame(left_column, text="⚙️ Options", padding="15")
        options_frame.pack(fill=tk.X, pady=(0, 15))
        
        options_left = ttk.Frame(options_frame)
        options_left.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        ttk.Checkbutton(
            options_left,
            text="Archive files after processing",
            variable=self.archive_var
        ).pack(anchor=tk.W)
        
        ttk.Checkbutton(
            options_left,
            text="Auto-open result file after completion",
            variable=self.auto_open_file
        ).pack(anchor=tk.W)
        
        options_right = ttk.Frame(options_frame)
        options_right.pack(side=tk.RIGHT, fill=tk.X)
        
        ttk.Checkbutton(
            options_right,
            text="Verbose logging",
            variable=self.verbose_logging
        ).pack(anchor=tk.W)
        
        # ===== PROGRESS BAR (left column) =====
        progress_frame = ttk.Frame(left_column)
        progress_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.progress_bar = ttk.Progressbar(
            progress_frame,
            variable=self.progress_var,
            maximum=100,
            mode='determinate'
        )
        self.progress_bar.pack(fill=tk.X)
        
        status_frame = ttk.Frame(left_column)
        status_frame.pack(fill=tk.X, pady=(5, 10))
        
        self.status_label = ttk.Label(status_frame, textvariable=self.status_var, font=('Segoe UI', 10))
        self.status_label.pack(side=tk.LEFT)
        
        # ===== LOG FRAME (right column, bottom pane of the splitter) =====
        log_frame = ttk.LabelFrame(right_paned, text="📋 Activity Log", padding="10")
        right_paned.add(log_frame, weight=2)
        
        log_container = ttk.Frame(log_frame)
        log_container.pack(fill=tk.BOTH, expand=True)
        
        self.log_text = tk.Text(
            log_container,
            height=8,
            font=('Consolas', 9),
            bg='#0b1420',
            fg='#f8fafc',
            wrap=tk.WORD,
            state=tk.DISABLED
        )
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        log_scrollbar = ttk.Scrollbar(log_container, orient="vertical", command=self.log_text.yview)
        log_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.log_text.config(yscrollcommand=log_scrollbar.set)
        
        # ===== ACTION BUTTONS (left column, stacked to fit a narrower column) =====
        button_frame = ttk.Frame(left_column)
        button_frame.pack(fill=tk.X, pady=(10, 0))
        
        self.run_button = tk.Button(
            button_frame,
            text="🚀 Run Reconciliation",
            font=('Segoe UI', 12, 'bold'),
            bg=self.colors['secondary'],
            fg=self.colors['white'],
            pady=10,
            command=self.run_reconciliation,
            cursor="hand2"
        )
        self.run_button.pack(fill=tk.X, pady=(0, 8))
        
        secondary_button_row = ttk.Frame(button_frame)
        secondary_button_row.pack(fill=tk.X)
        
        ttk.Button(
            secondary_button_row,
            text="Clear All",
            command=self.clear_all
        ).pack(side=tk.LEFT, padx=(0, 10))
        
        ttk.Button(
            secondary_button_row,
            text="Clear Log",
            command=self.clear_log
        ).pack(side=tk.LEFT, padx=(0, 10))
        
        ttk.Button(
            secondary_button_row,
            text="Exit",
            command=self.on_app_close
        ).pack(side=tk.RIGHT)
        
        self.footer_label = tk.Label(
            main_frame,
            text="© 2026 Load Reconciliation v2.1 | Made with ❤️",
            font=('Segoe UI', 8),
            fg='#BDC3C7'
        )
        self.footer_label.grid(row=1, column=0, columnspan=2, pady=(15, 0))
        
        # Initial log message
        self.add_log("ℹ️ Application started. Select files and click Run Reconciliation.", "INFO")
        
        # Bind mouse wheel for scrolling anywhere within this tool window
        self._bind_mousewheel()

    def _bind_mousewheel(self):
        """Bind mouse wheel events so the user can scroll naturally anywhere within the window."""
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel, add="+")
        self.canvas.bind_all("<Button-4>", self._on_mousewheel, add="+")
        self.canvas.bind_all("<Button-5>", self._on_mousewheel, add="+")

    def _on_mousewheel(self, event):
        """Handle mouse wheel scrolling for the canvas and inner scrollable widgets."""
        try:
            # If the tool's canvas is not currently visible/mapped (e.g. on another tab), do nothing
            if not self.canvas.winfo_ismapped():
                return

            widget = event.widget
            if not widget:
                return

            # Make sure event is inside our window
            if widget.winfo_toplevel() != self.canvas.winfo_toplevel():
                return

            # Check if the hovered widget is inside our theme_target / canvas
            curr = widget
            inside_tool = False
            while curr:
                if curr == self.canvas or curr == self.theme_target:
                    inside_tool = True
                    break
                curr = getattr(curr, 'master', None)

            if not inside_tool:
                return

            # Calculate scroll units
            if hasattr(event, 'delta') and event.delta:
                # Windows delta is +/- 120 per notch.
                # Scrolling 3 units per notch provides a smooth and responsive experience.
                scroll_units = int(-1 * (event.delta / 120)) * 3
            elif getattr(event, 'num', None) == 4:
                scroll_units = -3
            elif getattr(event, 'num', None) == 5:
                scroll_units = 3
            else:
                return

            # If hovering directly over an inner scrollable widget (Treeviews or Log Text), scroll that
            if widget in (self.history_tree, self.plan_tree, self.log_text):
                line_units = int(scroll_units / 3) or (1 if scroll_units > 0 else -1)
                widget.yview_scroll(line_units, "units")
                return "break"

            # Otherwise, scroll the main tool canvas
            self.canvas.yview_scroll(scroll_units, "units")
            return "break"
        except Exception:
            pass

    def add_log(self, message, level="INFO"):
        """Add message to log with timestamp and color"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_entry = f"[{timestamp}] {message}\n"
        
        self.log_text.config(state=tk.NORMAL)
        
        tags = {
            "INFO": "info",
            "SUCCESS": "success",
            "WARNING": "warning",
            "ERROR": "error",
            "DEBUG": "debug"
        }
        
        tag = tags.get(level, "info")
        self.log_text.insert(tk.END, log_entry, tag)
        
        self.log_text.tag_configure("info", foreground="#ECF0F1")
        self.log_text.tag_configure("success", foreground="#2ECC71")
        self.log_text.tag_configure("warning", foreground="#F1C40F")
        self.log_text.tag_configure("error", foreground="#E74C3C")
        self.log_text.tag_configure("debug", foreground="#3498DB")
        
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)
        
        self.log_messages.append(log_entry)
        
    def clear_log(self):
        """Clear the log text widget"""
        self.log_text.config(state=tk.NORMAL)
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state=tk.DISABLED)
        self.log_messages.clear()
        self.add_log("📋 Log cleared", "INFO")
        
    # ===================== DATA PREVIEW =====================

    def build_preview_pane(self, parent):
        """Build a Treeview with vertical + horizontal scrollbars inside parent, return the Treeview"""
        container = ttk.Frame(parent)
        container.pack(fill=tk.BOTH, expand=True)

        vsb = ttk.Scrollbar(container, orient="vertical")
        hsb = ttk.Scrollbar(container, orient="horizontal")

        tree = ttk.Treeview(
            container,
            show="headings",
            height=8,
            yscrollcommand=vsb.set,
            xscrollcommand=hsb.set,
            selectmode="browse"
        )

        vsb.config(command=tree.yview)
        hsb.config(command=tree.xview)

        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        container.rowconfigure(0, weight=1)
        container.columnconfigure(0, weight=1)

        tree.tag_configure('oddrow', background='#F5F5F5')
        tree.tag_configure('evenrow', background='#FFFFFF')

        tree.bind("<<TreeviewSelect>>", lambda e, t=tree: self.on_preview_row_select(t))

        return tree

    def make_swatch(self, color, width=22, height=5):
        """Create a small solid-color PhotoImage to use as a Treeview heading marker"""
        img = tk.PhotoImage(width=width, height=height)
        img.put(color, to=(0, 0, width, height))
        return img

    def make_header_badge(self, text, bg_color="#F59E0B", fg_color="#ffffff", height=20, icon="🔑"):
        """Create a rounded colored background badge image for Treeview heading."""
        full_text = f"{icon} {text}" if icon else str(text)
        try:
            font = ImageFont.truetype("segoeuib.ttf", 11)
        except Exception:
            try:
                font = ImageFont.truetype("arialbd.ttf", 11)
            except Exception:
                try:
                    font = ImageFont.truetype("segoeui.ttf", 11)
                except Exception:
                    font = ImageFont.load_default()

        try:
            bbox = font.getbbox(full_text)
            text_w = bbox[2] - bbox[0]
            text_h = bbox[3] - bbox[1]
        except Exception:
            text_w = len(full_text) * 7
            text_h = 12

        width = max(text_w + 16, 50)
        img = Image.new("RGBA", (width, height), (0, 0, 0, 0))
        draw = ImageDraw.Draw(img)

        # Draw filled colored rectangle with rounded corners
        draw.rounded_rectangle([0, 1, width - 1, height - 1], radius=3, fill=bg_color)

        # Draw text
        text_x = 8
        text_y = max(1, (height - text_h) // 2 - 1)
        draw.text((text_x, text_y), full_text, fill=fg_color, font=font)

        return ImageTk.PhotoImage(img), width

    def populate_preview(self, file_type):
        """Fill the given preview Treeview with data from the loaded dataframe"""
        if file_type == 'history':
            tree = self.history_tree
            df = self.history_df_loaded
            count_var = self.history_preview_count
        else:
            tree = self.plan_tree
            df = self.plan_df_loaded
            count_var = self.plan_preview_count

        if tree is None:
            return

        # Clear existing
        tree.delete(*tree.get_children())
        tree['columns'] = ()

        if df is None or df.empty:
            count_var.set("No data loaded")
            return

        columns = [str(c) for c in df.columns]
        tree['columns'] = columns

        for col in columns:
            tree.heading(col, text=col, anchor=tk.W)
            col_str = str(col)
            try:
                data_max_len = int(df[col].astype(str).str.len().max()) if len(df) else len(col_str)
            except Exception:
                data_max_len = len(col_str)
            header_w = len(col_str) * 7.2 + 18
            data_w = min(220, data_max_len * 6.8 + 14)
            width = max(50, int(max(header_w, data_w)))
            tree.column(col, width=width, minwidth=40, anchor=tk.W, stretch=False)

        preview_df = df.head(self.PREVIEW_MAX_ROWS)
        for i, (_, row) in enumerate(preview_df.iterrows()):
            values = ["" if pd.isna(v) else str(v) for v in row.tolist()]
            tag = 'evenrow' if i % 2 == 0 else 'oddrow'
            tree.insert("", tk.END, iid=str(i), values=values, tags=(tag,))

        total_rows = len(df)
        shown = len(preview_df)
        if total_rows > shown:
            count_var.set(f"Showing {shown} of {total_rows} rows")
        else:
            count_var.set(f"{total_rows} rows")

        self.highlight_preview_columns(file_type)

    def highlight_preview_columns(self, file_type):
        """Visually mark the selected Primary Key (Load ID) and Quantity columns in the preview header with colored background"""
        if file_type == 'history':
            tree = self.history_tree
            load_id_col = self.history_load_id_col.get()
            qty_col = self.history_qty_col.get()
            img_store = self._history_marker_imgs
        else:
            tree = self.plan_tree
            load_id_col = self.plan_load_id_col.get()
            qty_col = self.plan_qty_col.get()
            img_store = self._plan_marker_imgs

        if tree is None:
            return

        columns = tree['columns']
        if not columns:
            return

        img_store.clear()

        for col in columns:
            if col == load_id_col and col == qty_col:
                img, badge_w = self.make_header_badge(col, bg_color="#F59E0B", fg_color="#ffffff", icon="🔑🔢")
                extra = badge_w + 8
            elif col == load_id_col:
                img, badge_w = self.make_header_badge(col, bg_color="#F59E0B", fg_color="#ffffff", icon="🔑")
                extra = badge_w + 8
            elif col == qty_col:
                img, badge_w = self.make_header_badge(col, bg_color="#10B981", fg_color="#ffffff", icon="🔢")
                extra = badge_w + 8
            else:
                img = None
                extra = 0

            # Sizing check: snugly fit header badge without over-extending
            if extra > 0:
                try:
                    current_w = int(tree.column(col, 'width'))
                    if current_w < extra:
                        tree.column(col, width=extra)
                except Exception:
                    pass

            if img is not None:
                img_store[col] = img
                tree.heading(col, image=img, text='', anchor=tk.W)
            else:
                tree.heading(col, text=col, image='', anchor=tk.W)

    def on_preview_row_select(self, tree):
        """Show the values of the selected preview row in the status bar"""
        selection = tree.selection()
        if not selection:
            return

        item = selection[0]
        values = tree.item(item, "values")
        columns = tree['columns']

        source = "Loading History" if tree is self.history_tree else "Load Plan"

        pairs = []
        for col, val in zip(columns, values):
            marker = ""
            if tree is self.history_tree:
                if col == self.history_load_id_col.get():
                    marker = "🔑 "
                elif col == self.history_qty_col.get():
                    marker = "🔢 "
            else:
                if col == self.plan_load_id_col.get():
                    marker = "🔑 "
                elif col == self.plan_qty_col.get():
                    marker = "🔢 "
            pairs.append(f"{marker}{col}: {val}" if marker else f"{col}: {val}")

        row_num = int(item) + 1
        self.preview_status_var.set(f"📌 {source} — Row {row_num}: " + " | ".join(pairs))

    def refresh_preview(self, file_type):
        """Manually refresh a preview pane from its currently loaded dataframe"""
        df = self.history_df_loaded if file_type == 'history' else self.plan_df_loaded
        if df is None:
            messagebox.showinfo("No Data", f"Please load a {'Loading History' if file_type == 'history' else 'Load Plan'} file first.")
            return
        self.populate_preview(file_type)
        self.add_log(f"🔄 {'Loading History' if file_type == 'history' else 'Load Plan'} preview refreshed", "INFO")

    def validate_excel_file(self, file_path):
        """Validate that the file is a valid Excel file"""
        if not file_path:
            return False, "No file selected"
        
        if not os.path.exists(file_path):
            return False, "File does not exist"
        
        valid_extensions = ('.xlsx', '.xls', '.xlsm', '.xlsb')
        if not file_path.lower().endswith(valid_extensions):
            return False, "File is not a valid Excel file (must be .xlsx or .xls)"
        
        try:
            pd.read_excel(file_path, nrows=1)
            return True, "Valid Excel file"
        except Exception as e:
            return False, f"Invalid Excel file: {str(e)}"
    
    def toggle_history_split(self):
        """Enable/disable history text-to-columns"""
        if self.history_split_enabled.get():
            self.history_split_combo['state'] = 'readonly'
            self.history_split_combo['values'] = self.history_columns
            for child in self.history_split_combo.master.winfo_children():
                if isinstance(child, ttk.Button) and child['text'] == 'Configure Split':
                    child['state'] = 'normal'
        else:
            self.history_split_combo['state'] = 'disabled'
            self.history_split_column.set("")
            for child in self.history_split_combo.master.winfo_children():
                if isinstance(child, ttk.Button) and child['text'] == 'Configure Split':
                    child['state'] = 'disabled'
    
    def toggle_plan_split(self):
        """Enable/disable plan text-to-columns"""
        if self.plan_split_enabled.get():
            self.plan_split_combo['state'] = 'readonly'
            self.plan_split_combo['values'] = self.plan_columns
            for child in self.plan_split_combo.master.winfo_children():
                if isinstance(child, ttk.Button) and child['text'] == 'Configure Split':
                    child['state'] = 'normal'
        else:
            self.plan_split_combo['state'] = 'disabled'
            self.plan_split_column.set("")
            for child in self.plan_split_combo.master.winfo_children():
                if isinstance(child, ttk.Button) and child['text'] == 'Configure Split':
                    child['state'] = 'disabled'
    
    def toggle_plan_divide(self):
        """Enable/disable dividing Plan Qty by a user-supplied number"""
        if self.plan_divide_enabled.get():
            self.plan_divide_entry['state'] = 'normal'
            if not self.plan_divide_value.get().strip():
                self.plan_divide_value.set("1")
        else:
            self.plan_divide_entry['state'] = 'disabled'
    
    def configure_split(self, file_type):
        """Open text-to-columns configuration dialog"""
        if file_type == 'history':
            if not self.history_split_column.get():
                messagebox.showwarning("Warning", "Please select a column to split")
                return
            if self.history_df_loaded is None:
                messagebox.showerror("Error", "Please load a file first")
                return
            
            column_name = self.history_split_column.get()
            if column_name not in self.history_df_loaded.columns:
                messagebox.showerror("Error", f"Column '{column_name}' not found")
                return
            
            column_data = self.history_df_loaded[column_name].astype(str).tolist()
            
            dialog = TextToColumnsDialog(
                self.root,
                f"Split History Column: {column_name}",
                column_data,
                column_name,
                app=self
            )
            self.root.wait_window(dialog.dialog)
            
            if dialog.result is not None:
                self.history_split_widths = dialog.widths
                self.history_original_data = dialog.result
                self.add_log(f"✅ History text-to-columns configured: {len(self.history_split_widths)} columns", "SUCCESS")
                self.add_log(f"   Widths: {self.history_split_widths}", "INFO")

                preview_df = self.apply_text_to_columns(
                    self.history_df_loaded, column_name, self.history_original_data, "History"
                )
                self.history_df_loaded = preview_df
                new_columns = preview_df.columns.tolist()
                self.history_columns = new_columns
                self.history_load_id_combo['values'] = new_columns
                self.history_qty_combo['values'] = new_columns
                self.history_split_combo['values'] = new_columns
                # Auto-select the second part of the split columns as Primary Key (Load ID)
                part_2_col = f"{column_name}_part_2"
                if part_2_col in new_columns:
                    self.history_load_id_col.set(part_2_col)
                elif len(new_columns) > 1:
                    self.history_load_id_col.set(new_columns[1])
                elif new_columns:
                    self.history_load_id_col.set(new_columns[0])
                if self.history_qty_col.get() not in new_columns:
                    self.history_qty_col.set("")
                self.populate_preview('history')
                self.add_log("👁️ Loading History preview updated with split columns at beginning", "INFO")
                
        elif file_type == 'plan':
            if not self.plan_split_column.get():
                messagebox.showwarning("Warning", "Please select a column to split")
                return
            if self.plan_df_loaded is None:
                messagebox.showerror("Error", "Please load a file first")
                return
            
            column_name = self.plan_split_column.get()
            if column_name not in self.plan_df_loaded.columns:
                messagebox.showerror("Error", f"Column '{column_name}' not found")
                return
            
            column_data = self.plan_df_loaded[column_name].astype(str).tolist()
            
            dialog = TextToColumnsDialog(
                self.root,
                f"Split Plan Column: {column_name}",
                column_data,
                column_name,
                app=self
            )
            self.root.wait_window(dialog.dialog)
            
            if dialog.result is not None:
                self.plan_split_widths = dialog.widths
                self.plan_original_data = dialog.result
                self.add_log(f"✅ Plan text-to-columns configured: {len(self.plan_split_widths)} columns", "SUCCESS")
                self.add_log(f"   Widths: {self.plan_split_widths}", "INFO")

                preview_df = self.apply_text_to_columns(
                    self.plan_df_loaded, column_name, self.plan_original_data, "Plan"
                )
                self.plan_df_loaded = preview_df
                new_columns = preview_df.columns.tolist()
                self.plan_columns = new_columns
                self.plan_load_id_combo['values'] = new_columns
                self.plan_qty_combo['values'] = new_columns
                self.plan_split_combo['values'] = new_columns
                # Auto-select the second part of the split columns as Primary Key (Load ID)
                part_2_col = f"{column_name}_part_2"
                if part_2_col in new_columns:
                    self.plan_load_id_col.set(part_2_col)
                elif len(new_columns) > 1:
                    self.plan_load_id_col.set(new_columns[1])
                elif new_columns:
                    self.plan_load_id_col.set(new_columns[0])
                if self.plan_qty_col.get() not in new_columns:
                    self.plan_qty_col.set("")
                self.populate_preview('plan')
                self.add_log("👁️ Load Plan preview updated with split columns at beginning", "INFO")
    
    def clean_whole_number_columns(self, df):
        """Excel stores plain numbers without an explicit format as floats, so
        pandas reads whole numbers (Load IDs, quantities, etc.) as e.g. 100.0
        instead of 100. This converts any numeric column that contains only
        whole numbers (ignoring blanks) to a nullable integer type so it
        displays and matches cleanly, without touching columns that
        genuinely have decimals."""
        for col in df.columns:
            if pd.api.types.is_float_dtype(df[col]):
                non_null = df[col].dropna()
                if len(non_null) > 0 and (non_null == non_null.round(0)).all():
                    try:
                        df[col] = df[col].astype('Int64')
                    except (TypeError, ValueError, OverflowError):
                        pass
        return df
    
    def load_column_names(self, file_path, combo_load_id, combo_qty, target_var_load_id, target_var_qty, split_combo=None):
        """Load column names from Excel file into comboboxes"""
        try:
            # Try reading with headers first
            df = pd.read_excel(file_path)
            
            # If no columns found, try reading without headers
            if df.empty or len(df.columns) == 0:
                df = pd.read_excel(file_path, header=None)
                if not df.empty:
                    # Create column names
                    columns = [f"Column_{i+1}" for i in range(len(df.columns))]
                    df.columns = columns
                else:
                    self.add_log(f"⚠️ File appears to be empty", "WARNING")
                    return []
            
            # Whole-number columns come back from Excel as floats (100.0
            # instead of 100) - clean those up before anything else uses them.
            df = self.clean_whole_number_columns(df)
            
            columns = df.columns.tolist()
            
            print(f"DEBUG: Loaded columns: {columns}")
            print(f"DEBUG: Number of rows: {len(df)}")
            
            # Store the dataframe
            if combo_load_id == self.history_load_id_combo:
                self.history_df_loaded = df
                self.history_columns = columns
                if split_combo:
                    split_combo['values'] = columns
            else:
                self.plan_df_loaded = df
                self.plan_columns = columns
                if split_combo:
                    split_combo['values'] = columns
            
            # Populate comboboxes with column names
            combo_load_id['values'] = columns
            combo_qty['values'] = columns
            
            # Auto-select columns
            target_var_load_id.set("")
            target_var_qty.set("")

            # 1. Primary Key / Load ID auto-selection
            for col in columns:
                col_lower = str(col).lower().strip()
                if 'load' in col_lower and ('id' in col_lower or '#' in col_lower or 'number' in col_lower or 'no' in col_lower):
                    target_var_load_id.set(col)
                    break

            if not target_var_load_id.get():
                for col in columns:
                    col_lower = str(col).lower().strip()
                    if 'load' in col_lower:
                        target_var_load_id.set(col)
                        break

            if not target_var_load_id.get():
                for col in columns:
                    col_lower = str(col).lower().strip()
                    if 'id' in col_lower or 'key' in col_lower:
                        target_var_load_id.set(col)
                        break

            if not target_var_load_id.get() and columns:
                target_var_load_id.set(columns[0])

            # 2. Quantity column auto-selection with priority for specified column names:
            # - QUANTITY
            # - Actual Qty
            # - Pln Qty
            # - Loaded Qty
            # - Planned Qty
            target_qty_patterns = [
                "quantity",
                "actual qty",
                "pln qty",
                "loaded qty",
                "planned qty"
            ]

            # Priority 1: Exact / normalized matches to user-specified quantity list
            for pattern in target_qty_patterns:
                pat_norm = re.sub(r'[^a-z0-9]', '', pattern.lower())
                for col in columns:
                    col_norm = re.sub(r'[^a-z0-9]', '', str(col).lower())
                    if col_norm == pat_norm:
                        target_var_qty.set(col)
                        break
                if target_var_qty.get():
                    break

            # Priority 2: Substring matches to user-specified quantity patterns
            if not target_var_qty.get():
                for pattern in target_qty_patterns:
                    pat_norm = re.sub(r'[^a-z0-9]', '', pattern.lower())
                    for col in columns:
                        col_norm = re.sub(r'[^a-z0-9]', '', str(col).lower())
                        if pat_norm in col_norm:
                            target_var_qty.set(col)
                            break
                    if target_var_qty.get():
                        break

            # Priority 3: General quantity heuristic keywords
            if not target_var_qty.get():
                for col in columns:
                    col_lower = str(col).lower().strip()
                    if any(term in col_lower for term in ['qty', 'quantity', 'pick', 'actual']):
                        target_var_qty.set(col)
                        break

            # Priority 4: Fallback selections
            if not target_var_qty.get() and columns:
                target_var_qty.set(columns[-1])
            
            self.add_log(f"✅ Loaded {len(columns)} columns: {', '.join(columns[:5])}{'...' if len(columns) > 5 else ''}", "INFO")
            
            # Force UI update
            self.root.update_idletasks()
            
            return columns
        except Exception as e:
            self.add_log(f"⚠️ Could not read column names: {e}", "WARNING")
            print(f"ERROR loading columns: {e}")
            import traceback
            traceback.print_exc()
            return []
    
    def apply_text_to_columns(self, df, split_column, split_data, file_type):
        """Apply the text-to-columns split to the dataframe and place new columns at the beginning"""
        if split_data is None or not split_column:
            return df
        
        self.add_log(f"🔄 Applying text-to-columns to {file_type}...", "INFO")
        
        df_copy = df.copy()
        df_copy = df_copy.drop(columns=[split_column])
        
        new_cols_created = []
        for i in range(len(split_data[0])):
            new_col_name = f"{split_column}_part_{i+1}"
            values = [row[i] if i < len(row) else "" for row in split_data]
            df_copy[new_col_name] = values
            new_cols_created.append(new_col_name)
            self.add_log(f"   Created column: {new_col_name}", "DEBUG")
        
        # Place newly split columns at the very beginning of the sheet / DataFrame
        remaining_cols = [c for c in df_copy.columns if c not in new_cols_created]
        df_copy = df_copy[new_cols_created + remaining_cols]
        
        self.add_log(f"✅ Text-to-columns applied: {len(split_data[0])} new columns placed at the top/beginning", "SUCCESS")
        return df_copy
    
    def open_concatenate_dialog(self, file_type):
        """Open the Concatenate Columns dialog for the given file type ('history' or 'plan')"""
        df = self.history_df_loaded if file_type == 'history' else self.plan_df_loaded
        if df is None or df.empty:
            messagebox.showerror("Error", "Please load a file first")
            return
        
        friendly_name = "Loading History" if file_type == 'history' else "Load Plan"
        dialog = ConcatenateColumnsDialog(
            self.root,
            f"Concatenate Columns - {friendly_name}",
            df,
            app=self
        )
        self.root.wait_window(dialog.dialog)
        
        if dialog.result is None:
            return
        
        columns = dialog.result['columns']
        separator = dialog.result['separator']
        new_column_name = dialog.result['new_column_name']
        drop_originals = dialog.result['drop_originals']
        
        new_df = self.apply_column_concatenation(df, columns, separator, new_column_name, drop_originals, friendly_name)
        new_columns = new_df.columns.tolist()
        
        # Remember this operation so it can be replayed on the fresh copy of
        # the file that's read when reconciliation actually runs.
        config = dict(dialog.result)
        
        if file_type == 'history':
            self.history_concat_configs.append(config)
            self.history_df_loaded = new_df
            self.history_columns = new_columns
            self.history_load_id_combo['values'] = new_columns
            self.history_qty_combo['values'] = new_columns
            self.history_split_combo['values'] = new_columns
            # Automatically set newly concatenated column as Primary Key (Load ID)
            self.history_load_id_col.set(new_column_name)
            if self.history_qty_col.get() not in new_columns:
                self.history_qty_col.set("")
            if self.history_split_column.get() not in new_columns:
                self.history_split_column.set("")
            self.populate_preview('history')
            self.add_log("👁️ Loading History preview updated with concatenated column at beginning", "INFO")
        else:
            self.plan_concat_configs.append(config)
            self.plan_df_loaded = new_df
            self.plan_columns = new_columns
            self.plan_load_id_combo['values'] = new_columns
            self.plan_qty_combo['values'] = new_columns
            self.plan_split_combo['values'] = new_columns
            # Automatically set newly concatenated column as Primary Key (Load ID)
            self.plan_load_id_col.set(new_column_name)
            if self.plan_qty_col.get() not in new_columns:
                self.plan_qty_col.set("")
            if self.plan_split_column.get() not in new_columns:
                self.plan_split_column.set("")
            self.populate_preview('plan')
            self.add_log("👁️ Load Plan preview updated with concatenated column at beginning", "INFO")
    
    def apply_column_concatenation(self, df, columns, separator, new_column_name, drop_originals, file_type):
        """Combine the given columns (in order) into a single new column, joined by separator, placed at the beginning"""
        self.add_log(f"🔗 Concatenating {file_type} columns into '{new_column_name}': {', '.join(columns)}", "INFO")
        
        df_copy = df.copy()
        
        def concat_row(row):
            parts = []
            for col in columns:
                val = row[col]
                parts.append("" if pd.isna(val) else str(val))
            return separator.join(parts)
        
        df_copy[new_column_name] = df_copy.apply(concat_row, axis=1)
        
        if drop_originals:
            cols_to_drop = [c for c in columns if c != new_column_name and c in df_copy.columns]
            df_copy = df_copy.drop(columns=cols_to_drop)
            self.add_log(f"   Removed original columns: {', '.join(cols_to_drop)}", "DEBUG")
        
        # Place newly concatenated column at the very beginning of the sheet / DataFrame
        remaining_cols = [c for c in df_copy.columns if c != new_column_name]
        df_copy = df_copy[[new_column_name] + remaining_cols]
        
        self.add_log(f"✅ Concatenation applied: '{new_column_name}' placed at the top/beginning", "SUCCESS")
        return df_copy
    
    def browse_history_file(self):
        file_path = filedialog.askopenfilename(
            title="Select Loading History File",
            filetypes=[("Excel files", "*.xlsx *.xls *.xlsm *.xlsb"), ("All files", "*.*")]
        )
        if file_path:
            self.history_file_path.set(file_path)
            display_name = os.path.basename(file_path)
            if len(display_name) > 50:
                display_name = display_name[:25] + "..." + display_name[-10:]
            self.history_file_info.set(f"📄 {display_name} ({self.get_file_size(file_path)})")
            self.error_message.set("")
            self.add_log(f"📂 Selected History file: {os.path.basename(file_path)}", "INFO")
            
            valid, msg = self.validate_excel_file(file_path)
            if not valid:
                self.error_message.set(f"⚠️ {msg}")
                self.add_log(f"⚠️ History file validation failed: {msg}", "WARNING")
            else:
                self.history_columns = self.load_column_names(
                    file_path,
                    self.history_load_id_combo,
                    self.history_qty_combo,
                    self.history_load_id_col,
                    self.history_qty_col,
                    self.history_split_combo
                )
                if self.history_columns:
                    self.add_log(f"📊 Loaded {len(self.history_columns)} columns from History file", "INFO")
                    
                    if self.history_split_enabled.get():
                        self.history_split_combo['values'] = self.history_columns
                        self.history_split_combo['state'] = 'readonly'

                    self.history_concat_button['state'] = 'normal'
                    self.populate_preview('history')
                    self.add_log("👁️ Loading History preview updated", "INFO")
                else:
                    self.error_message.set("⚠️ Could not read columns from file")
                    self.add_log("⚠️ Could not read columns from History file", "WARNING")
    
    def browse_plan_file(self):
        file_path = filedialog.askopenfilename(
            title="Select Load Plan File",
            filetypes=[("Excel files", "*.xlsx *.xls *.xlsm *.xlsb"), ("All files", "*.*")]
        )
        if file_path:
            self.plan_file_path.set(file_path)
            display_name = os.path.basename(file_path)
            if len(display_name) > 50:
                display_name = display_name[:25] + "..." + display_name[-10:]
            self.plan_file_info.set(f"📄 {display_name} ({self.get_file_size(file_path)})")
            self.error_message.set("")
            self.add_log(f"📂 Selected Plan file: {os.path.basename(file_path)}", "INFO")
            
            valid, msg = self.validate_excel_file(file_path)
            if not valid:
                self.error_message.set(f"⚠️ {msg}")
                self.add_log(f"⚠️ Plan file validation failed: {msg}", "WARNING")
            else:
                self.plan_columns = self.load_column_names(
                    file_path,
                    self.plan_load_id_combo,
                    self.plan_qty_combo,
                    self.plan_load_id_col,
                    self.plan_qty_col,
                    self.plan_split_combo
                )
                if self.plan_columns:
                    self.add_log(f"📊 Loaded {len(self.plan_columns)} columns from Plan file", "INFO")
                    
                    if self.plan_split_enabled.get():
                        self.plan_split_combo['values'] = self.plan_columns
                        self.plan_split_combo['state'] = 'readonly'

                    self.plan_concat_button['state'] = 'normal'
                    self.populate_preview('plan')
                    self.add_log("👁️ Load Plan preview updated", "INFO")
                else:
                    self.error_message.set("⚠️ Could not read columns from file")
                    self.add_log("⚠️ Could not read columns from Plan file", "WARNING")
    
    def get_file_size(self, file_path):
        size = os.path.getsize(file_path)
        if size < 1024:
            return f"{size} B"
        elif size < 1024 * 1024:
            return f"{size / 1024:.1f} KB"
        else:
            return f"{size / (1024 * 1024):.1f} MB"
    
    def browse_output_folder(self):
        folder_path = filedialog.askdirectory(title="Select Output Folder")
        if folder_path:
            self.output_folder.set(folder_path)
            self.add_log(f"📁 Output folder set to: {folder_path}", "INFO")
    
    def browse_archive_folder(self):
        folder_path = filedialog.askdirectory(title="Select Archive Folder")
        if folder_path:
            self.archive_folder.set(folder_path)
            self.add_log(f"📁 Archive folder set to: {folder_path}", "INFO")
    
    def clear_all(self):
        self.history_file_path.set("")
        self.plan_file_path.set("")
        self.history_file_info.set("No file selected")
        self.plan_file_info.set("No file selected")
        self.error_message.set("")
        self.history_load_id_col.set("")
        self.history_qty_col.set("")
        self.plan_load_id_col.set("")
        self.plan_qty_col.set("")
        self.history_split_column.set("")
        self.plan_split_column.set("")
        self.history_split_enabled.set(False)
        self.plan_split_enabled.set(False)
        self.plan_divide_enabled.set(False)
        self.plan_divide_value.set("1")
        self.plan_divide_entry['state'] = 'disabled'
        self.history_split_widths = ""
        self.plan_split_widths = ""
        self.history_original_data = None
        self.plan_original_data = None
        self.history_concat_configs = []
        self.plan_concat_configs = []
        self.history_df_loaded = None
        self.plan_df_loaded = None
        self.history_load_id_combo['values'] = []
        self.history_qty_combo['values'] = []
        self.plan_load_id_combo['values'] = []
        self.plan_qty_combo['values'] = []
        self.history_split_combo['values'] = []
        self.plan_split_combo['values'] = []
        self.history_split_combo['state'] = 'disabled'
        self.plan_split_combo['state'] = 'disabled'
        self.history_concat_button['state'] = 'disabled'
        self.plan_concat_button['state'] = 'disabled'
        self.output_folder.set(os.path.join(os.path.expanduser("~"), "Desktop"))
        self.archive_folder.set("C:\\Reconciliation\\Archive")
        self.progress_var.set(0)
        self.status_var.set("Ready")
        self.run_button.config(state=tk.NORMAL)

        if self.history_tree is not None:
            self.history_tree.delete(*self.history_tree.get_children())
            self.history_tree['columns'] = ()
        if self.plan_tree is not None:
            self.plan_tree.delete(*self.plan_tree.get_children())
            self.plan_tree['columns'] = ()
        self.history_preview_count.set("No data loaded")
        self.plan_preview_count.set("No data loaded")
        self.preview_status_var.set("👆 Click a row in either preview below to see its values here")

        self.clear_log()
        self.add_log("🧹 All fields cleared", "INFO")
    
    def update_progress(self, value, status):
        self.progress_var.set(value)
        self.status_var.set(status)
        self.root.update_idletasks()
    
    def safe_read_excel(self, file_path):
        """Safely read Excel file with multiple engine attempts"""
        self.add_log(f"📖 Reading file: {os.path.basename(file_path)}", "DEBUG")
        
        valid, msg = self.validate_excel_file(file_path)
        if not valid:
            raise Exception(f"Invalid file: {msg}")
        
        try:
            df = pd.read_excel(file_path, engine='openpyxl')
            df = self.clean_whole_number_columns(df)
            self.add_log(f"✅ Successfully read {len(df)} rows using openpyxl", "DEBUG")
            return df
        except Exception as e1:
            self.add_log(f"⚠️ openpyxl failed: {e1}", "WARNING")
            try:
                df = pd.read_excel(file_path)
                df = self.clean_whole_number_columns(df)
                self.add_log(f"✅ Successfully read {len(df)} rows using default engine", "DEBUG")
                return df
            except Exception as e2:
                self.add_log(f"❌ Failed to read file: {e2}", "ERROR")
                raise Exception(f"Could not read file: {e2}")
    
    def apply_excel_formatting(self, filepath):
        """Apply full Excel formatting with colors"""
        try:
            self.add_log("🎨 Applying Excel formatting...", "INFO")
            wb = load_workbook(filepath)
            
            header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            
            green_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            orange_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            blue_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
            purple_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
            
            white_font = Font(color="FFFFFF", bold=True)
            black_font = Font(color="000000", bold=True)
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            if 'variance' in wb.sheetnames:
                ws = wb['variance']
                
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=1, column=col)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                
                for row in range(2, ws.max_row + 1):
                    status_cell = ws.cell(row=row, column=5)
                    comments_cell = ws.cell(row=row, column=6)
                    variance_cell = ws.cell(row=row, column=4)
                    
                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row, column=col)
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    if status_cell.value:
                        status = str(status_cell.value).strip()
                        
                        color_map = {
                            'MATCH': (green_fill, white_font),
                            'MINOR VARIANCE': (yellow_fill, black_font),
                            'MODERATE VARIANCE': (orange_fill, black_font),
                            'MAJOR VARIANCE': (red_fill, white_font),
                            'MISSING IN HISTORY': (blue_fill, white_font),
                            'EXTRA IN HISTORY': (purple_fill, white_font)
                        }
                        
                        fill, font = color_map.get(status, (green_fill, white_font))
                        for cell in [status_cell, comments_cell, variance_cell]:
                            cell.fill = fill
                            cell.font = font
                
                for col in range(1, ws.max_column + 1):
                    max_length = 0
                    for row in range(1, ws.max_row + 1):
                        cell_value = ws.cell(row=row, column=col).value
                        if cell_value:
                            max_length = max(max_length, len(str(cell_value)))
                    adjusted_width = min(max_length + 3, 50)
                    ws.column_dimensions[get_column_letter(col)].width = adjusted_width
                
                ws.row_dimensions[1].height = 30
                for row in range(2, ws.max_row + 1):
                    ws.row_dimensions[row].height = 25
            
            for sheet_name in ['Loading_History_Original', 'Load_Plan_Original', 'Summary']:
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    
                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row=1, column=col)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = thin_border
                    
                    for col in range(1, ws.max_column + 1):
                        max_length = 0
                        for row in range(1, ws.max_row + 1):
                            cell_value = ws.cell(row=row, column=col).value
                            if cell_value:
                                max_length = max(max_length, len(str(cell_value)))
                        adjusted_width = min(max_length + 3, 50)
                        ws.column_dimensions[get_column_letter(col)].width = adjusted_width
                    
                    for row in range(1, ws.max_row + 1):
                        for col in range(1, ws.max_column + 1):
                            cell = ws.cell(row=row, column=col)
                            cell.border = thin_border
                            if row > 1:
                                cell.alignment = Alignment(horizontal='left', vertical='center')
            
            wb.save(filepath)
            self.add_log("✅ Excel formatting applied successfully", "SUCCESS")
            
        except Exception as e:
            self.add_log(f"⚠️ Formatting warning: {e}", "WARNING")
    
    def run_reconciliation(self):
        # Validate files
        if not self.history_file_path.get():
            self.error_message.set("⚠️ Please select a Loading History file")
            messagebox.showerror("Error", "Please select a Loading History file")
            return
        
        if not self.plan_file_path.get():
            self.error_message.set("⚠️ Please select a Load Plan file")
            messagebox.showerror("Error", "Please select a Load Plan file")
            return
        
        # Validate column selections
        if not self.history_load_id_col.get():
            self.error_message.set("⚠️ Please select a Load ID column for History file")
            messagebox.showerror("Error", "Please select a Load ID column for the History file")
            return
        
        if not self.history_qty_col.get():
            self.error_message.set("⚠️ Please select a Quantity column for History file")
            messagebox.showerror("Error", "Please select a Quantity column for the History file")
            return
        
        if not self.plan_load_id_col.get():
            self.error_message.set("⚠️ Please select a Load ID column for Plan file")
            messagebox.showerror("Error", "Please select a Load ID column for the Plan file")
            return
        
        if not self.plan_qty_col.get():
            self.error_message.set("⚠️ Please select a Quantity column for Plan file")
            messagebox.showerror("Error", "Please select a Quantity column for the Plan file")
            return
        
        # Validate text-to-columns
        if self.history_split_enabled.get() and not self.history_split_column.get():
            self.error_message.set("⚠️ Please select a column to split for History")
            messagebox.showerror("Error", "Please select a column to split for the History file")
            return
        
        if self.plan_split_enabled.get() and not self.plan_split_column.get():
            self.error_message.set("⚠️ Please select a column to split for Plan")
            messagebox.showerror("Error", "Please select a column to split for the Plan file")
            return
        
        # Validate divide-by value
        if self.plan_divide_enabled.get():
            try:
                divide_value = float(self.plan_divide_value.get())
            except ValueError:
                self.error_message.set("⚠️ Please enter a valid number to divide Plan Qty by")
                messagebox.showerror("Error", "Please enter a valid number to divide the Plan Quantity by")
                return
            if divide_value == 0:
                self.error_message.set("⚠️ Divide value cannot be zero")
                messagebox.showerror("Error", "The number to divide Plan Qty by cannot be zero")
                return
        
        for file_path, file_type in [
            (self.history_file_path.get(), "History"),
            (self.plan_file_path.get(), "Plan")
        ]:
            valid, msg = self.validate_excel_file(file_path)
            if not valid:
                self.error_message.set(f"⚠️ {file_type} file: {msg}")
                messagebox.showerror("Error", f"Invalid {file_type} file:\n{msg}")
                return
        
        self.error_message.set("")
        
        try:
            import subprocess
            subprocess.call(['taskkill', '/F', '/IM', 'EXCEL.EXE'], 
                          stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            self.add_log("📌 Closed any open Excel instances", "DEBUG")
        except:
            pass
        
        self.run_button.config(state=tk.DISABLED)
        self.add_log("🚀 Starting reconciliation process...", "INFO")
        thread = threading.Thread(target=self.process_reconciliation)
        thread.daemon = True
        thread.start()
    
    def process_reconciliation(self):
        try:
            history_path = self.history_file_path.get()
            plan_path = self.plan_file_path.get()
            
            history_load_id_col = self.history_load_id_col.get()
            history_qty_col = self.history_qty_col.get()
            plan_load_id_col = self.plan_load_id_col.get()
            plan_qty_col = self.plan_qty_col.get()
            
            self.update_progress(5, "📖 Reading files...")
            self.add_log("📖 Reading input files...", "INFO")
            
            history_df = self.safe_read_excel(history_path)
            plan_df = self.safe_read_excel(plan_path)
            
            self.add_log(f"📊 History: {len(history_df)} rows, {len(history_df.columns)} columns", "INFO")
            self.add_log(f"📊 Plan: {len(plan_df)} rows, {len(plan_df.columns)} columns", "INFO")
            
            # Reapply any Concatenate Columns operations configured earlier -
            # the file was just re-read from disk, so those columns don't exist yet
            if self.history_concat_configs:
                for cfg in self.history_concat_configs:
                    missing = [c for c in cfg['columns'] if c not in history_df.columns]
                    if missing:
                        self.add_log(
                            f"⚠️ Could not reapply History concatenation '{cfg['new_column_name']}' - "
                            f"missing source column(s): {', '.join(missing)}", "WARNING"
                        )
                        continue
                    history_df = self.apply_column_concatenation(
                        history_df, cfg['columns'], cfg['separator'], cfg['new_column_name'],
                        cfg['drop_originals'], "History"
                    )
                new_columns = history_df.columns.tolist()
                self.history_load_id_combo['values'] = new_columns
                self.history_qty_combo['values'] = new_columns
                self.history_split_combo['values'] = new_columns
            
            if self.plan_concat_configs:
                for cfg in self.plan_concat_configs:
                    missing = [c for c in cfg['columns'] if c not in plan_df.columns]
                    if missing:
                        self.add_log(
                            f"⚠️ Could not reapply Plan concatenation '{cfg['new_column_name']}' - "
                            f"missing source column(s): {', '.join(missing)}", "WARNING"
                        )
                        continue
                    plan_df = self.apply_column_concatenation(
                        plan_df, cfg['columns'], cfg['separator'], cfg['new_column_name'],
                        cfg['drop_originals'], "Plan"
                    )
                new_columns = plan_df.columns.tolist()
                self.plan_load_id_combo['values'] = new_columns
                self.plan_qty_combo['values'] = new_columns
                self.plan_split_combo['values'] = new_columns
            
            # Apply text-to-columns if enabled
            if self.history_split_enabled.get() and self.history_original_data is not None:
                history_df = self.apply_text_to_columns(
                    history_df,
                    self.history_split_column.get(),
                    self.history_original_data,
                    "History"
                )
                
                new_columns = history_df.columns.tolist()
                self.history_load_id_combo['values'] = new_columns
                self.history_qty_combo['values'] = new_columns
                self.history_split_combo['values'] = new_columns
                
                if history_load_id_col not in new_columns:
                    self.add_log(f"⚠️ Column '{history_load_id_col}' was removed by split. Please re-select.", "WARNING")
                    self.history_load_id_col.set("")
                    history_load_id_col = ""
                if history_qty_col not in new_columns:
                    self.add_log(f"⚠️ Column '{history_qty_col}' was removed by split. Please re-select.", "WARNING")
                    self.history_qty_col.set("")
                    history_qty_col = ""
            
            if self.plan_split_enabled.get() and self.plan_original_data is not None:
                plan_df = self.apply_text_to_columns(
                    plan_df,
                    self.plan_split_column.get(),
                    self.plan_original_data,
                    "Plan"
                )
                
                new_columns = plan_df.columns.tolist()
                self.plan_load_id_combo['values'] = new_columns
                self.plan_qty_combo['values'] = new_columns
                self.plan_split_combo['values'] = new_columns
                
                if plan_load_id_col not in new_columns:
                    self.add_log(f"⚠️ Column '{plan_load_id_col}' was removed by split. Please re-select.", "WARNING")
                    self.plan_load_id_col.set("")
                    plan_load_id_col = ""
                if plan_qty_col not in new_columns:
                    self.add_log(f"⚠️ Column '{plan_qty_col}' was removed by split. Please re-select.", "WARNING")
                    self.plan_qty_col.set("")
                    plan_qty_col = ""
            
            if not history_load_id_col or not history_qty_col:
                self.error_message.set("⚠️ Please re-select columns after text-to-columns split")
                messagebox.showerror("Error", "Please re-select the Load ID and Quantity columns after applying text-to-columns split.")
                return
            
            if not plan_load_id_col or not plan_qty_col:
                self.error_message.set("⚠️ Please re-select columns after text-to-columns split")
                messagebox.showerror("Error", "Please re-select the Load ID and Quantity columns after applying text-to-columns split.")
                return
            
            missing_history_cols = [c for c in (history_load_id_col, history_qty_col) if c not in history_df.columns]
            missing_plan_cols = [c for c in (plan_load_id_col, plan_qty_col) if c not in plan_df.columns]
            if missing_history_cols or missing_plan_cols:
                detail = []
                if missing_history_cols:
                    detail.append(f"History file is missing: {', '.join(missing_history_cols)}")
                if missing_plan_cols:
                    detail.append(f"Plan file is missing: {', '.join(missing_plan_cols)}")
                message = (
                    "The selected Load ID/Quantity column(s) could not be found in the file.\n\n"
                    + "\n".join(detail)
                    + "\n\nIf you used Concatenate Columns or Text-to-Columns, please re-check that "
                      "configuration and re-select the Load ID / Quantity columns."
                )
                self.error_message.set("⚠️ Selected column(s) not found - see error for details")
                self.add_log(f"❌ Missing column(s) before processing: {', '.join(missing_history_cols + missing_plan_cols)}", "ERROR")
                messagebox.showerror("Error", message)
                return
            
            self.update_progress(20, "🔄 Processing data...")
            
            history_df.columns = history_df.columns.str.strip()
            plan_df.columns = plan_df.columns.str.strip()
            
            self.add_log(f"📌 History - Load ID: '{history_load_id_col}', Quantity: '{history_qty_col}'", "INFO")
            self.add_log(f"📌 Plan - Load ID: '{plan_load_id_col}', Quantity: '{plan_qty_col}'", "INFO")
            
            # Convert Load ID columns to string for consistent merging
            self.add_log("🔄 Converting Load ID columns to string type...", "INFO")
            
            if history_load_id_col in history_df.columns:
                history_df[history_load_id_col] = history_df[history_load_id_col].astype(str).str.strip()
                history_df[history_load_id_col] = history_df[history_load_id_col].replace('nan', '')
                history_df[history_load_id_col] = history_df[history_load_id_col].replace('None', '')
            
            if plan_load_id_col in plan_df.columns:
                plan_df[plan_load_id_col] = plan_df[plan_load_id_col].astype(str).str.strip()
                plan_df[plan_load_id_col] = plan_df[plan_load_id_col].replace('nan', '')
                plan_df[plan_load_id_col] = plan_df[plan_load_id_col].replace('None', '')
            
            self.add_log("✅ Load ID columns converted to string type", "INFO")
            
            self.update_progress(30, "📊 Cleaning and grouping data...")
            
            history_df[history_qty_col] = pd.to_numeric(history_df[history_qty_col], errors='coerce').fillna(0)
            plan_df[plan_qty_col] = pd.to_numeric(plan_df[plan_qty_col], errors='coerce').fillna(0)
            
            plan_divide_value = None
            if self.plan_divide_enabled.get():
                plan_divide_value = float(self.plan_divide_value.get())
                plan_df[plan_qty_col] = plan_df[plan_qty_col] / plan_divide_value
                self.add_log(f"➗ Dividing Plan Qty ('{plan_qty_col}') by {plan_divide_value}", "INFO")
            
            history_grouped = history_df.groupby(history_load_id_col, as_index=False)[history_qty_col].sum()
            history_grouped.columns = ['Load ID', 'Total Loaded Qty']
            
            plan_grouped = plan_df.groupby(plan_load_id_col, as_index=False)[plan_qty_col].sum()
            plan_grouped.columns = ['Load ID', 'Total Plan Qty']
            
            self.add_log(f"📊 Unique Load IDs in Plan: {len(plan_grouped)}", "INFO")
            self.add_log(f"📊 Unique Load IDs in History: {len(history_grouped)}", "INFO")
            
            self.update_progress(50, "🔍 Finding variances...")
            
            merged_df = pd.merge(
                plan_grouped,
                history_grouped,
                on='Load ID',
                how='left'
            ).fillna(0)
            
            merged_df.columns = ['Load ID', 'Plan Qty', 'Loaded Qty']
            merged_df['Variance'] = merged_df['Plan Qty'] - merged_df['Loaded Qty']
            
            history_only = history_grouped[~history_grouped['Load ID'].isin(plan_grouped['Load ID'])].copy()
            if not history_only.empty:
                history_only['Plan Qty'] = 0
                history_only['Variance'] = -history_only['Total Loaded Qty']
                history_only.columns = ['Load ID', 'Loaded Qty', 'Plan Qty', 'Variance']
                merged_df = pd.concat([merged_df, history_only], ignore_index=True)
                self.add_log(f"⚠️ Found {len(history_only)} extra loads in History", "WARNING")
            
            def get_status(row):
                variance = row['Variance']
                plan_qty = row['Plan Qty']
                loaded_qty = row['Loaded Qty']
                
                if plan_qty == 0 and loaded_qty > 0:
                    return 'EXTRA IN HISTORY'
                elif plan_qty > 0 and loaded_qty == 0:
                    return 'MISSING IN HISTORY'
                elif abs(variance) < 0.01:
                    return 'MATCH'
                elif abs(variance) <= 100:
                    return 'MINOR VARIANCE'
                elif abs(variance) <= 1000:
                    return 'MODERATE VARIANCE'
                else:
                    return 'MAJOR VARIANCE'
            
            def get_comments(row):
                status = row['Status']
                variance = row['Variance']
                comments_map = {
                    'MATCH': 'Exact match ✅',
                    'EXTRA IN HISTORY': f'Load in History but not in Plan (variance: {variance:.2f}) ⚠️',
                    'MISSING IN HISTORY': f'Load in Plan but not in History (variance: {variance:.2f}) ⚠️',
                    'MINOR VARIANCE': f'Minor variance - check quantities (variance: {variance:.2f}) 🔶',
                    'MODERATE VARIANCE': f'Moderate variance - investigate (variance: {variance:.2f}) 🟠',
                    'MAJOR VARIANCE': f'LARGE VARIANCE - URGENT! (variance: {variance:.2f}) 🔴'
                }
                return comments_map.get(status, status)
            
            merged_df['Status'] = merged_df.apply(get_status, axis=1)
            merged_df['Comments'] = merged_df.apply(get_comments, axis=1)
            
            result_df = merged_df[[
                'Load ID',
                'Plan Qty',
                'Loaded Qty',
                'Variance',
                'Status',
                'Comments'
            ]]
            
            try:
                result_df['Load ID'] = pd.to_numeric(result_df['Load ID'])
                result_df = result_df.sort_values('Load ID').reset_index(drop=True)
            except:
                result_df = result_df.sort_values('Load ID').reset_index(drop=True)
            
            self.update_progress(70, "💾 Saving results...")
            
            output_path = self.output_folder.get()
            if not os.path.exists(output_path):
                os.makedirs(output_path)
                self.add_log(f"📁 Created output folder: {output_path}", "INFO")
            
            timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
            output_file = os.path.join(output_path, f"Reconciliations_{timestamp}.xlsx")
            
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                result_df.to_excel(writer, sheet_name='variance', index=False)
                history_df.to_excel(writer, sheet_name='Loading_History_Original', index=False)
                plan_df.to_excel(writer, sheet_name='Load_Plan_Original', index=False)
                
                match_count = len(result_df[result_df['Status'] == 'MATCH'])
                variance_count = len(result_df[result_df['Status'].str.contains('VARIANCE', case=False)])
                missing_count = len(result_df[result_df['Status'] == 'MISSING IN HISTORY'])
                extra_count = len(result_df[result_df['Status'] == 'EXTRA IN HISTORY'])
                match_rate = round((match_count / len(plan_grouped)) * 100 if len(plan_grouped) > 0 else 0, 2)
                
                summary_data = {
                    'Metric': [
                        'Total Loads in Plan (unique)',
                        'Total Loads in History (unique)',
                        'Exact Matches',
                        'Variances Found',
                        'Missing in History',
                        'Extra in History',
                        'Match Rate',
                        'Processing Date',
                        'History Load ID Column',
                        'History Quantity Column',
                        'Plan Load ID Column',
                        'Plan Quantity Column'
                    ],
                    'Value': [
                        len(plan_grouped),
                        len(history_grouped),
                        match_count,
                        variance_count,
                        missing_count,
                        extra_count,
                        f"{match_rate}%",
                        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        history_load_id_col,
                        history_qty_col,
                        plan_load_id_col,
                        plan_qty_col
                    ]
                }
                
                if self.history_split_enabled.get():
                    summary_data['Metric'].append('History Text-to-Columns')
                    summary_data['Value'].append(f"Column: {self.history_split_column.get()}, Widths: {self.history_split_widths}")
                
                if self.plan_split_enabled.get():
                    summary_data['Metric'].append('Plan Text-to-Columns')
                    summary_data['Value'].append(f"Column: {self.plan_split_column.get()}, Widths: {self.plan_split_widths}")
                
                if plan_divide_value is not None:
                    summary_data['Metric'].append('Plan Qty Divided By')
                    summary_data['Value'].append(plan_divide_value)
                
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            self.add_log(f"💾 Results saved to: {os.path.basename(output_file)}", "SUCCESS")
            self.last_output_file = output_file
            
            self.update_progress(85, "🎨 Applying formatting...")
            self.apply_excel_formatting(output_file)
            
            self.update_progress(95, "📂 Archiving files...")
            
            if self.archive_var.get():
                archive_path = self.archive_folder.get()
                if not os.path.exists(archive_path):
                    os.makedirs(archive_path)
                    self.add_log(f"📁 Created archive folder: {archive_path}", "INFO")
                
                date_str = datetime.now().strftime("%Y-%m-%d")
                archive_date_path = os.path.join(archive_path, date_str)
                if not os.path.exists(archive_date_path):
                    os.makedirs(archive_date_path)
                
                history_filename = os.path.basename(history_path)
                plan_filename = os.path.basename(plan_path)
                
                shutil.copy2(history_path, os.path.join(archive_date_path, f"{timestamp}_{history_filename}"))
                shutil.copy2(plan_path, os.path.join(archive_date_path, f"{timestamp}_{plan_filename}"))
                self.add_log(f"📦 Archived files to: {archive_date_path}", "INFO")
            
            self.update_progress(100, "✅ Done!")
            
            if self.auto_open_file.get():
                try:
                    os.startfile(output_file)
                    self.add_log("📂 Opened result file", "INFO")
                except:
                    self.add_log("⚠️ Could not auto-open file", "WARNING")
            
            summary = (
                f"✅ Reconciliation Complete!\n\n"
                f"📁 File saved: {output_file}\n\n"
                f"📊 SUMMARY:\n"
                f"  • Unique Loads in Plan: {len(plan_grouped)}\n"
                f"  • Unique Loads in History: {len(history_grouped)}\n"
                f"  • ✅ Exact Matches: {match_count}\n"
                f"  • ⚠️ Variances Found: {variance_count}\n"
                f"  • ❌ Missing in History: {missing_count}\n"
                f"  • ➕ Extra in History: {extra_count}\n"
                f"  • 🎯 Match Rate: {match_rate}%\n\n"
                f"📌 Column Mapping:\n"
                f"  • History Load ID: '{history_load_id_col}'\n"
                f"  • History Quantity: '{history_qty_col}'\n"
                f"  • Plan Load ID: '{plan_load_id_col}'\n"
                f"  • Plan Quantity: '{plan_qty_col}'"
            )
            
            if self.history_split_enabled.get():
                summary += f"\n  • History Text-to-Columns: {self.history_split_column.get()} (widths: {self.history_split_widths})"
            if self.plan_split_enabled.get():
                summary += f"\n  • Plan Text-to-Columns: {self.plan_split_column.get()} (widths: {self.plan_split_widths})"
            if plan_divide_value is not None:
                summary += f"\n  • Plan Qty divided by: {plan_divide_value}"
            
            self.add_log(f"✅ Reconciliation complete! Match rate: {match_rate}%", "SUCCESS")
            messagebox.showinfo("Success", summary)
            
        except Exception as e:
            error_msg = f"An error occurred:\n{str(e)}"
            self.add_log(f"❌ {error_msg}", "ERROR")
            self.error_message.set(f"⚠️ {str(e)}")
            messagebox.showerror("Error", error_msg)
        
        finally:
            self.run_button.config(state=tk.NORMAL)
            self.status_var.set("Ready")

if __name__ == "__main__":
    try:
        import ctypes
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID('efl.nexus.reconciliation')
    except Exception:
        pass

    root = tk.Tk()
    for icon_name in ("icon_2.ico", "icon.ico", "favicon.ico"):
        base_dir = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
        icon_path = os.path.join(base_dir, icon_name)
        if not os.path.exists(icon_path) and getattr(sys, 'frozen', False):
            icon_path = os.path.join(os.path.dirname(sys.executable), icon_name)
        if os.path.exists(icon_path):
            try:
                root.iconbitmap(default=icon_path)
                break
            except Exception:
                try:
                    root.iconbitmap(icon_path)
                    break
                except Exception:
                    pass
    
    app = ReconciliationApp(root)
    root.mainloop()