"""
Outlook Email Sender (GUI) - Google Sheets Cloud Log Version
------------------------------------------------------------
Sends email through the user's real, already-installed classic Outlook
desktop app via COM automation (pywin32) - not a separate SMTP relay - so the
message goes out using the user's actual account, signature, and Sent Items
folder.

Signature handling:
- A plain Tkinter text box cannot hold rich text (fonts, colors, logos), so
  pasting a signature into it would always strip its formatting.
- Instead of pasting anything, this app reads the user's real default
  signature straight off disk (the .htm file Outlook itself maintains under
  %APPDATA%\\Microsoft\\Signatures) and stitches the typed message in above
  it, all as one HTML email body.

Attachments:
- Multiple files can be attached at once (PDF and/or Excel), shown in a
  list, with the ability to remove individual files before sending.

Recipient group templates:
- Each template stores a group name plus its own 'To' and 'CC' recipient
  lists, saved in templates.xlsx (Excel) next to this script.
- A dropdown lets the user pick a saved template, which fills in To/CC.
- "Add New Template" opens a small form to save a new group.

GDN / GRN type selection:
- Two toggle buttons let the user explicitly mark the email being composed
  as a "GDN" or a "GRN" (or neither). Pressing one of these buttons also
  renames the Subject field to "GDN -" / "GRN -" (see _select_type below);
  the previously typed subject is remembered and restored if the type is
  cleared or switched back off.

Send tracking (GDN/GRN log):
- Whenever an email is sent with GDN or GRN selected, log data is sent
  silently over HTTP to a Google Apps Script Web App attached to an online
  Google Sheet.
- A running total ("GDN sent: N   GRN sent: N") is shown in the top right,
  fetched live from the Google Sheet via the Web App.

Notes:
- Requires classic ("old") Outlook for Windows.
- Requires pywin32 (`pip install pywin32`).
- Requires openpyxl for templates (`pip install openpyxl`).
- Requires requests for cloud logging (`pip install requests`).

Run:
    python outlook_email_gui.py
"""

import os
import re
import sys
import html
import json
import datetime
import threading
import requests
import webbrowser
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

try:
    import win32com.client
    WIN32COM_AVAILABLE = True
except ImportError:
    WIN32COM_AVAILABLE = False

try:
    import winreg
    WINREG_AVAILABLE = True
except ImportError:
    WINREG_AVAILABLE = False

try:
    from openpyxl import Workbook, load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# --------------------------------------------------------------------------
# Shared visual style
# --------------------------------------------------------------------------

BG_COLOR = "#f4f6f9"
PANEL_COLOR = "#ffffff"
ACCENT_COLOR = "#2f6feb"
ACCENT_COLOR_DARK = "#255ac2"
TEXT_COLOR = "#1f2937"
MUTED_COLOR = "#6b7280"
DANGER_COLOR = "#d64545"
FONT_FAMILY = "Segoe UI"
FONT_NORMAL = (FONT_FAMILY, 10)
FONT_BOLD = (FONT_FAMILY, 10, "bold")
FONT_HEADER = (FONT_FAMILY, 14, "bold")

OUTLOOK_MAIL_ITEM = 0  # olMailItem


def _get_base_dir():
    """Where templates.xlsx and config.json should live."""
    if getattr(sys, "frozen", False):
        return os.path.dirname(os.path.abspath(sys.executable))
    return os.path.dirname(os.path.abspath(__file__))


BASE_DIR = _get_base_dir()
TEMPLATES_XLSX_PATH = os.path.join(BASE_DIR, "templates.xlsx")
CONFIG_JSON_PATH = os.path.join(BASE_DIR, "config.json")

TEMPLATE_PLACEHOLDER = "-- Select Template --"
DEFAULT_BODY_TEXT = "Dear all,\n\nFYI."

SEND_TYPE_GDN = "GDN"
SEND_TYPE_GRN = "GRN"

ATTACHMENT_FILETYPES = [
    ("PDF and Excel files", "*.pdf;*.xlsx;*.xls"),
    ("PDF files", "*.pdf"),
    ("Excel files", "*.xlsx;*.xls"),
    ("All files", "*.*"),
]

SIGNATURE_REGISTRY_PATHS = [
    r"Software\Microsoft\Office\16.0\Common\MailSettings",
    r"Software\Microsoft\Office\15.0\Common\MailSettings",
]


def style_button(btn, kind="primary"):
    colors = {
        "primary": (ACCENT_COLOR, "#ffffff", ACCENT_COLOR_DARK),
        "danger": (DANGER_COLOR, "#ffffff", "#b23a3a"),
        "secondary": ("#e5e7eb", TEXT_COLOR, "#d1d5db"),
    }
    bg, fg, active_bg = colors.get(kind, colors["primary"])
    btn.configure(
        bg=bg, fg=fg, activebackground=active_bg, activeforeground=fg,
        font=FONT_BOLD, relief="flat", bd=0, padx=14, pady=6, cursor="hand2",
    )


# --------------------------------------------------------------------------
# Config Store (JSON-based settings)
# --------------------------------------------------------------------------

class ConfigStore:
    """Handles saving and loading user-defined Google Sheets & Web App URLs."""

    def __init__(self, path):
        self.path = path
        self.config = self.load()

    def load(self):
        if os.path.exists(self.path):
            try:
                with open(self.path, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception:
                pass
        return {"webapp_url": "", "sheet_url": ""}

    def save(self, webapp_url, sheet_url):
        self.config["webapp_url"] = webapp_url.strip()
        self.config["sheet_url"] = sheet_url.strip()
        try:
            with open(self.path, "w", encoding="utf-8") as f:
                json.dump(self.config, f, indent=4)
        except Exception as e:
            print(f"Failed to save config: {e}")

    def get_webapp_url(self):
        return self.config.get("webapp_url", "")

    def get_sheet_url(self):
        return self.config.get("sheet_url", "")


# --------------------------------------------------------------------------
# Recipient helpers
# --------------------------------------------------------------------------

def normalize_recipients(raw_text):
    """Outlook expects recipients separated by semicolons."""
    if not raw_text:
        return ""
    text = raw_text.replace(",", ";")
    parts = [p.strip() for p in text.split(";")]
    parts = [p for p in parts if p]
    return "; ".join(parts)


# --------------------------------------------------------------------------
# Default signature (read directly from disk)
# --------------------------------------------------------------------------

def _fix_signature_image_paths(html_text, files_dir):
    if not files_dir or not os.path.isdir(files_dir):
        return html_text

    def repl(match):
        attr, quote, src = match.group(1), match.group(2), match.group(3)
        if src.lower().startswith(("http://", "https://", "file://", "cid:", "data:")):
            return match.group(0)
        abs_path = os.path.join(files_dir, src)
        if os.path.isfile(abs_path):
            uri = "file:///" + abs_path.replace(os.sep, "/")
            return f'{attr}={quote}{uri}{quote}'
        return match.group(0)

    return re.sub(r'(src)=(["\'])([^"\']+)\2', repl, html_text, flags=re.IGNORECASE)


def get_default_signature_html():
    appdata = os.environ.get("APPDATA")
    if not appdata:
        return ""
    sig_dir = os.path.join(appdata, "Microsoft", "Signatures")
    if not os.path.isdir(sig_dir):
        return ""

    sig_name = None
    if WINREG_AVAILABLE:
        for path in SIGNATURE_REGISTRY_PATHS:
            try:
                with winreg.OpenKey(winreg.HKEY_CURRENT_USER, path) as key:
                    value, _ = winreg.QueryValueEx(key, "NewSignature")
                    if value:
                        sig_name = value
                        break
            except OSError:
                continue

    if not sig_name:
        try:
            htm_files = [f for f in os.listdir(sig_dir) if f.lower().endswith(".htm")]
        except OSError:
            return ""
        if not htm_files:
            return ""
        htm_files.sort(
            key=lambda f: os.path.getmtime(os.path.join(sig_dir, f)), reverse=True)
        sig_name = os.path.splitext(htm_files[0])[0]

    sig_path = os.path.join(sig_dir, f"{sig_name}.htm")
    if not os.path.isfile(sig_path):
        return ""

    try:
        with open(sig_path, "r", encoding="utf-8", errors="ignore") as f:
            signature_html = f.read()
    except OSError:
        return ""

    files_dir = os.path.join(sig_dir, f"{sig_name}_files")
    return _fix_signature_image_paths(signature_html, files_dir)


# --------------------------------------------------------------------------
# Recipient-group template storage (Excel)
# --------------------------------------------------------------------------

class TemplateStore:
    def __init__(self, path):
        self.path = path
        self.templates = self.load()

    def load(self):
        if not OPENPYXL_AVAILABLE or not os.path.exists(self.path):
            return []
        try:
            wb = load_workbook(self.path)
            ws = wb.active
            templates = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                name = str(row[0]).strip()
                to = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                cc = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                templates.append({"name": name, "to": to, "cc": cc})
            return templates
        except OSError:
            return []

    def save(self):
        if not OPENPYXL_AVAILABLE:
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "Templates"
        ws.append(["Template Name", "To", "CC"])
        for t in self.templates:
            ws.append([t["name"], t["to"], t["cc"]])
        for col, width in zip("ABC", (24, 50, 50)):
            ws.column_dimensions[col].width = width
        wb.save(self.path)

    def get_names(self):
        return [t["name"] for t in self.templates]

    def get_template(self, name):
        for t in self.templates:
            if t["name"] == name:
                return t
        return None

    def add_or_update_template(self, name, to, cc):
        existing = self.get_template(name)
        if existing:
            existing["to"] = to
            existing["cc"] = cc
        else:
            self.templates.append({"name": name, "to": to, "cc": cc})
        self.save()


# --------------------------------------------------------------------------
# Sent-mail tracking (Google Sheets Web App Integration)
# --------------------------------------------------------------------------

class SentLogStore:
    """Sends log data to a Google Apps Script Web App."""

    def __init__(self, config_store):
        self.config_store = config_store

    @staticmethod
    def _split_subject(send_type, subject_text):
        text = (subject_text or "").strip()
        marker = f"{send_type} -"
        idx = text.find(marker)
        if idx == -1:
            return text, send_type
        prefix_end = idx + len(marker)
        if prefix_end < len(text) and text[prefix_end] == " ":
            prefix_end += 1
        number_part = text[prefix_end:].strip()
        type_part = text[:prefix_end].rstrip()
        if type_part.endswith("-"):
            type_part = type_part[:-1].rstrip()
        return number_part, type_part

    def log_send(self, send_type, subject_text):
        url = self.config_store.get_webapp_url()
        if not url or send_type not in ["GDN", "GRN"]:
            return

        try:
            now = datetime.datetime.now()
            number_part, type_part = self._split_subject(send_type, subject_text)

            payload = {
                "type": send_type,
                "date": now.strftime("%Y-%m-%d"),
                "time": now.strftime("%H:%M:%S"),
                "number": number_part,
                "type_part": type_part
            }

            requests.post(url, json=payload, timeout=10)
        except Exception as e:
            print(f"Logging failed: {e}")

    def get_counts(self):
        url = self.config_store.get_webapp_url()
        if not url:
            return {"GDN": 0, "GRN": 0}

        try:
            response = requests.get(url, timeout=5)
            if response.status_code == 200:
                return response.json()
        except Exception:
            pass
        return {"GDN": 0, "GRN": 0}


# --------------------------------------------------------------------------
# Outlook automation
# --------------------------------------------------------------------------

class OutlookClient:
    def __init__(self):
        self._app = None

    def _get_app(self):
        if self._app is None:
            if not WIN32COM_AVAILABLE:
                raise RuntimeError(
                    "pywin32 is not installed. Install it with:\n\n"
                    "pip install pywin32")
            self._app = win32com.client.Dispatch("Outlook.Application")
        return self._app

    def build_mail(self, to_addr, cc_addr, subject, body_text,
                    attachment_paths, include_signature=True):
        app = self._get_app()
        mail = app.CreateItem(OUTLOOK_MAIL_ITEM)

        try:
            mail.To = to_addr
            if cc_addr:
                mail.CC = cc_addr
            mail.Subject = subject
        except Exception as exc:
            raise RuntimeError(f"Could not set recipients/subject: {exc}")

        signature_html = ""
        if include_signature:
            try:
                signature_html = get_default_signature_html()
            except Exception:
                signature_html = ""

        try:
            mail.HTMLBody = self._compose_html(body_text, signature_html)
        except Exception as exc:
            raise RuntimeError(f"Could not set the email body: {exc}")

        for path in attachment_paths:
            if not os.path.isfile(path):
                raise RuntimeError(f"Attachment file not found on disk: {path}")
            try:
                mail.Attachments.Add(path)
            except Exception as exc:
                raise RuntimeError(
                    f"Could not attach '{os.path.basename(path)}': {exc}")

        return mail

    @staticmethod
    def _compose_html(body_text, signature_html):
        message_html = (
            '<div style="font-family:Calibri,Arial,sans-serif;'
            'font-size:11pt;color:#000000;">'
            + html.escape(body_text).replace("\n", "<br>")
            + "</div><br>"
        )
        if signature_html:
            match = re.search(r"(<body[^>]*>)", signature_html, re.IGNORECASE)
            if match:
                insertion_point = match.end()
                return (signature_html[:insertion_point] + message_html
                         + signature_html[insertion_point:])
            return message_html + signature_html
        return f"<html><body>{message_html}</body></html>"

    def preview(self, mail):
        mail.Display(True)

    def send(self, mail):
        mail.Send()


# --------------------------------------------------------------------------
# Settings Dialog
# --------------------------------------------------------------------------

class SettingsDialog(tk.Toplevel):
    def __init__(self, master, config_store, on_save):
        super().__init__(master)
        self.config_store = config_store
        self.on_save = on_save
        self.title("Settings")
        self.configure(bg=BG_COLOR)
        self.resizable(False, False)
        self.transient(master)
        self.grab_set()
        self._build_ui()
        self._center_window(master)

    def _center_window(self, master):
        self.update_idletasks()
        master.update_idletasks()
        mw = master.winfo_width() or 800
        mh = master.winfo_height() or 600
        mx = master.winfo_rootx()
        my = master.winfo_rooty()
        dw = self.winfo_reqwidth()
        dh = self.winfo_reqheight()
        x = max(0, mx + (mw - dw) // 2)
        y = max(0, my + (mh - dh) // 2)
        self.geometry(f"+{x}+{y}")

    def _build_ui(self):
        tk.Label(self, text="Google Sheets Settings", font=FONT_HEADER, bg=BG_COLOR,
                  fg=TEXT_COLOR).grid(row=0, column=0, columnspan=2,
                                       padx=20, pady=(16, 10), sticky="w")

        tk.Label(self, text="Web App URL:", font=FONT_NORMAL, bg=BG_COLOR,
                  fg=TEXT_COLOR).grid(row=1, column=0, sticky="w", padx=20, pady=6)
        self.webapp_entry = tk.Entry(self, font=FONT_NORMAL, width=50)
        self.webapp_entry.insert(0, self.config_store.get_webapp_url())
        self.webapp_entry.grid(row=1, column=1, padx=20, pady=6)

        tk.Label(self, text="Google Sheet URL:", font=FONT_NORMAL, bg=BG_COLOR,
                  fg=TEXT_COLOR).grid(row=2, column=0, sticky="w", padx=20, pady=6)
        self.sheet_entry = tk.Entry(self, font=FONT_NORMAL, width=50)
        self.sheet_entry.insert(0, self.config_store.get_sheet_url())
        self.sheet_entry.grid(row=2, column=1, padx=20, pady=6)

        btn_frame = tk.Frame(self, bg=BG_COLOR)
        btn_frame.grid(row=3, column=0, columnspan=2, pady=16)

        save_btn = tk.Button(btn_frame, text="Save Settings", command=self._save)
        style_button(save_btn, "primary")
        save_btn.pack(side="left", padx=6)

        cancel_btn = tk.Button(btn_frame, text="Cancel", command=self.destroy)
        style_button(cancel_btn, "secondary")
        cancel_btn.pack(side="left", padx=6)

    def _save(self):
        webapp_url = self.webapp_entry.get().strip()
        sheet_url = self.sheet_entry.get().strip()

        self.config_store.save(webapp_url, sheet_url)
        self.on_save()
        messagebox.showinfo("Saved", "Settings saved successfully.")
        self.destroy()


# --------------------------------------------------------------------------
# Add New Template dialog
# --------------------------------------------------------------------------

class AddTemplateDialog(tk.Toplevel):
    def __init__(self, master, on_submit):
        super().__init__(master)
        self.on_submit = on_submit
        self.title("Add New Template")
        self.configure(bg=BG_COLOR)
        self.resizable(False, False)
        self.transient(master)
        self.grab_set()
        self._build_ui()
        self._center_window(master)

    def _center_window(self, master):
        self.update_idletasks()
        master.update_idletasks()
        mw = master.winfo_width() or 800
        mh = master.winfo_height() or 600
        mx = master.winfo_rootx()
        my = master.winfo_rooty()
        dw = self.winfo_reqwidth()
        dh = self.winfo_reqheight()
        x = max(0, mx + (mw - dw) // 2)
        y = max(0, my + (mh - dh) // 2)
        self.geometry(f"+{x}+{y}")

    def _build_ui(self):
        tk.Label(self, text="Add New Template", font=FONT_HEADER, bg=BG_COLOR,
                  fg=TEXT_COLOR).grid(row=0, column=0, columnspan=2,
                                       padx=20, pady=(16, 10), sticky="w")

        tk.Label(self, text="Template name:", font=FONT_NORMAL, bg=BG_COLOR,
                  fg=TEXT_COLOR).grid(row=1, column=0, sticky="w", padx=20, pady=6)
        self.name_entry = tk.Entry(self, font=FONT_NORMAL, width=40)
        self.name_entry.grid(row=1, column=1, padx=20, pady=6)

        tk.Label(self, text="To:", font=FONT_NORMAL, bg=BG_COLOR,
                  fg=TEXT_COLOR).grid(row=2, column=0, sticky="w", padx=20, pady=6)
        self.to_entry = tk.Entry(self, font=FONT_NORMAL, width=40)
        self.to_entry.grid(row=2, column=1, padx=20, pady=6)

        tk.Label(self, text="CC:", font=FONT_NORMAL, bg=BG_COLOR,
                  fg=TEXT_COLOR).grid(row=3, column=0, sticky="w", padx=20, pady=6)
        self.cc_entry = tk.Entry(self, font=FONT_NORMAL, width=40)
        self.cc_entry.grid(row=3, column=1, padx=20, pady=6)

        tk.Label(self, text="(separate multiple recipients with , or ;)",
                  font=(FONT_FAMILY, 8), bg=BG_COLOR, fg=MUTED_COLOR).grid(
            row=4, column=1, sticky="w", padx=20)

        btn_frame = tk.Frame(self, bg=BG_COLOR)
        btn_frame.grid(row=5, column=0, columnspan=2, pady=16)

        save_btn = tk.Button(btn_frame, text="Save Template", command=self._save)
        style_button(save_btn, "primary")
        save_btn.pack(side="left", padx=6)

        cancel_btn = tk.Button(btn_frame, text="Cancel", command=self.destroy)
        style_button(cancel_btn, "secondary")
        cancel_btn.pack(side="left", padx=6)

    def _save(self):
        name = self.name_entry.get().strip()
        to_addr = normalize_recipients(self.to_entry.get().strip())
        cc_addr = normalize_recipients(self.cc_entry.get().strip())

        if not name:
            messagebox.showwarning("Missing name", "Please enter a template name.")
            return
        if not to_addr:
            messagebox.showwarning("Missing recipients",
                                    "Please enter at least one 'To:' recipient.")
            return

        self.on_submit(name, to_addr, cc_addr)
        self.destroy()


# --------------------------------------------------------------------------
# Main application window
# --------------------------------------------------------------------------

class OutlookEmailApp:
    def __init__(self, root, container=None, standalone=True):
        """root: the Tk instance (used for dialogs / Toplevel parenting).
        container: the frame this tool's UI should be built into. Defaults
        to root itself, which reproduces the original standalone behavior.
        standalone: when False (i.e. embedded in the multi-tool shell),
        window-level chrome (title, size, background) is left alone since
        the shell owns it -- same convention as korber_tool.KorberApp /
        reconciliation_tool.ReconciliationApp."""
        self.root = root
        target = container if container is not None else root

        if standalone:
            root.title("Outlook Email Sender")
            root.geometry("780x760")
            root.configure(bg=BG_COLOR)
            root.minsize(700, 640)
            for icon_name in ("favicon.ico", "icon.ico"):
                icon_path = os.path.join(BASE_DIR, icon_name)
                if not os.path.exists(icon_path) and getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
                    icon_path = os.path.join(sys._MEIPASS, icon_name)
                if os.path.exists(icon_path):
                    try:
                        root.iconbitmap(default=icon_path)
                        break
                    except Exception:
                        try:
                            root.iconbitmap(icon_path)
                            break
                        except Exception:
                            pass

        self.config_store = ConfigStore(CONFIG_JSON_PATH)
        self.outlook = OutlookClient()
        self.template_store = TemplateStore(TEMPLATES_XLSX_PATH)
        self.sent_log = SentLogStore(self.config_store)
        self.attachment_paths = []

        self._subject_before_type = None
        self._last_prefix = None

        self._build_ui(target)
        self._refresh_counts_label()

        if not WIN32COM_AVAILABLE:
            messagebox.showwarning(
                "pywin32 not found",
                "pywin32 is not installed, so this app cannot talk to "
                "Outlook yet.\n\nInstall it with:\n\npip install pywin32")
        if not OPENPYXL_AVAILABLE:
            messagebox.showwarning(
                "openpyxl not found",
                "openpyxl is not installed, so recipient templates "
                "can't be saved/loaded yet.\n\n"
                "Install it with:\n\npip install openpyxl")

    # -- UI construction ---------------------------------------------------

    def _build_ui(self, target):
        header = tk.Frame(target, bg=PANEL_COLOR, height=64)
        header.pack(fill="x", side="top")
        header.pack_propagate(False)
        tk.Label(header, text="Outlook Email Sender", font=FONT_HEADER,
                  bg=PANEL_COLOR, fg=TEXT_COLOR).pack(side="left", padx=20)

        settings_btn = tk.Button(header, text="⚙ Settings",
                                  command=self._open_settings_dialog)
        style_button(settings_btn, "secondary")
        settings_btn.pack(side="right", padx=(0, 20))

        view_records_btn = tk.Button(header, text="View Records",
                                      command=self._open_records_window)
        style_button(view_records_btn, "secondary")
        view_records_btn.pack(side="right", padx=(0, 10))

        self.counts_var = tk.StringVar(value="GDN sent: 0    GRN sent: 0")
        tk.Label(header, textvariable=self.counts_var, font=FONT_NORMAL,
                  bg=PANEL_COLOR, fg=MUTED_COLOR).pack(side="right", padx=(20, 10))

        body = tk.Frame(target, bg=BG_COLOR)
        body.pack(fill="both", expand=True, padx=20, pady=16)
        body.grid_columnconfigure(1, weight=1)

        # Template picker
        tk.Label(body, text="Template:", font=FONT_NORMAL, bg=BG_COLOR,
                  fg=TEXT_COLOR).grid(row=0, column=0, sticky="w", pady=6)
        template_frame = tk.Frame(body, bg=BG_COLOR)
        template_frame.grid(row=0, column=1, sticky="ew", padx=(10, 0), pady=6)
        template_frame.grid_columnconfigure(0, weight=1)

        self.template_var = tk.StringVar(value=TEMPLATE_PLACEHOLDER)
        self.template_combo = ttk.Combobox(
            template_frame, textvariable=self.template_var,
            state="readonly")
        self.template_combo.grid(row=0, column=0, sticky="ew")
        self.template_combo.bind("<<ComboboxSelected>>", self._on_template_selected)

        add_template_btn = tk.Button(template_frame, text="Add New Template",
                                      command=self._open_add_template_dialog)
        style_button(add_template_btn, "secondary")
        add_template_btn.grid(row=0, column=1, padx=(10, 0))

        self._refresh_template_dropdown()

        # To
        tk.Label(body, text="To:", font=FONT_NORMAL, bg=BG_COLOR,
                  fg=TEXT_COLOR).grid(row=1, column=0, sticky="w", pady=6)
        self.to_entry = tk.Entry(body, font=FONT_NORMAL)
        self.to_entry.grid(row=1, column=1, sticky="ew", padx=(10, 0), pady=6)

        # CC
        tk.Label(body, text="CC:", font=FONT_NORMAL, bg=BG_COLOR,
                  fg=TEXT_COLOR).grid(row=2, column=0, sticky="w", pady=6)
        self.cc_entry = tk.Entry(body, font=FONT_NORMAL)
        self.cc_entry.grid(row=2, column=1, sticky="ew", padx=(10, 0), pady=6)

        tk.Label(body, text="(multiple recipients: separate with , or ;)",
                  font=(FONT_FAMILY, 8), bg=BG_COLOR, fg=MUTED_COLOR).grid(
            row=3, column=1, sticky="w", padx=(10, 0))

        # Subject
        tk.Label(body, text="Subject:", font=FONT_NORMAL, bg=BG_COLOR,
                  fg=TEXT_COLOR).grid(row=4, column=0, sticky="w", pady=(12, 6))
        self.subject_entry = tk.Entry(body, font=FONT_NORMAL)
        self.subject_entry.grid(row=4, column=1, sticky="ew", padx=(10, 0), pady=(12, 6))

        # GDN / GRN type selection
        tk.Label(body, text="Type:", font=FONT_NORMAL, bg=BG_COLOR,
                  fg=TEXT_COLOR).grid(row=5, column=0, sticky="w", pady=(0, 6))
        type_frame = tk.Frame(body, bg=BG_COLOR)
        type_frame.grid(row=5, column=1, sticky="w", padx=(10, 0), pady=(0, 6))

        self.selected_type_var = tk.StringVar(value="")

        self.gdn_btn = tk.Button(type_frame, text="GDN",
                                  command=lambda: self._select_type(SEND_TYPE_GDN))
        style_button(self.gdn_btn, "secondary")
        self.gdn_btn.pack(side="left")

        self.grn_btn = tk.Button(type_frame, text="GRN",
                                  command=lambda: self._select_type(SEND_TYPE_GRN))
        style_button(self.grn_btn, "secondary")
        self.grn_btn.pack(side="left", padx=(6, 0))

        clear_type_btn = tk.Button(type_frame, text="Clear",
                                    command=lambda: self._select_type(None))
        style_button(clear_type_btn, "secondary")
        clear_type_btn.pack(side="left", padx=(6, 0))

        tk.Label(type_frame,
                  text="(sets the subject to 'GDN -' / 'GRN -' and marks "
                       "this email for the send log)",
                  font=(FONT_FAMILY, 8), bg=BG_COLOR, fg=MUTED_COLOR).pack(
            side="left", padx=(10, 0))

        # Message
        tk.Label(body, text="Message:", font=FONT_NORMAL, bg=BG_COLOR,
                  fg=TEXT_COLOR).grid(row=6, column=0, sticky="nw", pady=(12, 6))
        msg_frame = tk.Frame(body, bg=BG_COLOR)
        msg_frame.grid(row=6, column=1, sticky="nsew", padx=(10, 0), pady=(12, 6))
        body.grid_rowconfigure(6, weight=1)

        self.message_text = tk.Text(msg_frame, font=FONT_NORMAL, wrap="word",
                                     height=12, relief="solid", bd=1)
        self.message_text.insert("1.0", DEFAULT_BODY_TEXT)
        self.message_text.pack(side="left", fill="both", expand=True)
        msg_scroll = tk.Scrollbar(msg_frame, command=self.message_text.yview)
        msg_scroll.pack(side="right", fill="y")
        self.message_text.configure(yscrollcommand=msg_scroll.set)

        # Signature toggle
        sig_frame = tk.Frame(body, bg=BG_COLOR)
        sig_frame.grid(row=7, column=1, sticky="w", padx=(10, 0), pady=(6, 0))
        self.include_signature_var = tk.BooleanVar(value=True)
        sig_check = tk.Checkbutton(
            sig_frame, text="Include my default Outlook signature",
            variable=self.include_signature_var, font=FONT_NORMAL, bg=BG_COLOR,
            fg=TEXT_COLOR, activebackground=BG_COLOR, selectcolor=PANEL_COLOR)
        sig_check.pack(side="left")
        tk.Label(sig_frame,
                  text="(read straight from your Signatures folder - no paste "
                       "needed, keeps its original fonts/artwork)",
                  font=(FONT_FAMILY, 8), bg=BG_COLOR, fg=MUTED_COLOR).pack(
            side="left", padx=(6, 0))

        # Attachments
        tk.Label(body, text="Attachments:", font=FONT_NORMAL, bg=BG_COLOR,
                  fg=TEXT_COLOR).grid(row=8, column=0, sticky="nw", pady=(12, 6))
        attach_frame = tk.Frame(body, bg=BG_COLOR)
        attach_frame.grid(row=8, column=1, sticky="ew", padx=(10, 0), pady=(12, 6))
        attach_frame.grid_columnconfigure(0, weight=1)

        list_frame = tk.Frame(attach_frame, bg=BG_COLOR)
        list_frame.grid(row=0, column=0, sticky="ew")
        self.attachment_listbox = tk.Listbox(
            list_frame, font=FONT_NORMAL, height=4, relief="solid", bd=1,
            selectbackground=ACCENT_COLOR)
        self.attachment_listbox.pack(side="left", fill="both", expand=True)
        attach_scroll = tk.Scrollbar(list_frame, command=self.attachment_listbox.yview)
        attach_scroll.pack(side="right", fill="y")
        self.attachment_listbox.configure(yscrollcommand=attach_scroll.set)

        attach_btn_frame = tk.Frame(attach_frame, bg=BG_COLOR)
        attach_btn_frame.grid(row=1, column=0, sticky="w", pady=(8, 0))

        attach_btn = tk.Button(attach_btn_frame, text="Attach PDF(s) / Excel(s)...",
                                command=self._attach_files)
        style_button(attach_btn, "secondary")
        attach_btn.pack(side="left")

        remove_btn = tk.Button(attach_btn_frame, text="Remove Selected",
                                command=self._remove_selected_attachment)
        style_button(remove_btn, "danger")
        remove_btn.pack(side="left", padx=(8, 0))

        clear_btn = tk.Button(attach_btn_frame, text="Clear All",
                               command=self._clear_attachments)
        style_button(clear_btn, "secondary")
        clear_btn.pack(side="left", padx=(8, 0))

        # Actions
        action_frame = tk.Frame(target, bg=BG_COLOR)
        action_frame.pack(fill="x", padx=20, pady=(0, 20))

        preview_btn = tk.Button(action_frame, text="Preview in Outlook",
                                 command=self._preview_email)
        style_button(preview_btn, "secondary")
        preview_btn.pack(side="left")

        send_btn = tk.Button(action_frame, text="Send", command=self._send_email)
        style_button(send_btn, "primary")
        send_btn.pack(side="left", padx=10)

    # -- Settings handling ----------------------------------------------------

    def _open_settings_dialog(self):
        SettingsDialog(self.root, self.config_store, on_save=self._refresh_counts_label)

    # -- Template handling ----------------------------------------------------

    def _refresh_template_dropdown(self, select_name=None):
        names = self.template_store.get_names()
        self.template_combo.configure(values=[TEMPLATE_PLACEHOLDER] + names)
        if select_name and select_name in names:
            self.template_var.set(select_name)
            self._on_template_selected()
        else:
            self.template_var.set(TEMPLATE_PLACEHOLDER)

    def _on_template_selected(self, event=None):
        name = self.template_var.get()
        if not name or name == TEMPLATE_PLACEHOLDER:
            return
        template = self.template_store.get_template(name)
        if not template:
            return
        self.to_entry.delete(0, tk.END)
        self.to_entry.insert(0, template["to"])
        self.cc_entry.delete(0, tk.END)
        self.cc_entry.insert(0, template["cc"])

        current_type = self.selected_type_var.get()
        if current_type:
            self._apply_subject_prefix(current_type)

    def _open_add_template_dialog(self):
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror(
                "openpyxl not installed",
                "Install openpyxl to save templates:\n\npip install openpyxl")
            return

        def submit(name, to_addr, cc_addr):
            self.template_store.add_or_update_template(name, to_addr, cc_addr)
            self._refresh_template_dropdown(select_name=name)

        AddTemplateDialog(self.root, on_submit=submit)

    # -- GDN/GRN type selection -------------------------------------------------

    def _build_subject_prefix(self, type_name):
        template_name = self.template_var.get()
        if template_name and template_name != TEMPLATE_PLACEHOLDER:
            return f"{template_name} {type_name} - "
        return f"{type_name} - "

    def _apply_subject_prefix(self, type_name):
        current_text = self.subject_entry.get()
        if self._last_prefix and current_text.startswith(self._last_prefix):
            number_part = current_text[len(self._last_prefix):]
        else:
            number_part = ""
        prefix = self._build_subject_prefix(type_name)
        self.subject_entry.delete(0, tk.END)
        self.subject_entry.insert(0, prefix + number_part)
        self._last_prefix = prefix

    def _select_type(self, type_name):
        turning_off = type_name is not None and self.selected_type_var.get() == type_name
        new_type = "" if turning_off else (type_name or "")
        self.selected_type_var.set(new_type)

        if new_type:
            if self._subject_before_type is None:
                self._subject_before_type = self.subject_entry.get()
            self._apply_subject_prefix(new_type)
        else:
            if self._subject_before_type is not None:
                self.subject_entry.delete(0, tk.END)
                self.subject_entry.insert(0, self._subject_before_type)
            self._subject_before_type = None
            self._last_prefix = None

        style_button(self.gdn_btn,
                     "primary" if self.selected_type_var.get() == SEND_TYPE_GDN
                     else "secondary")
        style_button(self.grn_btn,
                     "primary" if self.selected_type_var.get() == SEND_TYPE_GRN
                     else "secondary")

    # -- Send-tracking (GDN/GRN log) -------------------------------------------

    def _refresh_counts_label(self):
        def _fetch():
            counts = self.sent_log.get_counts()
            text = f"GDN sent: {counts.get('GDN', 0)}    GRN sent: {counts.get('GRN', 0)}"
            try:
                self.root.after(0, lambda: self.counts_var.set(text))
            except Exception:
                pass

        threading.Thread(target=_fetch, daemon=True).start()

    def open_records_window(self):
        """Public wrapper around _open_records_window(), for the shell
        (main_app.py) to call from the Settings page without reaching
        into a private method -- same convention as KorberApp's
        terminate_session() / show_about() and ReconciliationApp's
        open_settings_dialog()."""
        self._open_records_window()

    def _open_records_window(self):
        """Opens the Google Sheet directly in the user's default web browser."""
        sheet_url = self.config_store.get_sheet_url()
        if not sheet_url:
            messagebox.showwarning(
                "Missing URL",
                "Please set your Google Sheet URL in ⚙ Settings first."
            )
            return

        try:
            webbrowser.open(sheet_url)
        except Exception as exc:
            messagebox.showerror(
                "Browser Error",
                f"Could not open the web browser:\n\n{exc}"
            )

    # -- Attachment handling -------------------------------------------------

    def _attach_files(self):
        paths = filedialog.askopenfilenames(
            title="Select PDF and/or Excel files to attach",
            filetypes=ATTACHMENT_FILETYPES,
        )
        if not paths:
            return
        for path in paths:
            if path not in self.attachment_paths:
                self.attachment_paths.append(path)
        self._refresh_attachment_list()

    def _remove_selected_attachment(self):
        selection = list(self.attachment_listbox.curselection())
        if not selection:
            return
        for index in reversed(selection):
            del self.attachment_paths[index]
        self._refresh_attachment_list()

    def _clear_attachments(self):
        self.attachment_paths = []
        self._refresh_attachment_list()

    def _refresh_attachment_list(self):
        self.attachment_listbox.delete(0, tk.END)
        for path in self.attachment_paths:
            self.attachment_listbox.insert(tk.END, os.path.basename(path))

    # -- Gathering / validating form data ------------------------------------

    def _gather_fields(self):
        to_addr = normalize_recipients(self.to_entry.get().strip())
        cc_addr = normalize_recipients(self.cc_entry.get().strip())
        subject = self.subject_entry.get().strip()
        body = self.message_text.get("1.0", "end-1c")
        return to_addr, cc_addr, subject, body

    def _validate(self, to_addr, subject):
        if not to_addr:
            messagebox.showwarning("Missing recipient",
                                    "Please enter at least one 'To:' recipient.")
            return False
        if not subject:
            if not messagebox.askyesno(
                    "No subject",
                    "The subject line is empty. Send anyway?"):
                return False
        return True

    # -- Actions -------------------------------------------------------------

    def _preview_email(self):
        to_addr, cc_addr, subject, body = self._gather_fields()
        if not self._validate(to_addr, subject):
            return
        try:
            mail = self.outlook.build_mail(
                to_addr, cc_addr, subject, body, self.attachment_paths,
                include_signature=self.include_signature_var.get())
            self.outlook.preview(mail)
        except Exception as exc:
            messagebox.showerror("Outlook error",
                                  f"Could not open the preview in Outlook:\n\n{exc}")

    def _send_email(self):
        to_addr, cc_addr, subject, body = self._gather_fields()
        if not self._validate(to_addr, subject):
            return

        if self.attachment_paths:
            attachment_summary = "\n".join(
                f"  - {os.path.basename(p)}" for p in self.attachment_paths)
        else:
            attachment_summary = "  (none)"
        send_type = self.selected_type_var.get() or None
        confirm_text = (
            f"To: {to_addr}\n"
            f"CC: {cc_addr if cc_addr else '(none)'}\n"
            f"Subject: {subject if subject else '(none)'}\n"
            f"Type: {send_type if send_type else '(none)'}\n"
            f"Attachments:\n{attachment_summary}\n\n"
            "Send this email now?"
        )
        if not messagebox.askyesno("Confirm send", confirm_text):
            return

        try:
            mail = self.outlook.build_mail(
                to_addr, cc_addr, subject, body, self.attachment_paths,
                include_signature=self.include_signature_var.get())
            self.outlook.send(mail)
        except Exception as exc:
            messagebox.showerror("Outlook error",
                                  f"Could not send the email:\n\n{exc}")
            return

        if send_type:
            self.sent_log.log_send(send_type, subject)
            self._refresh_counts_label()

        messagebox.showinfo("Sent", "Email sent successfully.")
        self._reset_after_send()

    def _reset_after_send(self):
        self.to_entry.delete(0, tk.END)
        self.cc_entry.delete(0, tk.END)
        self.subject_entry.delete(0, tk.END)
        self.template_var.set(TEMPLATE_PLACEHOLDER)
        self._clear_attachments()
        self._subject_before_type = None
        self._select_type(None)


# --------------------------------------------------------------------------

if __name__ == "__main__":
    try:
        import ctypes
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID('efl.nexus.outlook')
    except Exception:
        pass

    root = tk.Tk()
    app = OutlookEmailApp(root)
    root.mainloop()
